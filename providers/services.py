# ============================================================
# 📂 providers/services.py
# 🧠 Primey Care | Providers Services
# ------------------------------------------------------------
# ✅ خدمات مساعدة رسمية لموديول الجهات المقدمة للخدمة
# ✅ Parsing / Validation / Serialization
# ✅ Filtering / Pagination
# ✅ Create / Update helpers
# ✅ Excel Import / Dry Run / Upsert
# ✅ دعم الاسم العربي والإنجليزي بشكل مستقل
# ✅ دعم السجل التجاري والرقم الضريبي
# ✅ دعم شعار وصورة مقدم الخدمة عبر Google Drive
# ✅ دعم ملفات ومرفقات مقدم الخدمة ProviderDocument
# ✅ Customer medical network fields from active marketing contracts
# ------------------------------------------------------------
# ملاحظات مهمة:
# - name يبقى موجودًا للتوافق الخلفي مع الاستيراد والواجهات القديمة.
# - name_ar و name_en تستخدم للفصل الواضح بين الاسم العربي والإنجليزي.
# - code يبقى unique لأنه كود داخلي للنظام.
# - import_key هو مفتاح منع التكرار عند الاستيراد من Excel.
# - الاستيراد يدعم .xlsx رسميًا.
# - الاستيراد يدعم كشف صف العناوين تلقائيًا داخل أول 20 صف.
# ============================================================

from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, QuerySet
from django.utils import timezone

from providers.models import (
    Provider,
    ProviderDocument,
    ProviderDocumentType,
    ProviderStatus,
    ProviderType,
)


# ============================================================
# 🔹 Constants
# ============================================================

MEDICAL_NETWORK_IMPORT_SOURCE = "medical_network_excel"
EXCEL_SCAN_HEADER_ROWS = 20

EXCEL_HEADER_ALIASES = {
    # External reference / serial
    "م": "external_reference",
    "م.": "external_reference",
    "#": "external_reference",
    "id": "external_reference",
    "ID": "external_reference",
    "Id": "external_reference",
    "no": "external_reference",
    "No": "external_reference",
    "NO": "external_reference",
    "number": "external_reference",
    "Number": "external_reference",
    "serial": "external_reference",
    "Serial": "external_reference",
    "reference": "external_reference",
    "Reference": "external_reference",
    "الرقم": "external_reference",
    "رقم": "external_reference",
    "رقم السجل": "external_reference",
    "رقم المرجع": "external_reference",
    "المرجع": "external_reference",
    "كود": "external_reference",
    "الكود": "external_reference",

    # Region
    "المنطقة": "region",
    "المنطقه": "region",
    "المنطقة الإدارية": "region",
    "المنطقه الاداريه": "region",
    "region": "region",
    "Region": "region",
    "REGION": "region",
    "region name": "region",
    "Region Name": "region",

    # City
    "المدينة": "city",
    "المدينه": "city",
    "المحافظة": "city",
    "المحافظه": "city",
    "city": "city",
    "City": "city",
    "CITY": "city",
    "city name": "city",
    "City Name": "city",
    "province": "city",
    "Province": "city",

    # Provider name
    "اسم المركز": "name",
    "اسم الجهة": "name",
    "اسم الجهه": "name",
    "اسم مقدم الخدمة": "name",
    "اسم مقدم الخدمه": "name",
    "اسم المنشأة": "name",
    "اسم المنشأه": "name",
    "اسم المؤسسة": "name",
    "اسم المؤسسه": "name",
    "اسم المستشفى": "name",
    "اسم العيادة": "name",
    "اسم العياده": "name",
    "اسم المختبر": "name",
    "اسم الصيدلية": "name",
    "اسم الصيدليه": "name",
    "المركز": "name",
    "الجهة": "name",
    "الجهه": "name",
    "مقدم الخدمة": "name",
    "مقدم الخدمه": "name",
    "المنشأة": "name",
    "المنشأه": "name",
    "provider": "name",
    "Provider": "name",
    "provider name": "name",
    "Provider Name": "name",
    "service provider": "name",
    "Service Provider": "name",
    "serviceproviders": "name",
    "ServiceProviders": "name",
    "service providers": "name",
    "Service Providers": "name",
    "facility": "name",
    "Facility": "name",
    "facility name": "name",
    "Facility Name": "name",
    "center": "name",
    "Center": "name",
    "center name": "name",
    "Center Name": "name",
    "centre": "name",
    "Centre": "name",
    "centre name": "name",
    "Centre Name": "name",
    "hospital": "name",
    "Hospital": "name",
    "hospital name": "name",
    "Hospital Name": "name",
    "clinic": "name",
    "Clinic": "name",
    "clinic name": "name",
    "Clinic Name": "name",
    "name": "name",
    "Name": "name",
    "NAME": "name",

    # Arabic / English split names
    "الاسم العربي": "name_ar",
    "اسم الجهة بالعربي": "name_ar",
    "اسم الجهه بالعربي": "name_ar",
    "اسم مقدم الخدمة بالعربي": "name_ar",
    "اسم مقدم الخدمه بالعربي": "name_ar",
    "arabic name": "name_ar",
    "Arabic Name": "name_ar",
    "name ar": "name_ar",
    "Name AR": "name_ar",
    "name_ar": "name_ar",

    "الاسم الإنجليزي": "name_en",
    "الاسم الانجليزي": "name_en",
    "اسم الجهة بالإنجليزي": "name_en",
    "اسم الجهة بالانجليزي": "name_en",
    "اسم الجهه بالانجليزي": "name_en",
    "اسم مقدم الخدمة بالإنجليزي": "name_en",
    "اسم مقدم الخدمة بالانجليزي": "name_en",
    "english name": "name_en",
    "English Name": "name_en",
    "name en": "name_en",
    "Name EN": "name_en",
    "name_en": "name_en",

    # Legal / tax
    "السجل التجاري": "commercial_registration",
    "رقم السجل التجاري": "commercial_registration",
    "سجل تجاري": "commercial_registration",
    "commercial registration": "commercial_registration",
    "Commercial Registration": "commercial_registration",
    "cr": "commercial_registration",
    "CR": "commercial_registration",
    "cr number": "commercial_registration",
    "CR Number": "commercial_registration",

    "الرقم الضريبي": "tax_number",
    "رقم ضريبي": "tax_number",
    "الرقم الضريبي / رقم ضريبة القيمة المضافة": "tax_number",
    "vat": "tax_number",
    "VAT": "tax_number",
    "vat number": "tax_number",
    "VAT Number": "tax_number",
    "tax number": "tax_number",
    "Tax Number": "tax_number",

    # Category / classification
    "التصنيف": "source_category",
    "الفئة": "source_category",
    "الفئه": "source_category",
    "النوع": "source_category",
    "نوع الجهة": "source_category",
    "نوع الجهه": "source_category",
    "نوع مقدم الخدمة": "source_category",
    "نوع مقدم الخدمه": "source_category",
    "نوع المنشأة": "source_category",
    "نوع المنشأه": "source_category",
    "التخصص": "source_category",
    "النشاط": "source_category",
    "category": "source_category",
    "Category": "source_category",
    "type": "source_category",
    "Type": "source_category",
    "classification": "source_category",
    "Classification": "source_category",
    "specialty": "source_category",
    "Specialty": "source_category",
    "speciality": "source_category",
    "Speciality": "source_category",
    "activity": "source_category",
    "Activity": "source_category",

    # Address
    "العنوان الكامل": "address",
    "العنوان": "address",
    "الموقع": "address",
    "اللوكيشن": "address",
    "address": "address",
    "Address": "address",
    "full address": "address",
    "Full Address": "address",
    "location": "address",
    "Location": "address",

    # Area / district
    "الحي": "area",
    "الحى": "area",
    "المنطقة داخل المدينة": "area",
    "المنطقه داخل المدينه": "area",
    "area": "area",
    "Area": "area",
    "district": "area",
    "District": "area",
    "neighborhood": "area",
    "Neighborhood": "area",
    "neighbourhood": "area",
    "Neighbourhood": "area",

    # Street
    "الشارع": "street",
    "street": "street",
    "Street": "street",
    "road": "street",
    "Road": "street",

    # Phone
    "الهاتف": "phone",
    "هاتف": "phone",
    "رقم الهاتف": "phone",
    "تليفون": "phone",
    "تلفون": "phone",
    "phone": "phone",
    "Phone": "phone",
    "PHONE": "phone",
    "telephone": "phone",
    "Telephone": "phone",
    "tel": "phone",
    "Tel": "phone",

    # Mobile
    "الجوال": "mobile",
    "جوال": "mobile",
    "رقم الجوال": "mobile",
    "الموبايل": "mobile",
    "موبايل": "mobile",
    "mobile": "mobile",
    "Mobile": "mobile",
    "MOBILE": "mobile",
    "cell": "mobile",
    "Cell": "mobile",

    # Notes
    "الخصومات / الملاحظات": "notes",
    "الخصومات": "notes",
    "الملاحظات": "notes",
    "ملاحظات": "notes",
    "ملاحظة": "notes",
    "خصم": "notes",
    "discount": "notes",
    "Discount": "notes",
    "discounts": "notes",
    "Discounts": "notes",
    "notes": "notes",
    "Notes": "notes",
    "NOTE": "notes",
    "remarks": "notes",
    "Remarks": "notes",
}


# ============================================================
# 🔹 JSON Helpers
# ============================================================

def parse_json_body(request) -> dict[str, Any]:
    if not request.body:
        return {}

    try:
        return json.loads(request.body.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise ValidationError(f"Invalid JSON body: {exc}")


# ============================================================
# 🔹 Primitive Parsers
# ============================================================

def normalize_text(value: Any, default: str = "") -> str:
    if value is None:
        return default

    value_str = str(value).strip()

    if value_str.lower() in {"none", "nan", "null"}:
        return default

    return value_str


def normalize_spaces(value: Any, default: str = "") -> str:
    value_str = normalize_text(value, default)
    value_str = re.sub(r"\s+", " ", value_str)
    return value_str.strip()


def normalize_identifier(value: Any, max_length: int = 100) -> str:
    value_str = normalize_text(value)
    value_str = value_str.replace(" ", "")
    value_str = re.sub(r"[^0-9A-Za-z\u0600-\u06FF\-_/]", "", value_str)
    return value_str[:max_length]


def normalize_url(value: Any) -> str:
    value_str = normalize_text(value)
    return value_str[:1000]


def normalize_phone(value: Any) -> str:
    value_str = normalize_text(value)

    if not value_str:
        return ""

    # Excel may read numbers as floats like 550000000.0
    if re.fullmatch(r"\d+\.0", value_str):
        value_str = value_str[:-2]

    value_str = value_str.replace(" ", "")
    value_str = value_str.replace("-", "")
    value_str = value_str.replace("(", "")
    value_str = value_str.replace(")", "")

    # Keep plus for international numbers.
    value_str = re.sub(r"[^0-9+]", "", value_str)

    return value_str[:30]


def parse_bool(value: Any, default: bool | None = None) -> bool | None:
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    value_str = str(value).strip().lower()

    if value_str in {"1", "true", "yes", "on", "نعم", "صح"}:
        return True

    if value_str in {"0", "false", "no", "off", "لا", "خطأ"}:
        return False

    return default


def parse_int(value: Any, default: int | None = None) -> int | None:
    if value in (None, ""):
        return default

    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _money(value: Any, default: Decimal | None = None) -> Decimal:
    if default is None:
        default = Decimal("0.00")

    if value in (None, ""):
        return default.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return default.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _percent(value: Any, default: Decimal | None = None) -> Decimal:
    if default is None:
        default = Decimal("0.00")

    if value in (None, ""):
        return default.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    try:
        parsed = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return default.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    if parsed < Decimal("0.00"):
        return Decimal("0.00")

    if parsed > Decimal("100.00"):
        return Decimal("100.00")

    return parsed


def _decimal_to_str(value: Any) -> str:
    return str(_money(value))


def _percent_to_str(value: Any) -> str:
    return str(_percent(value))


def _iso_date(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return str(value)


# ============================================================
# 🔹 Contract / Medical Network Helpers
# ============================================================

def _active_contract_filter(prefix: str = "contracts") -> Q:
    today = timezone.localdate()

    return (
        Q(**{f"{prefix}__status": "ACTIVE"})
        & (
            Q(**{f"{prefix}__start_date__isnull": True})
            | Q(**{f"{prefix}__start_date__lte": today})
        )
        & (
            Q(**{f"{prefix}__end_date__isnull": True})
            | Q(**{f"{prefix}__end_date__gte": today})
        )
    )


def _active_contract_product_filter() -> Q:
    today = timezone.localdate()

    return (
        Q(contracts__status="ACTIVE")
        & Q(contracts__contract_products__is_active=True)
        & (
            Q(contracts__start_date__isnull=True)
            | Q(contracts__start_date__lte=today)
        )
        & (
            Q(contracts__end_date__isnull=True)
            | Q(contracts__end_date__gte=today)
        )
    )


def _calculate_discount_from_special_price(
    *,
    before_price: Decimal,
    after_price: Decimal,
) -> Decimal:
    before_price = _money(before_price)
    after_price = _money(after_price)

    if before_price <= Decimal("0.00"):
        return Decimal("0.00")

    if after_price >= before_price:
        return Decimal("0.00")

    return _percent(((before_price - after_price) * Decimal("100.00")) / before_price)


def _get_provider_contract_queryset(provider: Provider):
    try:
        from contracts.models import Contract
    except Exception:
        return None

    today = timezone.localdate()

    return (
        Contract.objects
        .filter(provider_id=provider.pk, status="ACTIVE")
        .filter(
            Q(start_date__isnull=True) | Q(start_date__lte=today),
            Q(end_date__isnull=True) | Q(end_date__gte=today),
        )
        .prefetch_related("contract_products", "contract_products__product")
        .order_by("-created_at", "-id")
    )


def _serialize_active_contract(contract: Any) -> dict[str, Any]:
    return {
        "id": getattr(contract, "id", None),
        "contract_number": getattr(contract, "contract_number", ""),
        "title": getattr(contract, "title", ""),
        "status": getattr(contract, "status", ""),
        "start_date": _iso_date(getattr(contract, "start_date", None)),
        "end_date": _iso_date(getattr(contract, "end_date", None)),
        "signed_at": _iso_date(getattr(contract, "signed_at", None)),
        "pricing_model": getattr(contract, "pricing_model", ""),
        "discount_percentage": _percent_to_str(
            getattr(contract, "discount_percentage", Decimal("0.00"))
        ),
        "system_commission_percentage": _percent_to_str(
            getattr(contract, "system_commission_percentage", Decimal("0.00"))
        ),
    }


def _calculate_provider_network_context(provider: Provider) -> dict[str, Any]:
    """
    يحسب بيانات الشبكة الطبية للعميل من العقود النشطة.

    الحقول الناتجة تستخدم في:
    - /customer/network
    - /customer/page
    - أي عرض عام لمقدمي الخدمة المتعاقدين
    """

    default_context = {
        "has_active_contract": False,
        "has_active_contracts": False,
        "active_contracts_count": 0,
        "contracted_products_count": 0,
        "highest_discount_percent": Decimal("0.00"),
        "max_discount_percent": Decimal("0.00"),
        "discount_percent": Decimal("0.00"),
        "active_contract": None,
        "active_contracts": [],
    }

    if not provider or not getattr(provider, "pk", None):
        return default_context

    queryset = _get_provider_contract_queryset(provider)

    if queryset is None:
        return default_context

    contracts = list(queryset)

    if not contracts:
        return default_context

    active_contracts: list[dict[str, Any]] = []
    product_ids: set[int] = set()
    highest_discount = Decimal("0.00")

    for contract in contracts:
        serialized_contract = _serialize_active_contract(contract)
        contract_level_discount = _percent(
            getattr(contract, "discount_percentage", Decimal("0.00"))
        )
        highest_discount = max(highest_discount, contract_level_discount)

        active_contract_products = []

        try:
            contract_products = contract.contract_products.all()
        except Exception:
            contract_products = []

        for contract_product in contract_products:
            if not getattr(contract_product, "is_active", False):
                continue

            product = getattr(contract_product, "product", None)
            product_id = getattr(contract_product, "product_id", None)

            if product_id:
                product_ids.add(int(product_id))

            product_price = _money(getattr(product, "price", Decimal("0.00"))) if product else Decimal("0.00")
            product_discount = _percent(
                getattr(contract_product, "discount_percentage", Decimal("0.00"))
            )
            special_price = getattr(contract_product, "special_price", None)

            if special_price is not None:
                special_price_discount = _calculate_discount_from_special_price(
                    before_price=product_price,
                    after_price=_money(special_price),
                )
                product_discount = max(product_discount, special_price_discount)

            effective_discount = max(product_discount, contract_level_discount)
            highest_discount = max(highest_discount, effective_discount)

            active_contract_products.append(
                {
                    "id": getattr(contract_product, "id", None),
                    "product_id": product_id,
                    "product_name": getattr(product, "name", "") if product else "",
                    "product_code": getattr(product, "code", "") if product else "",
                    "product_type": getattr(product, "product_type", "") if product else "",
                    "is_active": getattr(contract_product, "is_active", False),
                    "special_price": (
                        _decimal_to_str(special_price)
                        if special_price is not None
                        else None
                    ),
                    "discount_percentage": _percent_to_str(
                        getattr(contract_product, "discount_percentage", Decimal("0.00"))
                    ),
                    "effective_discount_percent": _percent_to_str(effective_discount),
                    "coverage_notes": getattr(contract_product, "coverage_notes", ""),
                }
            )

        serialized_contract["contracted_products_count"] = len(active_contract_products)
        serialized_contract["contract_products"] = active_contract_products[:10]
        active_contracts.append(serialized_contract)

    has_active = bool(active_contracts)

    return {
        "has_active_contract": has_active,
        "has_active_contracts": has_active,
        "active_contracts_count": len(active_contracts),
        "contracted_products_count": len(product_ids),
        "highest_discount_percent": highest_discount,
        "max_discount_percent": highest_discount,
        "discount_percent": highest_discount,
        "active_contract": active_contracts[0] if active_contracts else None,
        "active_contracts": active_contracts[:5],
    }


# ============================================================
# 🔹 Query Helpers
# ============================================================

def paginate_queryset(queryset: QuerySet, page: int = 1, page_size: int = 20) -> dict[str, Any]:
    safe_page = max(page or 1, 1)
    safe_page_size = min(max(page_size or 20, 1), 100)

    paginator = Paginator(queryset, safe_page_size)
    current_page = paginator.get_page(safe_page)

    return {
        "items": list(current_page.object_list),
        "pagination": {
            "page": current_page.number,
            "page_size": safe_page_size,
            "total_pages": paginator.num_pages,
            "total_items": paginator.count,
            "has_next": current_page.has_next(),
            "has_previous": current_page.has_previous(),
        },
    }


def apply_provider_filters(queryset: QuerySet[Provider], params) -> QuerySet[Provider]:
    q = normalize_text(params.get("q") or params.get("search"))
    provider_type = normalize_text(params.get("provider_type"))
    status = normalize_text(params.get("status"))
    region = normalize_text(params.get("region"))
    city = normalize_text(params.get("city"))
    area = normalize_text(params.get("area"))
    source_category = normalize_text(params.get("source_category") or params.get("category"))
    import_source = normalize_text(params.get("import_source"))
    is_featured = parse_bool(params.get("is_featured"))
    is_active = parse_bool(params.get("is_active"))

    has_logo = parse_bool(params.get("has_logo"))
    has_image = parse_bool(params.get("has_image"))
    has_drive_folder = parse_bool(params.get("has_drive_folder"))

    has_active_contract = parse_bool(params.get("has_active_contract"))
    has_active_contracts = parse_bool(params.get("has_active_contracts"))
    has_contract = parse_bool(params.get("has_contract"))
    contracted = parse_bool(params.get("contracted"))

    if q:
        queryset = queryset.filter(
            Q(name__icontains=q)
            | Q(name_ar__icontains=q)
            | Q(name_en__icontains=q)
            | Q(code__icontains=q)
            | Q(commercial_registration__icontains=q)
            | Q(tax_number__icontains=q)
            | Q(contact_person__icontains=q)
            | Q(phone__icontains=q)
            | Q(mobile__icontains=q)
            | Q(email__icontains=q)
            | Q(region__icontains=q)
            | Q(city__icontains=q)
            | Q(area__icontains=q)
            | Q(street__icontains=q)
            | Q(address__icontains=q)
            | Q(source_category__icontains=q)
            | Q(import_source__icontains=q)
            | Q(external_reference__icontains=q)
            | Q(drive_folder_id__icontains=q)
            | Q(logo_drive_file_id__icontains=q)
            | Q(image_drive_file_id__icontains=q)
            | Q(notes__icontains=q)
        )

    if is_active is True:
        queryset = queryset.filter(status=ProviderStatus.ACTIVE)

    if is_active is False:
        queryset = queryset.exclude(status=ProviderStatus.ACTIVE)

    if provider_type:
        queryset = queryset.filter(provider_type=provider_type)

    if status:
        queryset = queryset.filter(status=status)

    if region:
        queryset = queryset.filter(region__icontains=region)

    if city:
        queryset = queryset.filter(city__icontains=city)

    if area:
        queryset = queryset.filter(area__icontains=area)

    if source_category:
        queryset = queryset.filter(source_category__icontains=source_category)

    if import_source:
        queryset = queryset.filter(import_source=import_source)

    if is_featured is not None:
        queryset = queryset.filter(is_featured=is_featured)

    if has_logo is True:
        queryset = queryset.filter(Q(logo_url__gt="") | Q(logo_drive_file_id__gt=""))

    if has_logo is False:
        queryset = queryset.filter(logo_url="", logo_drive_file_id="")

    if has_image is True:
        queryset = queryset.filter(Q(image_url__gt="") | Q(image_drive_file_id__gt=""))

    if has_image is False:
        queryset = queryset.filter(image_url="", image_drive_file_id="")

    if has_drive_folder is True:
        queryset = queryset.exclude(drive_folder_id="")

    if has_drive_folder is False:
        queryset = queryset.filter(drive_folder_id="")

    contract_filter_value = (
        has_active_contract
        if has_active_contract is not None
        else has_active_contracts
        if has_active_contracts is not None
        else has_contract
        if has_contract is not None
        else contracted
    )

    if contract_filter_value is True:
        queryset = queryset.filter(_active_contract_product_filter())

    if contract_filter_value is False:
        queryset = queryset.exclude(_active_contract_product_filter())

    ordering = normalize_text(params.get("ordering") or params.get("order_by"))

    allowed_ordering = {
        "name": "name",
        "-name": "-name",
        "city": "city",
        "-city": "-city",
        "region": "region",
        "-region": "-region",
        "created_at": "created_at",
        "-created_at": "-created_at",
        "updated_at": "updated_at",
        "-updated_at": "-updated_at",
    }

    if ordering in allowed_ordering:
        queryset = queryset.order_by(allowed_ordering[ordering], "-id")

    return queryset.distinct()


# ============================================================
# 🔹 Validation Helpers
# ============================================================

def _validate_provider_type(value: str) -> str:
    allowed = {choice for choice, _ in ProviderType.choices}
    if value not in allowed:
        raise ValidationError(f"Invalid provider_type. Allowed values: {sorted(allowed)}")
    return value


def _validate_provider_status(value: str) -> str:
    allowed = {choice for choice, _ in ProviderStatus.choices}
    if value not in allowed:
        raise ValidationError(f"Invalid status. Allowed values: {sorted(allowed)}")
    return value


def _validate_provider_document_type(value: str) -> str:
    allowed = {choice for choice, _ in ProviderDocumentType.choices}
    if value not in allowed:
        raise ValidationError(f"Invalid file_type. Allowed values: {sorted(allowed)}")
    return value


def _validate_unique_code(code: str, *, exclude_id: int | None = None) -> None:
    queryset = Provider.objects.filter(code__iexact=code)
    if exclude_id:
        queryset = queryset.exclude(pk=exclude_id)

    if queryset.exists():
        raise ValidationError("A provider with this code already exists.")


def _validate_unique_import_key(import_key: str, *, exclude_id: int | None = None) -> None:
    if not import_key:
        return

    queryset = Provider.objects.filter(import_key=import_key)
    if exclude_id:
        queryset = queryset.exclude(pk=exclude_id)

    if queryset.exists():
        raise ValidationError("A provider with this import key already exists.")


# ============================================================
# 🔹 Provider Type Mapping
# ============================================================

def guess_provider_type(source_category: str, name: str = "") -> str:
    value = f"{source_category} {name}".strip().lower()

    if any(word in value for word in ["مستشفى", "مستشفي", "hospital"]):
        return ProviderType.HOSPITAL

    if any(word in value for word in ["صيدلية", "صيدليه", "pharmacy"]):
        return ProviderType.PHARMACY

    if any(word in value for word in ["مختبر", "معمل", "lab", "laboratory"]):
        return ProviderType.LAB

    if any(word in value for word in ["عيادة", "عياده", "clinic"]):
        return ProviderType.CLINIC

    if any(word in value for word in ["مركز", "مجمع", "طبي", "medical center", "polyclinic"]):
        return ProviderType.MEDICAL_CENTER

    return ProviderType.OTHER


# ============================================================
# 🔹 Code / Import Key Helpers
# ============================================================

def build_provider_import_key(
    *,
    name: str,
    region: str = "",
    city: str = "",
    area: str = "",
    address: str = "",
    source_category: str = "",
    import_source: str = MEDICAL_NETWORK_IMPORT_SOURCE,
) -> str:
    raw_key = "|".join(
        [
            normalize_spaces(import_source).lower(),
            normalize_spaces(name).lower(),
            normalize_spaces(region).lower(),
            normalize_spaces(city).lower(),
            normalize_spaces(area).lower(),
            normalize_spaces(address).lower(),
            normalize_spaces(source_category).lower(),
        ]
    )

    digest = hashlib.sha256(raw_key.encode("utf-8")).hexdigest()
    return f"prov_{digest[:48]}"


def build_provider_code_from_import_key(import_key: str) -> str:
    digest = hashlib.sha1(import_key.encode("utf-8")).hexdigest().upper()
    return f"PV-{digest[:12]}"


def generate_unique_provider_code(*, base: str | None = None) -> str:
    if base:
        normalized = normalize_text(base).upper()
        normalized = re.sub(r"[^A-Z0-9-]", "-", normalized)
        normalized = re.sub(r"-+", "-", normalized).strip("-")
        if normalized:
            candidate = normalized[:50]
            if not Provider.objects.filter(code__iexact=candidate).exists():
                return candidate

    for index in range(1, 10000):
        candidate = f"PV-{timezone.now().strftime('%Y%m%d%H%M%S')}-{index}"
        candidate = candidate[:50]
        if not Provider.objects.filter(code__iexact=candidate).exists():
            return candidate

    raise ValidationError("Unable to generate a unique provider code.")


# ============================================================
# 🔹 Display Name Helpers
# ============================================================

def resolve_provider_name(*, payload: dict[str, Any]) -> str:
    name = normalize_spaces(payload.get("name"))
    name_ar = normalize_spaces(payload.get("name_ar"))
    name_en = normalize_spaces(payload.get("name_en"))

    return name or name_ar or name_en


def sync_provider_legacy_name(instance: Provider) -> None:
    """
    Keep legacy name populated to avoid breaking old screens/imports.
    """

    if not normalize_spaces(instance.name):
        instance.name = normalize_spaces(instance.name_ar) or normalize_spaces(instance.name_en)

    if not normalize_spaces(instance.name_ar) and normalize_spaces(instance.name):
        instance.name_ar = normalize_spaces(instance.name)

    if not normalize_spaces(instance.name_en):
        instance.name_en = ""


# ============================================================
# 🔹 Documents Serialization
# ============================================================

def serialize_provider_document(obj: ProviderDocument) -> dict[str, Any]:
    return {
        "id": obj.id,
        "provider_id": obj.provider_id,
        "file_type": obj.file_type,
        "file_type_label": obj.get_file_type_display() if obj.file_type else "",
        "title": obj.title,
        "description": obj.description,
        "file_url": obj.file_url,
        "drive_file_id": obj.drive_file_id,
        "drive_folder_id": obj.drive_folder_id,
        "original_filename": obj.original_filename,
        "content_type": obj.content_type,
        "size_bytes": obj.size_bytes,
        "is_primary": obj.is_primary,
        "uploaded_by_id": obj.uploaded_by_id,
        "uploaded_by": (
            getattr(obj.uploaded_by, "get_full_name", lambda: "")()
            or getattr(obj.uploaded_by, "username", "")
            if obj.uploaded_by_id
            else ""
        ),
        "created_at": obj.created_at.isoformat() if obj.created_at else None,
        "updated_at": obj.updated_at.isoformat() if obj.updated_at else None,
    }


def serialize_provider_documents(provider: Provider) -> list[dict[str, Any]]:
    documents = getattr(provider, "documents", None)

    if documents is None:
        return []

    try:
        queryset = documents.all().select_related("uploaded_by")
    except Exception:
        queryset = ProviderDocument.objects.filter(provider=provider).select_related("uploaded_by")

    return [serialize_provider_document(document) for document in queryset]


# ============================================================
# 🔹 Serialization
# ============================================================

def serialize_provider(obj: Provider, *, include_documents: bool = False) -> dict[str, Any]:
    network_context = _calculate_provider_network_context(obj)

    highest_discount_percent = network_context["highest_discount_percent"]
    has_active_contract = bool(network_context["has_active_contract"])

    data = {
        "id": obj.id,
        "name": obj.name,
        "name_ar": getattr(obj, "name_ar", ""),
        "name_en": getattr(obj, "name_en", ""),
        "display_name_ar": getattr(obj, "display_name_ar", "") or getattr(obj, "name_ar", "") or obj.name,
        "display_name_en": getattr(obj, "display_name_en", "") or getattr(obj, "name_en", "") or obj.name,
        "code": obj.code,
        "provider_type": obj.provider_type,
        "category": getattr(obj, "source_category", "") or obj.provider_type,
        "classification": getattr(obj, "source_category", "") or obj.provider_type,
        "status": obj.status,
        "is_active": obj.status == ProviderStatus.ACTIVE,

        "commercial_registration": getattr(obj, "commercial_registration", ""),
        "tax_number": getattr(obj, "tax_number", ""),

        "logo_url": getattr(obj, "logo_url", ""),
        "logo": getattr(obj, "logo_url", ""),
        "logo_drive_file_id": getattr(obj, "logo_drive_file_id", ""),
        "image_url": getattr(obj, "image_url", ""),
        "image": getattr(obj, "image_url", ""),
        "image_drive_file_id": getattr(obj, "image_drive_file_id", ""),

        "drive_folder_id": getattr(obj, "drive_folder_id", ""),
        "drive_folder_url": getattr(obj, "drive_folder_url", ""),

        "contact_person": obj.contact_person,
        "phone": obj.phone,
        "phone_number": obj.phone,
        "mobile": obj.mobile,
        "whatsapp_number": obj.mobile or obj.phone,
        "email": obj.email,
        "website": obj.website,
        "region": getattr(obj, "region", ""),
        "region_name": getattr(obj, "region", ""),
        "city": obj.city,
        "city_name": obj.city,
        "area": obj.area,
        "street": getattr(obj, "street", ""),
        "address": obj.address,
        "google_maps_link": obj.google_maps_link,
        "source_category": getattr(obj, "source_category", ""),
        "provider_category": getattr(obj, "source_category", ""),
        "import_key": getattr(obj, "import_key", None),
        "import_source": getattr(obj, "import_source", ""),
        "external_reference": getattr(obj, "external_reference", ""),
        "notes": obj.notes,
        "is_featured": obj.is_featured,
        "featured": obj.is_featured,
        "has_logo": bool(getattr(obj, "logo_url", "") or getattr(obj, "logo_drive_file_id", "")),
        "has_image": bool(getattr(obj, "image_url", "") or getattr(obj, "image_drive_file_id", "")),
        "has_drive_folder": bool(getattr(obj, "drive_folder_id", "")),

        # Customer medical network fields
        "has_active_contract": has_active_contract,
        "has_active_contracts": has_active_contract,
        "active_contract": network_context["active_contract"],
        "active_contracts": network_context["active_contracts"],
        "active_contracts_count": network_context["active_contracts_count"],
        "contracted_products_count": network_context["contracted_products_count"],
        "highest_discount_percent": _percent_to_str(highest_discount_percent),
        "max_discount_percent": _percent_to_str(network_context["max_discount_percent"]),
        "discount_percent": _percent_to_str(network_context["discount_percent"]),
        "discount_percentage": _percent_to_str(network_context["discount_percent"]),
        "discount_rate": _percent_to_str(network_context["discount_percent"]),

        "last_imported_at": obj.last_imported_at.isoformat() if getattr(obj, "last_imported_at", None) else None,
        "created_at": obj.created_at.isoformat() if obj.created_at else None,
        "updated_at": obj.updated_at.isoformat() if obj.updated_at else None,
    }

    if include_documents:
        data["documents"] = serialize_provider_documents(obj)

    return data


# ============================================================
# 🔹 Create Provider
# ============================================================

def create_provider(*, payload: dict[str, Any]) -> Provider:
    name = resolve_provider_name(payload=payload)
    name_ar = normalize_spaces(payload.get("name_ar"))
    name_en = normalize_spaces(payload.get("name_en"))
    code = normalize_text(payload.get("code")).upper()
    provider_type = normalize_text(payload.get("provider_type")) or ProviderType.OTHER
    status = normalize_text(payload.get("status")) or ProviderStatus.ACTIVE
    import_key = normalize_text(payload.get("import_key")) or None

    if not name:
        raise ValidationError("Provider name is required.")

    if not code:
        code = generate_unique_provider_code()

    provider_type = _validate_provider_type(provider_type)
    status = _validate_provider_status(status)

    _validate_unique_code(code)
    _validate_unique_import_key(import_key)

    provider = Provider(
        name=name,
        name_ar=name_ar or name,
        name_en=name_en,
        code=code,
        provider_type=provider_type,
        status=status,
        commercial_registration=normalize_identifier(payload.get("commercial_registration")),
        tax_number=normalize_identifier(payload.get("tax_number")),
        logo_url=normalize_url(payload.get("logo_url")),
        logo_drive_file_id=normalize_text(payload.get("logo_drive_file_id"))[:255],
        image_url=normalize_url(payload.get("image_url")),
        image_drive_file_id=normalize_text(payload.get("image_drive_file_id"))[:255],
        drive_folder_id=normalize_text(payload.get("drive_folder_id"))[:255],
        drive_folder_url=normalize_url(payload.get("drive_folder_url")),
        contact_person=normalize_spaces(payload.get("contact_person")),
        phone=normalize_phone(payload.get("phone")),
        mobile=normalize_phone(payload.get("mobile")),
        email=normalize_text(payload.get("email")),
        website=normalize_text(payload.get("website")),
        region=normalize_spaces(payload.get("region")),
        city=normalize_spaces(payload.get("city")),
        area=normalize_spaces(payload.get("area")),
        street=normalize_spaces(payload.get("street")),
        address=normalize_spaces(payload.get("address")),
        google_maps_link=normalize_text(payload.get("google_maps_link")),
        source_category=normalize_spaces(payload.get("source_category")),
        import_key=import_key,
        import_source=normalize_text(payload.get("import_source")),
        external_reference=normalize_text(payload.get("external_reference")),
        notes=normalize_spaces(payload.get("notes")),
        is_featured=parse_bool(payload.get("is_featured"), False) or False,
    )

    sync_provider_legacy_name(provider)

    if import_key:
        provider.last_imported_at = timezone.now()

    provider.full_clean()
    provider.save()
    return provider


# ============================================================
# 🔹 Update Provider
# ============================================================

def update_provider(*, instance: Provider, payload: dict[str, Any]) -> Provider:
    if "name" in payload:
        instance.name = normalize_spaces(payload.get("name"))

    if "name_ar" in payload:
        instance.name_ar = normalize_spaces(payload.get("name_ar"))

    if "name_en" in payload:
        instance.name_en = normalize_spaces(payload.get("name_en"))

    if "code" in payload:
        instance.code = normalize_text(payload.get("code")).upper()

    if "provider_type" in payload:
        instance.provider_type = normalize_text(payload.get("provider_type"))

    if "status" in payload:
        instance.status = normalize_text(payload.get("status"))

    if "commercial_registration" in payload:
        instance.commercial_registration = normalize_identifier(payload.get("commercial_registration"))

    if "tax_number" in payload:
        instance.tax_number = normalize_identifier(payload.get("tax_number"))

    if "logo_url" in payload:
        instance.logo_url = normalize_url(payload.get("logo_url"))

    if "logo_drive_file_id" in payload:
        instance.logo_drive_file_id = normalize_text(payload.get("logo_drive_file_id"))[:255]

    if "image_url" in payload:
        instance.image_url = normalize_url(payload.get("image_url"))

    if "image_drive_file_id" in payload:
        instance.image_drive_file_id = normalize_text(payload.get("image_drive_file_id"))[:255]

    if "drive_folder_id" in payload:
        instance.drive_folder_id = normalize_text(payload.get("drive_folder_id"))[:255]

    if "drive_folder_url" in payload:
        instance.drive_folder_url = normalize_url(payload.get("drive_folder_url"))

    if "contact_person" in payload:
        instance.contact_person = normalize_spaces(payload.get("contact_person"))

    if "phone" in payload:
        instance.phone = normalize_phone(payload.get("phone"))

    if "mobile" in payload:
        instance.mobile = normalize_phone(payload.get("mobile"))

    if "email" in payload:
        instance.email = normalize_text(payload.get("email"))

    if "website" in payload:
        instance.website = normalize_text(payload.get("website"))

    if "region" in payload:
        instance.region = normalize_spaces(payload.get("region"))

    if "city" in payload:
        instance.city = normalize_spaces(payload.get("city"))

    if "area" in payload:
        instance.area = normalize_spaces(payload.get("area"))

    if "street" in payload:
        instance.street = normalize_spaces(payload.get("street"))

    if "address" in payload:
        instance.address = normalize_spaces(payload.get("address"))

    if "google_maps_link" in payload:
        instance.google_maps_link = normalize_text(payload.get("google_maps_link"))

    if "source_category" in payload:
        instance.source_category = normalize_spaces(payload.get("source_category"))

    if "import_key" in payload:
        instance.import_key = normalize_text(payload.get("import_key")) or None

    if "import_source" in payload:
        instance.import_source = normalize_text(payload.get("import_source"))

    if "external_reference" in payload:
        instance.external_reference = normalize_text(payload.get("external_reference"))

    if "notes" in payload:
        instance.notes = normalize_spaces(payload.get("notes"))

    if "is_featured" in payload:
        instance.is_featured = parse_bool(payload.get("is_featured"), instance.is_featured)

    if "last_imported_at" in payload:
        instance.last_imported_at = payload.get("last_imported_at")

    sync_provider_legacy_name(instance)

    if not instance.name:
        raise ValidationError("Provider name is required.")

    if not instance.code:
        raise ValidationError("Provider code is required.")

    instance.provider_type = _validate_provider_type(instance.provider_type)
    instance.status = _validate_provider_status(instance.status)

    _validate_unique_code(instance.code, exclude_id=instance.id)
    _validate_unique_import_key(instance.import_key, exclude_id=instance.id)

    instance.full_clean()
    instance.save()
    return instance


# ============================================================
# 🔹 Provider Documents
# ============================================================

def create_provider_document(
    *,
    provider: Provider,
    file_type: str,
    file_url: str,
    drive_file_id: str = "",
    drive_folder_id: str = "",
    title: str = "",
    description: str = "",
    original_filename: str = "",
    content_type: str = "",
    size_bytes: int = 0,
    uploaded_by=None,
    is_primary: bool = False,
) -> ProviderDocument:
    file_type = _validate_provider_document_type(normalize_text(file_type) or ProviderDocumentType.OTHER)

    if not file_url:
        raise ValidationError("File URL is required.")

    document = ProviderDocument(
        provider=provider,
        file_type=file_type,
        title=normalize_spaces(title) or normalize_spaces(original_filename),
        description=normalize_spaces(description),
        file_url=normalize_url(file_url),
        drive_file_id=normalize_text(drive_file_id)[:255],
        drive_folder_id=normalize_text(drive_folder_id)[:255],
        original_filename=normalize_text(original_filename)[:255],
        content_type=normalize_text(content_type)[:150],
        size_bytes=max(int(size_bytes or 0), 0),
        uploaded_by=uploaded_by if getattr(uploaded_by, "is_authenticated", False) else None,
        is_primary=is_primary,
    )

    document.full_clean()
    document.save()

    if is_primary:
        ProviderDocument.objects.filter(
            provider=provider,
            file_type=file_type,
        ).exclude(pk=document.pk).update(is_primary=False)

    return document


def update_provider_main_file_from_document(*, provider: Provider, document: ProviderDocument) -> Provider:
    """
    Sync logo/image fields on Provider after uploading a primary document.
    """

    update_fields: list[str] = []

    if document.file_type == ProviderDocumentType.LOGO:
        provider.logo_url = document.file_url
        provider.logo_drive_file_id = document.drive_file_id
        update_fields.extend(["logo_url", "logo_drive_file_id", "updated_at"])

    if document.file_type == ProviderDocumentType.IMAGE:
        provider.image_url = document.file_url
        provider.image_drive_file_id = document.drive_file_id
        update_fields.extend(["image_url", "image_drive_file_id", "updated_at"])

    if document.drive_folder_id and not provider.drive_folder_id:
        provider.drive_folder_id = document.drive_folder_id
        update_fields.extend(["drive_folder_id", "updated_at"])

    if update_fields:
        provider.save(update_fields=list(dict.fromkeys(update_fields)))

    return provider


# ============================================================
# 🔹 Excel Import DTO
# ============================================================

@dataclass
class ProviderImportRow:
    row_number: int
    external_reference: str
    region: str
    city: str
    name: str
    source_category: str
    address: str
    area: str
    street: str
    phone: str
    mobile: str
    notes: str
    name_ar: str = ""
    name_en: str = ""
    commercial_registration: str = ""
    tax_number: str = ""


# ============================================================
# 🔹 Excel Helpers
# ============================================================

def _safe_cell_value(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return normalize_spaces(value)


def _normalize_excel_header(value: Any) -> str:
    return normalize_spaces(value)


def _header_lookup_key(value: Any) -> str:
    header = _normalize_excel_header(value)

    if not header:
        return ""

    header = header.replace("\ufeff", "")
    header = header.replace("\u200f", "")
    header = header.replace("\u200e", "")
    header = header.replace("ـ", "")
    header = header.strip().lower()

    # Normalize Arabic variants.
    header = header.replace("أ", "ا")
    header = header.replace("إ", "ا")
    header = header.replace("آ", "ا")
    header = header.replace("ى", "ي")
    header = header.replace("ة", "ه")

    # Remove punctuation and separators but keep Arabic/English letters and digits.
    header = re.sub(r"[\s\-_./\\|:;،,(){}\[\]#]+", "", header)
    return header


_NORMALIZED_EXCEL_HEADER_ALIASES = {
    _header_lookup_key(key): value
    for key, value in EXCEL_HEADER_ALIASES.items()
    if _header_lookup_key(key)
}


def _guess_excel_field_from_header(raw_header: Any) -> str | None:
    header = _normalize_excel_header(raw_header)
    key = _header_lookup_key(header)

    if not key:
        return None

    exact = _NORMALIZED_EXCEL_HEADER_ALIASES.get(key)
    if exact:
        return exact

    # Arabic heuristic detection.
    if "انجليزي" in key or "انجليزى" in key or "english" in key:
        if any(word in key for word in ["اسم", "name", "provider", "facility", "center", "hospital", "clinic"]):
            return "name_en"

    if "عربي" in key or "arabic" in key:
        if any(word in key for word in ["اسم", "name", "provider", "facility", "center", "hospital", "clinic"]):
            return "name_ar"

    if any(word in key for word in ["سجلتجاري", "السجلالتجاري", "commercialregistration", "crnumber"]):
        return "commercial_registration"

    if any(word in key for word in ["رقمضريبي", "الرقمالضريبي", "taxnumber", "vatnumber"]):
        return "tax_number"

    if any(word in key for word in ["اسم", "مقدم", "مزود", "مركز", "جهه", "منشاه", "مستشفي", "عياده", "مختبر", "صيدليه"]):
        return "name"

    if "منطقه" in key and "مدينه" not in key:
        return "region"

    if any(word in key for word in ["مدينه", "محافظه"]):
        return "city"

    if any(word in key for word in ["تصنيف", "فئه", "نوع", "تخصص", "نشاط"]):
        return "source_category"

    if any(word in key for word in ["عنوان", "موقع", "لوكيشن"]):
        return "address"

    if any(word in key for word in ["حي", "الحى"]):
        return "area"

    if "شارع" in key:
        return "street"

    if any(word in key for word in ["هاتف", "تلفون", "تليفون"]):
        return "phone"

    if any(word in key for word in ["جوال", "موبايل"]):
        return "mobile"

    if any(word in key for word in ["ملاحظ", "خصم", "خصومات"]):
        return "notes"

    if any(word in key for word in ["رقم", "مرجع", "كود"]) and not any(word in key for word in ["هاتف", "جوال", "ضريبي", "تجاري"]):
        return "external_reference"

    # English heuristic detection.
    if any(word in key for word in ["englishname", "nameen"]):
        return "name_en"

    if any(word in key for word in ["arabicname", "namear"]):
        return "name_ar"

    if any(word in key for word in ["commercialregistration", "crnumber"]):
        return "commercial_registration"

    if any(word in key for word in ["taxnumber", "vatnumber"]):
        return "tax_number"

    if any(word in key for word in ["providername", "serviceprovider", "facilityname", "centername", "centrename", "hospitalname", "clinicname"]):
        return "name"

    if key in {"provider", "serviceprovider", "facility", "center", "centre", "hospital", "clinic", "name"}:
        return "name"

    if "region" in key:
        return "region"

    if "city" in key or "province" in key:
        return "city"

    if any(word in key for word in ["category", "classification", "type", "specialty", "speciality", "activity"]):
        return "source_category"

    if "address" in key or "location" in key:
        return "address"

    if any(word in key for word in ["district", "neighborhood", "neighbourhood", "area"]):
        return "area"

    if "street" in key or "road" in key:
        return "street"

    if any(word in key for word in ["telephone", "phone", "tel"]):
        return "phone"

    if any(word in key for word in ["mobile", "cell"]):
        return "mobile"

    if any(word in key for word in ["note", "notes", "remark", "remarks", "discount"]):
        return "notes"

    if key in {"id", "no", "number", "serial", "reference"}:
        return "external_reference"

    return None


def _get_sheet(workbook, preferred_sheet_name: str | None = None):
    if preferred_sheet_name:
        if preferred_sheet_name not in workbook.sheetnames:
            raise ValidationError(f"Sheet '{preferred_sheet_name}' was not found in the Excel file.")
        return workbook[preferred_sheet_name]

    if "الشبكة كاملة" in workbook.sheetnames:
        return workbook["الشبكة كاملة"]

    if "ServiceProviders" in workbook.sheetnames:
        return workbook["ServiceProviders"]

    if "Providers" in workbook.sheetnames:
        return workbook["Providers"]

    if "providers" in workbook.sheetnames:
        return workbook["providers"]

    return workbook[workbook.sheetnames[0]]


def _build_header_map(sheet) -> tuple[int, dict[int, str]]:
    best_header_row_number = 1
    best_header_map: dict[int, str] = {}
    best_score = 0
    scanned_headers: list[str] = []

    max_scan_row = min(sheet.max_row or EXCEL_SCAN_HEADER_ROWS, EXCEL_SCAN_HEADER_ROWS)

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
        start=1,
    ):
        header_map: dict[int, str] = {}
        row_headers: list[str] = []

        for index, raw_header in enumerate(row):
            header = _normalize_excel_header(raw_header)
            if header:
                row_headers.append(header)

            field_name = _guess_excel_field_from_header(raw_header)
            if field_name and field_name not in header_map.values():
                header_map[index] = field_name

        found_fields = set(header_map.values())

        # Scoring:
        # - name/name_ar/name_en are critical.
        # - More mapped columns means more likely this is the real header row.
        score = len(found_fields)
        if "name" in found_fields or "name_ar" in found_fields or "name_en" in found_fields:
            score += 10
        if "city" in found_fields:
            score += 2
        if "phone" in found_fields or "mobile" in found_fields:
            score += 1

        if score > best_score:
            best_score = score
            best_header_row_number = row_number
            best_header_map = header_map
            scanned_headers = row_headers

    found_fields = set(best_header_map.values())

    if not ({"name", "name_ar", "name_en"} & found_fields):
        visible_headers = " | ".join(scanned_headers[:20]) if scanned_headers else "-"
        raise ValidationError(
            f"Excel file is missing required columns. Required: اسم المركز / اسم مقدم الخدمة / Provider Name. Detected headers: {visible_headers}"
        )

    return best_header_row_number, best_header_map


def read_providers_from_excel(
    *,
    uploaded_file,
    sheet_name: str | None = None,
) -> list[ProviderImportRow]:
    filename = normalize_text(getattr(uploaded_file, "name", ""))

    if filename.lower().endswith(".xls"):
        raise ValidationError("Unsupported Excel format .xls. Please save the file as .xlsx and upload it again.")

    if not filename.lower().endswith(".xlsx"):
        raise ValidationError("Unsupported file format. Please upload an .xlsx Excel file.")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("openpyxl is required to import Excel files. Please install it first.") from exc

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(f"Unable to read Excel file: {exc}") from exc

    sheet = _get_sheet(workbook, sheet_name)
    header_row_number, header_map = _build_header_map(sheet)

    rows: list[ProviderImportRow] = []

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
        start=header_row_number + 1,
    ):
        raw_data: dict[str, str] = {
            "external_reference": "",
            "region": "",
            "city": "",
            "name": "",
            "name_ar": "",
            "name_en": "",
            "commercial_registration": "",
            "tax_number": "",
            "source_category": "",
            "address": "",
            "area": "",
            "street": "",
            "phone": "",
            "mobile": "",
            "notes": "",
        }

        for index, value in enumerate(row):
            field_name = header_map.get(index)
            if not field_name:
                continue

            raw_data[field_name] = _safe_cell_value(value)

        name = normalize_spaces(raw_data.get("name"))
        name_ar = normalize_spaces(raw_data.get("name_ar"))
        name_en = normalize_spaces(raw_data.get("name_en"))
        resolved_name = name or name_ar or name_en

        # Skip fully empty rows.
        if not any(normalize_text(value) for value in raw_data.values()):
            continue

        rows.append(
            ProviderImportRow(
                row_number=row_number,
                external_reference=normalize_text(raw_data.get("external_reference")),
                region=normalize_spaces(raw_data.get("region")),
                city=normalize_spaces(raw_data.get("city")),
                name=resolved_name,
                name_ar=name_ar or resolved_name,
                name_en=name_en,
                commercial_registration=normalize_identifier(raw_data.get("commercial_registration")),
                tax_number=normalize_identifier(raw_data.get("tax_number")),
                source_category=normalize_spaces(raw_data.get("source_category")),
                address=normalize_spaces(raw_data.get("address")),
                area=normalize_spaces(raw_data.get("area")),
                street=normalize_spaces(raw_data.get("street")),
                phone=normalize_phone(raw_data.get("phone")),
                mobile=normalize_phone(raw_data.get("mobile")),
                notes=normalize_spaces(raw_data.get("notes")),
            )
        )

    return rows


# ============================================================
# 🔹 Excel Upsert
# ============================================================

def _provider_payload_from_import_row(row: ProviderImportRow) -> dict[str, Any]:
    provider_type = guess_provider_type(row.source_category, row.name)

    import_key = build_provider_import_key(
        name=row.name,
        region=row.region,
        city=row.city,
        area=row.area,
        address=row.address,
        source_category=row.source_category,
        import_source=MEDICAL_NETWORK_IMPORT_SOURCE,
    )

    return {
        "name": row.name,
        "name_ar": row.name_ar or row.name,
        "name_en": row.name_en,
        "code": build_provider_code_from_import_key(import_key),
        "provider_type": provider_type,
        "status": ProviderStatus.ACTIVE,
        "commercial_registration": row.commercial_registration,
        "tax_number": row.tax_number,
        "contact_person": "",
        "phone": row.phone,
        "mobile": row.mobile,
        "email": "",
        "website": "",
        "region": row.region,
        "city": row.city,
        "area": row.area,
        "street": row.street,
        "address": row.address,
        "google_maps_link": "",
        "source_category": row.source_category,
        "import_key": import_key,
        "import_source": MEDICAL_NETWORK_IMPORT_SOURCE,
        "external_reference": row.external_reference,
        "notes": row.notes,
        "is_featured": False,
        "last_imported_at": timezone.now(),
    }


def upsert_provider_from_import_row(
    *,
    row: ProviderImportRow,
    dry_run: bool = False,
) -> dict[str, Any]:
    if not row.name:
        return {
            "ok": False,
            "action": "skipped",
            "row_number": row.row_number,
            "message": "Provider name is required.",
            "provider": None,
        }

    payload = _provider_payload_from_import_row(row)
    import_key = payload["import_key"]

    existing = Provider.objects.filter(import_key=import_key).first()

    if dry_run:
        return {
            "ok": True,
            "action": "updated" if existing else "created",
            "row_number": row.row_number,
            "message": "Dry run only. No changes were saved.",
            "provider": {
                "id": existing.id if existing else None,
                "name": payload["name"],
                "name_ar": payload["name_ar"],
                "name_en": payload["name_en"],
                "code": existing.code if existing else payload["code"],
                "city": payload["city"],
                "region": payload["region"],
                "commercial_registration": payload["commercial_registration"],
                "tax_number": payload["tax_number"],
                "source_category": payload["source_category"],
            },
        }

    if existing:
        # Keep manual code stable if it already exists.
        payload["code"] = existing.code
        provider = update_provider(instance=existing, payload=payload)
        action = "updated"
    else:
        provider = create_provider(payload=payload)
        action = "created"

    return {
        "ok": True,
        "action": action,
        "row_number": row.row_number,
        "message": "Provider imported successfully.",
        "provider": serialize_provider(provider),
    }


def import_providers_from_excel(
    *,
    uploaded_file,
    sheet_name: str | None = None,
    dry_run: bool = False,
    sample_limit: int = 25,
) -> dict[str, Any]:
    rows = read_providers_from_excel(uploaded_file=uploaded_file, sheet_name=sheet_name)

    summary = {
        "ok": True,
        "dry_run": dry_run,
        "total_rows": len(rows),
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "failed": 0,
        "errors": [],
        "sample": [],
    }

    # Prevent duplicate rows inside the same Excel file from creating noise.
    seen_import_keys: set[str] = set()

    def process_rows() -> None:
        for row in rows:
            try:
                if not row.name:
                    summary["skipped"] += 1
                    summary["errors"].append(
                        {
                            "row_number": row.row_number,
                            "message": "Skipped row because provider name is empty.",
                        }
                    )
                    continue

                payload = _provider_payload_from_import_row(row)
                import_key = payload["import_key"]

                if import_key in seen_import_keys:
                    summary["skipped"] += 1
                    summary["errors"].append(
                        {
                            "row_number": row.row_number,
                            "message": "Skipped duplicate row inside the same Excel file.",
                            "name": row.name,
                        }
                    )
                    continue

                seen_import_keys.add(import_key)

                result = upsert_provider_from_import_row(row=row, dry_run=dry_run)

                action = result.get("action")

                if action == "created":
                    summary["created"] += 1
                elif action == "updated":
                    summary["updated"] += 1
                elif action == "skipped":
                    summary["skipped"] += 1
                else:
                    summary["failed"] += 1

                if len(summary["sample"]) < sample_limit:
                    summary["sample"].append(result)

            except ValidationError as exc:
                summary["failed"] += 1
                summary["errors"].append(
                    {
                        "row_number": row.row_number,
                        "name": row.name,
                        "message": "Validation failed.",
                        "errors": exc.messages,
                    }
                )
            except Exception as exc:
                summary["failed"] += 1
                summary["errors"].append(
                    {
                        "row_number": row.row_number,
                        "name": row.name,
                        "message": str(exc),
                    }
                )

    if dry_run:
        process_rows()
    else:
        with transaction.atomic():
            process_rows()

    return summary
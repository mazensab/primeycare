# ===================================================================
# 📂 performance_center/reports.py
# 🧭 Primey Care - Performance Center Reports Engine
# ===================================================================

from django.http import HttpResponse

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .models import PerformanceAnswer, PerformanceReview


# ================================================================
# 🅰️ إعداد خط PDF
# ================================================================
try:
    pdfmetrics.registerFont(TTFont("Tajawal", "static/fonts/Tajawal-Regular.ttf"))
except Exception:
    pass


def draw_rtl_text(c, text, x, y, size=12):
    c.setFont("Tajawal", size)
    c.drawRightString(x, y, str(text or ""))


# ===================================================================
# 📘 generate_review_pdf
# ===================================================================
def generate_review_pdf(review_id: int):
    review = PerformanceReview.objects.get(id=review_id)
    answers = PerformanceAnswer.objects.filter(review=review).select_related("item")

    response = HttpResponse(content_type="application/pdf")
    filename = f"performance_review_{review.id}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    margin_x = 190 * mm
    current_y = 270 * mm

    subject_name = review.employee_name or (
        f"Subject #{review.employee_id}" if review.employee_id else "Unknown Subject"
    )

    draw_rtl_text(c, "تقرير تقييم الأداء", margin_x, current_y, size=20)
    current_y -= 20

    draw_rtl_text(c, f"الهدف: {subject_name}", margin_x, current_y, size=13)
    current_y -= 10

    draw_rtl_text(c, f"القالب: {review.template.name}", margin_x, current_y, size=12)
    current_y -= 10

    draw_rtl_text(c, f"الفترة: {review.period_label}", margin_x, current_y, size=12)
    current_y -= 25

    draw_rtl_text(c, "تفاصيل التقييم:", margin_x, current_y, size=15)
    current_y -= 15

    for ans in answers:
        if current_y < 40:
            c.showPage()
            current_y = 270 * mm
            try:
                c.setFont("Tajawal", 12)
            except Exception:
                pass

        draw_rtl_text(c, f"السؤال: {ans.item.question}", margin_x, current_y)
        current_y -= 8

        draw_rtl_text(c, f"درجة التقييم الذاتي: {ans.self_score if ans.self_score is not None else '—'}", margin_x, current_y)
        current_y -= 8

        draw_rtl_text(c, f"درجة المدير: {ans.manager_score if ans.manager_score is not None else '—'}", margin_x, current_y)
        current_y -= 8

        draw_rtl_text(c, f"درجة HR: {ans.hr_score if ans.hr_score is not None else '—'}", margin_x, current_y)
        current_y -= 8

        note_text = ans.hr_answer or ans.manager_answer or ans.self_answer or "—"
        draw_rtl_text(c, f"ملاحظات: {note_text}", margin_x, current_y)
        current_y -= 15

    c.save()
    return response


# ===================================================================
# 📘 generate_employee_summary_pdf
# ===================================================================
def generate_employee_summary_pdf(employee_id: int):
    reviews = PerformanceReview.objects.filter(employee_id=employee_id)

    response = HttpResponse(content_type="application/pdf")
    filename = f"subject_summary_{employee_id}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    margin_x = 190 * mm
    current_y = 270 * mm

    draw_rtl_text(c, f"تقرير شامل — الهدف رقم {employee_id}", margin_x, current_y, size=20)
    current_y -= 20

    for review in reviews:
        draw_rtl_text(c, f"- قالب: {review.template.name}", margin_x, current_y)
        current_y -= 10

        draw_rtl_text(c, f"  النتيجة النهائية: {review.final_score or '—'}", margin_x, current_y)
        current_y -= 10

        draw_rtl_text(c, f"  الحالة: {review.status}", margin_x, current_y)
        current_y -= 15

        if current_y < 40:
            c.showPage()
            current_y = 270 * mm
            try:
                c.setFont("Tajawal", 12)
            except Exception:
                pass

    c.save()
    return response


# ===================================================================
# 📘 export_reviews_excel
# ===================================================================
def export_reviews_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Reviews"

    headers = [
        "الهدف",
        "القالب",
        "الفترة",
        "الحالة",
        "النتيجة النهائية",
        "آخر تحديث",
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for review in PerformanceReview.objects.select_related("template").all():
        subject_name = review.employee_name or (
            f"Subject #{review.employee_id}" if review.employee_id else "Unknown Subject"
        )
        ws.append([
            subject_name,
            review.template.name,
            review.period_label,
            review.status,
            review.final_score,
            review.updated_at.strftime("%Y-%m-%d") if review.updated_at else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="performance_reviews.xlsx"'
    wb.save(response)
    return response
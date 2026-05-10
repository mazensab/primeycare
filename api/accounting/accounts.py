# ============================================================
# 📂 api/accounting/accounts.py
# 🧠 Accounting Accounts API — Primey Care V2
# ------------------------------------------------------------
# ✅ قائمة دليل الحسابات
# ✅ شجرة الحسابات
# ✅ ملخص الحسابات
# ✅ Excel Export
# ✅ متوافق مع Accounting Backend الجديد
# ✅ يدعم:
#    - search
#    - account_type
#    - nature
#    - is_group
#    - is_active
#    - allow_manual_posting
#    - is_system
#    - can_post
#    - parent_id
#    - level
#    - currency
#    - tree
#    - page / page_size
#    - ordering
# ============================================================

from __future__ import annotations

import logging
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from django.core.paginator import EmptyPage, Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import Account, AccountNature, AccountType


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Constants
# ============================================================

DEFAULT_PAGE = 1
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 500

ALLOWED_ORDERING = {
    "code": "code",
    "-code": "-code",
    "name": "name",
    "-name": "-name",
    "name_en": "name_en",
    "-name_en": "-name_en",
    "level": "level",
    "-level": "-level",
    "account_type": "account_type",
    "-account_type": "-account_type",
    "nature": "nature",
    "-nature": "-nature",
    "currency": "currency",
    "-currency": "-currency",
    "opening_balance": "opening_balance",
    "-opening_balance": "-opening_balance",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "updated_at": "updated_at",
    "-updated_at": "-updated_at",
}


# ============================================================
# 🔧 Helpers
# ============================================================

def _money(value: Decimal | int | float | str | None) -> Decimal:
    amount = Decimal(str(value or "0.00"))
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool | None = None) -> bool | None:
    if value in {None, ""}:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    raise ValueError("قيمة boolean غير صحيحة. استخدم true أو false.")


def _parse_int(value: str | None, field_name: str) -> int | None:
    if value in {None, ""}:
        return None

    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed <= 0:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر من صفر.")

    return parsed


def _parse_positive_int(
    value: str | None,
    field_name: str,
    *,
    default: int,
    min_value: int = 1,
    max_value: int | None = None,
) -> int:
    if value in {None, ""}:
        parsed = default
    else:
        try:
            parsed = int(str(value).strip())
        except ValueError as exc:
            raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed < min_value:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر أو تساوي {min_value}.")

    if max_value is not None and parsed > max_value:
        raise ValueError(f"قيمة {field_name} يجب ألا تتجاوز {max_value}.")

    return parsed


def _parse_ordering(value: str | None, default: str = "code") -> str:
    raw = (value or "").strip() or default

    if raw not in ALLOWED_ORDERING:
        allowed = ", ".join(ALLOWED_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_ORDERING[raw]


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any]) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=200,
        json_dumps_params={"ensure_ascii": False},
    )


def _safe_attr(obj: Any, attr_name: str, default: Any = None) -> Any:
    try:
        return getattr(obj, attr_name, default)
    except Exception:
        return default


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _account_display_name(account: Account | None) -> str | None:
    if not account:
        return None

    return (
        _safe_attr(account, "name", None)
        or _safe_attr(account, "name_ar", None)
        or _safe_attr(account, "name_en", None)
        or f"Account #{account.pk}"
    )


def _can_post(account: Account) -> bool:
    return bool(
        _safe_attr(account, "is_active", True)
        and not _safe_attr(account, "is_group", False)
    )


def _can_manual_post(account: Account) -> bool:
    return bool(
        _can_post(account)
        and _safe_attr(account, "allow_manual_posting", True)
    )


def _serialize_account(
    account: Account,
    *,
    include_children_count: bool = True,
    include_metadata: bool = False,
) -> dict[str, Any]:
    children_count = 0

    if include_children_count:
        try:
            children_count = account.children.count()
        except Exception:
            children_count = 0

    parent = _safe_attr(account, "parent", None)

    payload = {
        "id": account.id,
        "code": _safe_attr(account, "code", None),
        "name": _account_display_name(account),
        "name_ar": _safe_attr(account, "name", None),
        "name_en": _safe_attr(account, "name_en", ""),
        "description": _safe_attr(account, "description", ""),
        "account_type": _safe_attr(account, "account_type", None),
        "account_type_label": (
            account.get_account_type_display()
            if hasattr(account, "get_account_type_display")
            else None
        ),
        "nature": _safe_attr(account, "nature", None),
        "nature_label": (
            account.get_nature_display()
            if hasattr(account, "get_nature_display")
            else None
        ),
        "parent_id": _safe_attr(account, "parent_id", None),
        "parent_code": _safe_attr(parent, "code", None),
        "parent_name": _account_display_name(parent) if parent else None,
        "level": int(_safe_attr(account, "level", 1) or 1),
        "is_group": bool(_safe_attr(account, "is_group", False)),
        "is_active": bool(_safe_attr(account, "is_active", True)),
        "allow_manual_posting": bool(_safe_attr(account, "allow_manual_posting", True)),
        "is_system": bool(_safe_attr(account, "is_system", False)),
        "currency": _safe_attr(account, "currency", "SAR"),
        "opening_balance": _money(_safe_attr(account, "opening_balance", "0.00")),
        "can_post": _can_post(account),
        "can_manual_post": _can_manual_post(account),
        "children_count": children_count,
        "has_children": children_count > 0,
        "created_at": (
            _safe_attr(account, "created_at", None).isoformat()
            if _safe_attr(account, "created_at", None)
            else None
        ),
        "updated_at": (
            _safe_attr(account, "updated_at", None).isoformat()
            if _safe_attr(account, "updated_at", None)
            else None
        ),
    }

    if include_metadata:
        payload["metadata"] = _safe_attr(account, "metadata", {}) or {}

    return payload


def _build_accounts_queryset(request):
    search = (request.GET.get("search") or "").strip()
    account_type = (request.GET.get("account_type") or "").strip()
    nature = (request.GET.get("nature") or "").strip()
    currency = (request.GET.get("currency") or "").strip().upper()

    is_group = _parse_bool(request.GET.get("is_group"), default=None)
    is_active = _parse_bool(request.GET.get("is_active"), default=None)
    allow_manual_posting = _parse_bool(request.GET.get("allow_manual_posting"), default=None)
    is_system = _parse_bool(request.GET.get("is_system"), default=None)
    can_post = _parse_bool(request.GET.get("can_post"), default=None)

    parent_id = _parse_int(request.GET.get("parent_id"), "parent_id")
    level = _parse_int(request.GET.get("level"), "level")

    qs = Account.objects.select_related("parent").all()

    if search:
        qs = qs.filter(
            Q(code__icontains=search)
            | Q(name__icontains=search)
            | Q(name_en__icontains=search)
            | Q(description__icontains=search)
            | Q(currency__icontains=search)
            | Q(parent__code__icontains=search)
            | Q(parent__name__icontains=search)
            | Q(parent__name_en__icontains=search)
        )

    if account_type:
        valid_types = {choice[0] for choice in AccountType.choices}
        if account_type not in valid_types:
            raise ValueError("نوع الحساب غير صحيح.")
        qs = qs.filter(account_type=account_type)

    if nature:
        valid_natures = {choice[0] for choice in AccountNature.choices}
        if nature not in valid_natures:
            raise ValueError("طبيعة الحساب غير صحيحة.")
        qs = qs.filter(nature=nature)

    if currency:
        qs = qs.filter(currency__iexact=currency)

    if is_group is not None:
        qs = qs.filter(is_group=is_group)

    if is_active is not None:
        qs = qs.filter(is_active=is_active)

    if allow_manual_posting is not None:
        qs = qs.filter(allow_manual_posting=allow_manual_posting)

    if is_system is not None:
        qs = qs.filter(is_system=is_system)

    if can_post is True:
        qs = qs.filter(is_active=True, is_group=False)

    if can_post is False:
        qs = qs.exclude(is_active=True, is_group=False)

    if parent_id:
        qs = qs.filter(parent_id=parent_id)

    if level:
        qs = qs.filter(level=level)

    return {
        "qs": qs,
        "filters": {
            "search": search or None,
            "account_type": account_type or None,
            "nature": nature or None,
            "currency": currency or None,
            "is_group": is_group,
            "is_active": is_active,
            "allow_manual_posting": allow_manual_posting,
            "is_system": is_system,
            "can_post": can_post,
            "parent_id": parent_id,
            "level": level,
        },
    }


def _build_accounts_tree(accounts: list[Account]) -> list[dict[str, Any]]:
    account_map: dict[int, dict[str, Any]] = {}
    roots: list[dict[str, Any]] = []

    for account in accounts:
        item = _serialize_account(
            account,
            include_children_count=False,
            include_metadata=False,
        )
        item["children"] = []
        account_map[account.id] = item

    for account in accounts:
        item = account_map[account.id]
        parent_id = _safe_attr(account, "parent_id", None)

        if parent_id and parent_id in account_map:
            account_map[parent_id]["children"].append(item)
        else:
            roots.append(item)

    return roots


def _build_accounts_summary(accounts: list[Account]) -> dict[str, Any]:
    total_accounts = len(accounts)
    active_accounts = sum(1 for item in accounts if item.is_active)
    group_accounts = sum(1 for item in accounts if item.is_group)
    posting_accounts = sum(1 for item in accounts if item.is_active and not item.is_group)
    manual_posting_accounts = sum(
        1
        for item in accounts
        if item.is_active and not item.is_group and item.allow_manual_posting
    )
    system_accounts = sum(1 for item in accounts if item.is_system)

    summary_by_type: dict[str, dict[str, Any]] = {}
    summary_by_nature: dict[str, dict[str, Any]] = {}

    for account in accounts:
        account_type = str(account.account_type or "")
        nature = str(account.nature or "")

        if account_type not in summary_by_type:
            summary_by_type[account_type] = {
                "account_type": account_type,
                "account_type_label": account.get_account_type_display()
                if hasattr(account, "get_account_type_display")
                else account_type,
                "total": 0,
                "posting_accounts": 0,
                "group_accounts": 0,
                "active_accounts": 0,
            }

        summary_by_type[account_type]["total"] += 1
        summary_by_type[account_type]["posting_accounts"] += 1 if _can_post(account) else 0
        summary_by_type[account_type]["group_accounts"] += 1 if account.is_group else 0
        summary_by_type[account_type]["active_accounts"] += 1 if account.is_active else 0

        if nature not in summary_by_nature:
            summary_by_nature[nature] = {
                "nature": nature,
                "nature_label": account.get_nature_display()
                if hasattr(account, "get_nature_display")
                else nature,
                "total": 0,
            }

        summary_by_nature[nature]["total"] += 1

    return {
        "total_accounts": total_accounts,
        "active_accounts": active_accounts,
        "inactive_accounts": total_accounts - active_accounts,
        "group_accounts": group_accounts,
        "posting_accounts": posting_accounts,
        "manual_posting_accounts": manual_posting_accounts,
        "system_accounts": system_accounts,
        "non_system_accounts": total_accounts - system_accounts,
        "total_opening_balance": _money(
            sum(
                (_money(_safe_attr(item, "opening_balance", "0.00")) for item in accounts),
                Decimal("0.00"),
            )
        ),
        "by_type": list(summary_by_type.values()),
        "by_nature": list(summary_by_nature.values()),
    }


def _choices_payload() -> dict[str, Any]:
    return {
        "account_types": [
            {"value": value, "label": label}
            for value, label in AccountType.choices
        ],
        "natures": [
            {"value": value, "label": label}
            for value, label in AccountNature.choices
        ],
        "orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_ORDERING.keys()
        ],
    }


def _build_accounts_excel(
    accounts: list[Account],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Accounts")

    ws.append(["Chart of Accounts"])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)

    meta_rows = [
        ("Search", filters.get("search")),
        ("Account Type", filters.get("account_type")),
        ("Nature", filters.get("nature")),
        ("Currency", filters.get("currency")),
        ("Is Group", filters.get("is_group")),
        ("Is Active", filters.get("is_active")),
        ("Allow Manual Posting", filters.get("allow_manual_posting")),
        ("Is System", filters.get("is_system")),
        ("Can Post", filters.get("can_post")),
        ("Parent ID", filters.get("parent_id")),
        ("Level", filters.get("level")),
        ("Total Accounts", summary.get("total_accounts")),
        ("Posting Accounts", summary.get("posting_accounts")),
        ("Manual Posting Accounts", summary.get("manual_posting_accounts")),
        ("Group Accounts", summary.get("group_accounts")),
        ("System Accounts", summary.get("system_accounts")),
        ("Total Opening Balance", summary.get("total_opening_balance")),
    ]

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    start_row = current_row + 1

    headers = [
        "ID",
        "Code",
        "Name AR",
        "Name EN",
        "Display Name",
        "Type",
        "Type Label",
        "Nature",
        "Nature Label",
        "Parent ID",
        "Parent Code",
        "Parent Name",
        "Level",
        "Currency",
        "Opening Balance",
        "Is Group",
        "Is Active",
        "Allow Manual Posting",
        "Is System",
        "Can Post",
        "Can Manual Post",
        "Children Count",
        "Description",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1
    for account in accounts:
        row = _serialize_account(account)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("code"))
        ws.cell(row=current_row, column=3, value=row.get("name_ar"))
        ws.cell(row=current_row, column=4, value=row.get("name_en"))
        ws.cell(row=current_row, column=5, value=row.get("name"))
        ws.cell(row=current_row, column=6, value=row.get("account_type"))
        ws.cell(row=current_row, column=7, value=row.get("account_type_label"))
        ws.cell(row=current_row, column=8, value=row.get("nature"))
        ws.cell(row=current_row, column=9, value=row.get("nature_label"))
        ws.cell(row=current_row, column=10, value=row.get("parent_id"))
        ws.cell(row=current_row, column=11, value=row.get("parent_code"))
        ws.cell(row=current_row, column=12, value=row.get("parent_name"))
        ws.cell(row=current_row, column=13, value=row.get("level"))
        ws.cell(row=current_row, column=14, value=row.get("currency"))
        ws.cell(row=current_row, column=15, value=float(row.get("opening_balance", 0) or 0))
        ws.cell(row=current_row, column=16, value=str(row.get("is_group")))
        ws.cell(row=current_row, column=17, value=str(row.get("is_active")))
        ws.cell(row=current_row, column=18, value=str(row.get("allow_manual_posting")))
        ws.cell(row=current_row, column=19, value=str(row.get("is_system")))
        ws.cell(row=current_row, column=20, value=str(row.get("can_post")))
        ws.cell(row=current_row, column=21, value=str(row.get("can_manual_post")))
        ws.cell(row=current_row, column=22, value=row.get("children_count"))
        ws.cell(row=current_row, column=23, value=row.get("description"))
        ws.cell(row=current_row, column=24, value=row.get("created_at"))
        ws.cell(row=current_row, column=25, value=row.get("updated_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "chart_of_accounts.xlsx")


# ============================================================
# 📘 Accounts List API
# ============================================================

@require_GET
def accounting_accounts_api(request):
    """
    قائمة دليل الحسابات.
    """
    try:
        query_data = _build_accounts_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_ordering(request.GET.get("ordering"), default="code")
        as_tree = _parse_bool(request.GET.get("tree"), default=False)
        include_metadata = _parse_bool(request.GET.get("include_metadata"), default=False)

        if ordering.lstrip("-") == "code":
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by(ordering, "code")

        accounts_all = list(qs)
        summary = _build_accounts_summary(accounts_all)

        if as_tree:
            payload = {
                "filters": {
                    **query_data["filters"],
                    "ordering": ordering,
                    "tree": True,
                    "include_metadata": False,
                },
                "summary": summary,
                "choices": _choices_payload(),
                "results": _build_accounts_tree(accounts_all),
            }
            return _success_response(payload)

        paginator = Paginator(accounts_all, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        payload = {
            "filters": {
                **query_data["filters"],
                "ordering": ordering,
                "tree": False,
                "include_metadata": include_metadata,
            },
            "summary": summary,
            "choices": _choices_payload(),
            "pagination": {
                "page": page_obj.number,
                "page_size": page_size,
                "total_pages": paginator.num_pages,
                "total_items": paginator.count,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
                "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
            },
            "results": [
                _serialize_account(account, include_metadata=include_metadata)
                for account in page_obj.object_list
            ],
        }

        return _success_response(payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load accounting accounts: %s", exc)
        return _error_response(
            "تعذر تحميل دليل الحسابات.",
            status=500,
        )


@require_GET
def accounting_accounts_excel_api(request):
    """
    تصدير دليل الحسابات إلى Excel.
    """
    try:
        query_data = _build_accounts_queryset(request)
        ordering = _parse_ordering(request.GET.get("ordering"), default="code")

        if ordering.lstrip("-") == "code":
            qs = query_data["qs"].order_by(ordering)
        else:
            qs = query_data["qs"].order_by(ordering, "code")

        accounts_all = list(qs)
        summary = _build_accounts_summary(accounts_all)

        return _build_accounts_excel(
            accounts_all,
            query_data["filters"],
            summary,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export accounting accounts excel: %s", exc)
        return _error_response(
            "تعذر تصدير دليل الحسابات إلى Excel.",
            status=500,
        )
# ============================================================
# 📂 api/accounting/routing.py
# 🧠 Accounting Routing API — Primey Care V2
# ------------------------------------------------------------
# ✅ إدارة قواعد التوجيه المحاسبي Accounting Routing Rules
# ✅ إدارة إعدادات المحاسبة العامة Accounting Settings
# ✅ قائمة / تفاصيل / إنشاء / تحديث
# ✅ تفعيل / تعطيل قاعدة توجيه
# ✅ Excel Export
# ✅ متوافق مع Accounting Backend الجديد
# ============================================================

from __future__ import annotations

import json
import logging
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import (
    Account,
    AccountingAccountPurpose,
    AccountingRoutingRule,
    AccountingRoutingSource,
    AccountingSettings,
    CostCenter,
    TaxRate,
)


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Constants
# ============================================================

DEFAULT_PAGE = 1
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 500

ALLOWED_ROUTING_ORDERING = {
    "source": "source",
    "-source": "-source",
    "purpose": "purpose",
    "-purpose": "-purpose",
    "priority": "priority",
    "-priority": "-priority",
    "is_active": "is_active",
    "-is_active": "-is_active",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "updated_at": "updated_at",
    "-updated_at": "-updated_at",
}


# ============================================================
# 🔧 Helpers
# ============================================================

def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool | None = None) -> bool | None:
    if value in {None, ""}:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    raise ValueError("قيمة boolean غير صحيحة. استخدم true أو false.")


def _parse_int(value: str | None, field_name: str) -> int | None:
    if value in {None, ""}:
        return None

    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed <= 0:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر من صفر.")

    return parsed


def _parse_positive_int(
    value: str | None,
    field_name: str,
    *,
    default: int,
    min_value: int = 1,
    max_value: int | None = None,
) -> int:
    if value in {None, ""}:
        parsed = default
    else:
        try:
            parsed = int(str(value).strip())
        except ValueError as exc:
            raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed < min_value:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر أو تساوي {min_value}.")

    if max_value is not None and parsed > max_value:
        raise ValueError(f"قيمة {field_name} يجب ألا تتجاوز {max_value}.")

    return parsed


def _parse_ordering(value: str | None, default: str = "source") -> str:
    raw = (value or "").strip() or default

    if raw not in ALLOWED_ROUTING_ORDERING:
        allowed = ", ".join(ALLOWED_ROUTING_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_ROUTING_ORDERING[raw]


def _json_body(request) -> dict[str, Any]:
    try:
        body = request.body.decode("utf-8") if request.body else "{}"
        data = json.loads(body or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError("صيغة JSON غير صحيحة.") from exc

    if not isinstance(data, dict):
        raise ValueError("جسم الطلب يجب أن يكون JSON Object.")

    return data


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any], status: int = 200) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _validation_error_payload(exc: ValidationError) -> dict[str, Any]:
    if hasattr(exc, "message_dict"):
        return {
            "ok": False,
            "message": "تعذر حفظ البيانات.",
            "errors": exc.message_dict,
        }

    return {
        "ok": False,
        "message": "تعذر حفظ البيانات.",
        "errors": exc.messages if hasattr(exc, "messages") else [str(exc)],
    }


def _status_label(value: str, choices) -> str:
    for choice_value, label in choices:
        if choice_value == value:
            return label
    return value


def _safe_attr(obj: Any, attr_name: str, default: Any = None) -> Any:
    try:
        return getattr(obj, attr_name, default)
    except Exception:
        return default


def _iso_datetime(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return None


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# 🧾 Shared Serializers / Resolvers
# ============================================================

def _resolve_account(account_id: Any, field_name: str = "account_id") -> Account | None:
    if account_id in {None, ""}:
        return None

    try:
        parsed_id = int(account_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({field_name: "معرف الحساب غير صحيح."}) from exc

    account = Account.objects.filter(id=parsed_id).first()

    if not account:
        raise ValidationError({field_name: "الحساب المطلوب غير موجود."})

    if account.is_group:
        raise ValidationError({field_name: "لا يمكن استخدام حساب تجميعي في التوجيه."})

    if not account.is_active:
        raise ValidationError({field_name: "لا يمكن استخدام حساب غير نشط في التوجيه."})

    return account


def _resolve_tax_rate(tax_rate_id: Any) -> TaxRate | None:
    if tax_rate_id in {None, ""}:
        return None

    try:
        parsed_id = int(tax_rate_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({"tax_rate_id": "معرف الضريبة غير صحيح."}) from exc

    tax_rate = TaxRate.objects.filter(id=parsed_id).first()

    if not tax_rate:
        raise ValidationError({"tax_rate_id": "الضريبة المطلوبة غير موجودة."})

    if not tax_rate.is_active:
        raise ValidationError({"tax_rate_id": "لا يمكن استخدام ضريبة غير نشطة."})

    return tax_rate


def _resolve_cost_center(cost_center_id: Any) -> CostCenter | None:
    if cost_center_id in {None, ""}:
        return None

    try:
        parsed_id = int(cost_center_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({"cost_center_id": "معرف مركز التكلفة غير صحيح."}) from exc

    cost_center = CostCenter.objects.filter(id=parsed_id).first()

    if not cost_center:
        raise ValidationError({"cost_center_id": "مركز التكلفة المطلوب غير موجود."})

    if cost_center.is_group:
        raise ValidationError({"cost_center_id": "لا يمكن استخدام مركز تكلفة تجميعي في التوجيه."})

    if not cost_center.can_post:
        raise ValidationError({"cost_center_id": "مركز التكلفة غير قابل للترحيل."})

    return cost_center


def _serialize_account(account: Account | None) -> dict[str, Any] | None:
    if not account:
        return None

    return {
        "id": account.id,
        "code": account.code,
        "name": account.name,
        "name_en": account.name_en,
        "account_type": account.account_type,
        "account_type_label": (
            account.get_account_type_display()
            if hasattr(account, "get_account_type_display")
            else account.account_type
        ),
        "nature": account.nature,
        "nature_label": (
            account.get_nature_display()
            if hasattr(account, "get_nature_display")
            else account.nature
        ),
        "is_group": account.is_group,
        "is_active": account.is_active,
        "allow_manual_posting": account.allow_manual_posting,
        "is_system": account.is_system,
        "currency": account.currency,
    }


def _serialize_tax_rate(tax_rate: TaxRate | None) -> dict[str, Any] | None:
    if not tax_rate:
        return None

    return {
        "id": tax_rate.id,
        "code": tax_rate.code,
        "name": tax_rate.name,
        "tax_type": tax_rate.tax_type,
        "rate": tax_rate.rate,
        "is_active": tax_rate.is_active,
        "is_default": tax_rate.is_default,
    }


def _serialize_cost_center(cost_center: CostCenter | None) -> dict[str, Any] | None:
    if not cost_center:
        return None

    return {
        "id": cost_center.id,
        "code": cost_center.code,
        "name": cost_center.name,
        "name_en": cost_center.name_en,
        "level": cost_center.level,
        "is_group": cost_center.is_group,
        "status": cost_center.status,
        "can_post": cost_center.can_post,
    }


# ============================================================
# 🧭 Routing Rule Serializers
# ============================================================

def _serialize_routing_rule(rule: AccountingRoutingRule) -> dict[str, Any]:
    return {
        "id": rule.id,
        "source": rule.source,
        "source_label": _status_label(rule.source, AccountingRoutingSource.choices),
        "purpose": rule.purpose,
        "purpose_label": _status_label(rule.purpose, AccountingAccountPurpose.choices),
        "account_id": rule.account_id,
        "account": _serialize_account(rule.account),
        "tax_rate_id": rule.tax_rate_id,
        "tax_rate": _serialize_tax_rate(rule.tax_rate),
        "cost_center_id": rule.cost_center_id,
        "cost_center": _serialize_cost_center(rule.cost_center),
        "is_active": rule.is_active,
        "priority": rule.priority,
        "description": rule.description,
        "metadata": rule.metadata or {},
        "created_at": _iso_datetime(rule.created_at),
        "updated_at": _iso_datetime(rule.updated_at),
    }


def _build_routing_queryset(request):
    search = (request.GET.get("search") or "").strip()
    source = (request.GET.get("source") or "").strip()
    purpose = (request.GET.get("purpose") or "").strip()
    is_active = _parse_bool(request.GET.get("is_active"), default=None)

    account_id = _parse_int(request.GET.get("account_id"), "account_id")
    tax_rate_id = _parse_int(request.GET.get("tax_rate_id"), "tax_rate_id")
    cost_center_id = _parse_int(request.GET.get("cost_center_id"), "cost_center_id")

    qs = AccountingRoutingRule.objects.select_related(
        "account",
        "tax_rate",
        "cost_center",
    ).all()

    if search:
        qs = qs.filter(
            Q(source__icontains=search)
            | Q(purpose__icontains=search)
            | Q(description__icontains=search)
            | Q(account__code__icontains=search)
            | Q(account__name__icontains=search)
            | Q(account__name_en__icontains=search)
            | Q(tax_rate__code__icontains=search)
            | Q(tax_rate__name__icontains=search)
            | Q(cost_center__code__icontains=search)
            | Q(cost_center__name__icontains=search)
        )

    if source:
        valid_sources = {choice[0] for choice in AccountingRoutingSource.choices}
        if source not in valid_sources:
            raise ValueError("مصدر قاعدة التوجيه غير صحيح.")
        qs = qs.filter(source=source)

    if purpose:
        valid_purposes = {choice[0] for choice in AccountingAccountPurpose.choices}
        if purpose not in valid_purposes:
            raise ValueError("غرض قاعدة التوجيه غير صحيح.")
        qs = qs.filter(purpose=purpose)

    if is_active is not None:
        qs = qs.filter(is_active=is_active)

    if account_id:
        qs = qs.filter(account_id=account_id)

    if tax_rate_id:
        qs = qs.filter(tax_rate_id=tax_rate_id)

    if cost_center_id:
        qs = qs.filter(cost_center_id=cost_center_id)

    return {
        "qs": qs,
        "filters": {
            "search": search or None,
            "source": source or None,
            "purpose": purpose or None,
            "is_active": is_active,
            "account_id": account_id,
            "tax_rate_id": tax_rate_id,
            "cost_center_id": cost_center_id,
        },
    }


def _build_routing_summary(rules: list[AccountingRoutingRule]) -> dict[str, Any]:
    total = len(rules)
    active = sum(1 for item in rules if item.is_active)
    inactive = total - active

    by_source: dict[str, int] = {}
    by_purpose: dict[str, int] = {}

    for item in rules:
        by_source[item.source] = by_source.get(item.source, 0) + 1
        by_purpose[item.purpose] = by_purpose.get(item.purpose, 0) + 1

    return {
        "total_rules": total,
        "active_rules": active,
        "inactive_rules": inactive,
        "by_source": by_source,
        "by_purpose": by_purpose,
    }


def _apply_payload_to_routing_rule(
    rule: AccountingRoutingRule,
    data: dict[str, Any],
) -> AccountingRoutingRule:
    if "source" in data:
        source = _clean_text(data.get("source"))
        valid_sources = {choice[0] for choice in AccountingRoutingSource.choices}
        if source not in valid_sources:
            raise ValidationError({"source": "مصدر قاعدة التوجيه غير صحيح."})
        rule.source = source

    if "purpose" in data:
        purpose = _clean_text(data.get("purpose"))
        valid_purposes = {choice[0] for choice in AccountingAccountPurpose.choices}
        if purpose not in valid_purposes:
            raise ValidationError({"purpose": "غرض قاعدة التوجيه غير صحيح."})
        rule.purpose = purpose

    if "account_id" in data:
        account = _resolve_account(data.get("account_id"), "account_id")
        if not account:
            raise ValidationError({"account_id": "الحساب مطلوب لقاعدة التوجيه."})
        rule.account = account

    if "tax_rate_id" in data:
        rule.tax_rate = _resolve_tax_rate(data.get("tax_rate_id"))

    if "cost_center_id" in data:
        rule.cost_center = _resolve_cost_center(data.get("cost_center_id"))

    if "is_active" in data:
        is_active = data.get("is_active")
        if not isinstance(is_active, bool):
            raise ValidationError({"is_active": "قيمة is_active يجب أن تكون true أو false."})
        rule.is_active = is_active

    if "priority" in data:
        try:
            priority = int(data.get("priority") or 100)
        except (TypeError, ValueError) as exc:
            raise ValidationError({"priority": "الأولوية يجب أن تكون رقمًا صحيحًا."}) from exc

        if priority < 0:
            raise ValidationError({"priority": "الأولوية لا يمكن أن تكون سالبة."})

        rule.priority = priority

    if "description" in data:
        rule.description = _clean_text(data.get("description"))

    if "metadata" in data:
        metadata = data.get("metadata")
        if metadata in {None, ""}:
            rule.metadata = {}
        elif not isinstance(metadata, dict):
            raise ValidationError({"metadata": "metadata يجب أن تكون JSON Object."})
        else:
            rule.metadata = metadata

    return rule


def _get_routing_rule_or_none(rule_id: int) -> AccountingRoutingRule | None:
    return (
        AccountingRoutingRule.objects.select_related(
            "account",
            "tax_rate",
            "cost_center",
        )
        .filter(id=rule_id)
        .first()
    )


# ============================================================
# ⚙️ Accounting Settings Serializers
# ============================================================

def _get_or_create_settings() -> AccountingSettings:
    settings_obj = AccountingSettings.objects.select_related("default_tax_rate").first()

    if settings_obj:
        return settings_obj

    return AccountingSettings.objects.create()


def _serialize_settings(settings_obj: AccountingSettings) -> dict[str, Any]:
    return {
        "id": settings_obj.id,
        "default_currency": settings_obj.default_currency,
        "default_tax_rate_id": settings_obj.default_tax_rate_id,
        "default_tax_rate": _serialize_tax_rate(settings_obj.default_tax_rate),
        "auto_post_invoices": settings_obj.auto_post_invoices,
        "auto_post_payments": settings_obj.auto_post_payments,
        "auto_post_treasury": settings_obj.auto_post_treasury,
        "require_period_for_posting": settings_obj.require_period_for_posting,
        "allow_posting_without_cost_center": settings_obj.allow_posting_without_cost_center,
        "metadata": settings_obj.metadata or {},
        "created_at": _iso_datetime(settings_obj.created_at),
        "updated_at": _iso_datetime(settings_obj.updated_at),
    }


def _apply_payload_to_settings(
    settings_obj: AccountingSettings,
    data: dict[str, Any],
) -> AccountingSettings:
    if "default_currency" in data:
        currency = _clean_text(data.get("default_currency")).upper()
        if not currency:
            raise ValidationError({"default_currency": "العملة الافتراضية مطلوبة."})
        settings_obj.default_currency = currency

    if "default_tax_rate_id" in data:
        settings_obj.default_tax_rate = _resolve_tax_rate(data.get("default_tax_rate_id"))

    for field_name in [
        "auto_post_invoices",
        "auto_post_payments",
        "auto_post_treasury",
        "require_period_for_posting",
        "allow_posting_without_cost_center",
    ]:
        if field_name in data:
            value = data.get(field_name)
            if not isinstance(value, bool):
                raise ValidationError({field_name: f"قيمة {field_name} يجب أن تكون true أو false."})
            setattr(settings_obj, field_name, value)

    if "metadata" in data:
        metadata = data.get("metadata")
        if metadata in {None, ""}:
            settings_obj.metadata = {}
        elif not isinstance(metadata, dict):
            raise ValidationError({"metadata": "metadata يجب أن تكون JSON Object."})
        else:
            settings_obj.metadata = metadata

    return settings_obj


# ============================================================
# 📤 Excel
# ============================================================

def _build_routing_rules_excel(
    rules: list[AccountingRoutingRule],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Routing Rules")

    ws.append(["Accounting Routing Rules"])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)

    meta_rows = [
        ("Search", filters.get("search")),
        ("Source", filters.get("source")),
        ("Purpose", filters.get("purpose")),
        ("Is Active", filters.get("is_active")),
        ("Account ID", filters.get("account_id")),
        ("Tax Rate ID", filters.get("tax_rate_id")),
        ("Cost Center ID", filters.get("cost_center_id")),
        ("Total Rules", summary.get("total_rules")),
        ("Active Rules", summary.get("active_rules")),
        ("Inactive Rules", summary.get("inactive_rules")),
    ]

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    start_row = current_row + 1

    headers = [
        "ID",
        "Source",
        "Source Label",
        "Purpose",
        "Purpose Label",
        "Account Code",
        "Account Name",
        "Tax Rate",
        "Cost Center",
        "Is Active",
        "Priority",
        "Description",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for rule in rules:
        row = _serialize_routing_rule(rule)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("source"))
        ws.cell(row=current_row, column=3, value=row.get("source_label"))
        ws.cell(row=current_row, column=4, value=row.get("purpose"))
        ws.cell(row=current_row, column=5, value=row.get("purpose_label"))
        ws.cell(row=current_row, column=6, value=(row.get("account") or {}).get("code") if row.get("account") else "")
        ws.cell(row=current_row, column=7, value=(row.get("account") or {}).get("name") if row.get("account") else "")
        ws.cell(row=current_row, column=8, value=(row.get("tax_rate") or {}).get("code") if row.get("tax_rate") else "")
        ws.cell(row=current_row, column=9, value=(row.get("cost_center") or {}).get("name") if row.get("cost_center") else "")
        ws.cell(row=current_row, column=10, value=str(row.get("is_active")))
        ws.cell(row=current_row, column=11, value=row.get("priority"))
        ws.cell(row=current_row, column=12, value=row.get("description"))
        ws.cell(row=current_row, column=13, value=row.get("created_at"))
        ws.cell(row=current_row, column=14, value=row.get("updated_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "accounting_routing_rules.xlsx")


# ============================================================
# 🧭 Routing Rules API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "POST"])
def accounting_routing_rules_api(request):
    """
    GET:
      قائمة قواعد التوجيه المحاسبي.

    POST:
      إنشاء قاعدة توجيه.
    """
    if request.method == "POST":
        return _create_routing_rule(request)

    return _list_routing_rules(request)


def _list_routing_rules(request):
    try:
        query_data = _build_routing_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_ordering(request.GET.get("ordering"), default="source")

        if ordering.lstrip("-") == "source":
            qs = qs.order_by(ordering, "purpose", "priority", "id")
        else:
            qs = qs.order_by(ordering, "source", "purpose", "priority", "id")

        rules = list(qs)
        summary = _build_routing_summary(rules)
        paginator = Paginator(rules, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        return _success_response(
            {
                "filters": {
                    **query_data["filters"],
                    "ordering": ordering,
                },
                "summary": summary,
                "choices": _choices_payload(),
                "pagination": {
                    "page": page_obj.number,
                    "page_size": page_size,
                    "total_pages": paginator.num_pages,
                    "total_items": paginator.count,
                    "has_next": page_obj.has_next(),
                    "has_previous": page_obj.has_previous(),
                    "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                    "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
                },
                "results": [
                    _serialize_routing_rule(item)
                    for item in page_obj.object_list
                ],
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load accounting routing rules: %s", exc)
        return _error_response("تعذر تحميل قواعد التوجيه المحاسبي.", status=500)


@transaction.atomic
def _create_routing_rule(request):
    try:
        data = _json_body(request)

        rule = AccountingRoutingRule()
        _apply_payload_to_routing_rule(rule, data)
        rule.full_clean()
        rule.save()

        return _success_response(
            {
                "routing_rule": _serialize_routing_rule(rule),
                "message": "تم إنشاء قاعدة التوجيه المحاسبي بنجاح.",
            },
            status=201,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to create accounting routing rule: %s", exc)
        return _error_response("تعذر إنشاء قاعدة التوجيه المحاسبي.", status=500)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT"])
def accounting_routing_rule_detail_api(request, rule_id: int):
    """
    GET:
      تفاصيل قاعدة توجيه.

    PATCH/PUT:
      تحديث قاعدة توجيه.
    """
    if request.method in {"PATCH", "PUT"}:
        return _update_routing_rule(request, rule_id)

    return _detail_routing_rule(request, rule_id)


def _detail_routing_rule(request, rule_id: int):
    try:
        rule = _get_routing_rule_or_none(rule_id)

        if not rule:
            return _error_response("قاعدة التوجيه المطلوبة غير موجودة.", status=404)

        return _success_response(
            {
                "routing_rule": _serialize_routing_rule(rule),
            }
        )

    except Exception as exc:
        logger.exception("Failed to load accounting routing rule detail: %s", exc)
        return _error_response("تعذر تحميل تفاصيل قاعدة التوجيه المحاسبي.", status=500)


@transaction.atomic
def _update_routing_rule(request, rule_id: int):
    try:
        data = _json_body(request)
        rule = _get_routing_rule_or_none(rule_id)

        if not rule:
            return _error_response("قاعدة التوجيه المطلوبة غير موجودة.", status=404)

        _apply_payload_to_routing_rule(rule, data)
        rule.full_clean()
        rule.save()

        return _success_response(
            {
                "routing_rule": _serialize_routing_rule(rule),
                "message": "تم تحديث قاعدة التوجيه المحاسبي بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update accounting routing rule: %s", exc)
        return _error_response("تعذر تحديث قاعدة التوجيه المحاسبي.", status=500)


@csrf_exempt
@require_http_methods(["POST"])
def accounting_routing_rule_status_api(request, rule_id: int):
    """
    تغيير حالة قاعدة التوجيه.
    body:
    {
      "is_active": true | false
    }
    """
    try:
        data = _json_body(request)

        if "is_active" not in data or not isinstance(data.get("is_active"), bool):
            raise ValueError("قيمة is_active مطلوبة ويجب أن تكون true أو false.")

        rule = _get_routing_rule_or_none(rule_id)

        if not rule:
            return _error_response("قاعدة التوجيه المطلوبة غير موجودة.", status=404)

        rule.is_active = data["is_active"]
        rule.full_clean()
        rule.save(update_fields=["is_active", "updated_at"])

        return _success_response(
            {
                "routing_rule": _serialize_routing_rule(rule),
                "message": "تم تحديث حالة قاعدة التوجيه المحاسبي بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update accounting routing rule status: %s", exc)
        return _error_response("تعذر تحديث حالة قاعدة التوجيه المحاسبي.", status=500)


@require_GET
def accounting_routing_rules_excel_api(request):
    """
    تصدير قواعد التوجيه المحاسبي إلى Excel.
    """
    try:
        query_data = _build_routing_queryset(request)
        ordering = _parse_ordering(request.GET.get("ordering"), default="source")

        if ordering.lstrip("-") == "source":
            qs = query_data["qs"].order_by(ordering, "purpose", "priority", "id")
        else:
            qs = query_data["qs"].order_by(ordering, "source", "purpose", "priority", "id")

        rules = list(qs)
        summary = _build_routing_summary(rules)

        return _build_routing_rules_excel(
            rules,
            query_data["filters"],
            summary,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export accounting routing rules excel: %s", exc)
        return _error_response("تعذر تصدير قواعد التوجيه المحاسبي إلى Excel.", status=500)


# ============================================================
# ⚙️ Accounting Settings API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT"])
def accounting_settings_api(request):
    """
    GET:
      إعدادات المحاسبة العامة.

    PATCH/PUT:
      تحديث إعدادات المحاسبة.
    """
    if request.method in {"PATCH", "PUT"}:
        return _update_accounting_settings(request)

    return _detail_accounting_settings(request)


def _detail_accounting_settings(request):
    try:
        settings_obj = _get_or_create_settings()

        return _success_response(
            {
                "settings": _serialize_settings(settings_obj),
                "choices": _choices_payload(),
            }
        )

    except Exception as exc:
        logger.exception("Failed to load accounting settings: %s", exc)
        return _error_response("تعذر تحميل إعدادات المحاسبة.", status=500)


@transaction.atomic
def _update_accounting_settings(request):
    try:
        data = _json_body(request)
        settings_obj = _get_or_create_settings()

        _apply_payload_to_settings(settings_obj, data)
        settings_obj.full_clean()
        settings_obj.save()

        return _success_response(
            {
                "settings": _serialize_settings(settings_obj),
                "message": "تم تحديث إعدادات المحاسبة بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update accounting settings: %s", exc)
        return _error_response("تعذر تحديث إعدادات المحاسبة.", status=500)


# ============================================================
# 🧩 Choices
# ============================================================

def _choices_payload() -> dict[str, Any]:
    return {
        "routing_sources": [
            {"value": value, "label": label}
            for value, label in AccountingRoutingSource.choices
        ],
        "routing_purposes": [
            {"value": value, "label": label}
            for value, label in AccountingAccountPurpose.choices
        ],
        "orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_ROUTING_ORDERING.keys()
        ],
    }
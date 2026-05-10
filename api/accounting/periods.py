# ============================================================
# 📂 api/accounting/periods.py
# 🧠 Accounting Periods API — Primey Care V2
# ------------------------------------------------------------
# ✅ إدارة السنوات المالية Fiscal Years
# ✅ إدارة الفترات المحاسبية Accounting Periods
# ✅ قائمة / تفاصيل / إنشاء / تحديث
# ✅ فتح / إغلاق / قفل الفترة
# ✅ تعيين السنة المالية الحالية
# ✅ Excel Export
# ✅ متوافق مع Accounting Backend الجديد
# ============================================================

from __future__ import annotations

import json
import logging
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import (
    AccountingPeriod,
    AccountingPeriodStatus,
    FiscalYear,
    FiscalYearStatus,
    JournalEntry,
)


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Constants
# ============================================================

DEFAULT_PAGE = 1
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 500

ALLOWED_FISCAL_YEAR_ORDERING = {
    "start_date": "start_date",
    "-start_date": "-start_date",
    "end_date": "end_date",
    "-end_date": "-end_date",
    "name": "name",
    "-name": "-name",
    "status": "status",
    "-status": "-status",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "updated_at": "updated_at",
    "-updated_at": "-updated_at",
}

ALLOWED_PERIOD_ORDERING = {
    "start_date": "start_date",
    "-start_date": "-start_date",
    "end_date": "end_date",
    "-end_date": "-end_date",
    "name": "name",
    "-name": "-name",
    "status": "status",
    "-status": "-status",
    "fiscal_year": "fiscal_year__start_date",
    "-fiscal_year": "-fiscal_year__start_date",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "updated_at": "updated_at",
    "-updated_at": "-updated_at",
}


# ============================================================
# 🔧 Helpers
# ============================================================

def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool | None = None) -> bool | None:
    if value in {None, ""}:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    raise ValueError("قيمة boolean غير صحيحة. استخدم true أو false.")


def _parse_int(value: str | None, field_name: str) -> int | None:
    if value in {None, ""}:
        return None

    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed <= 0:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر من صفر.")

    return parsed


def _parse_positive_int(
    value: str | None,
    field_name: str,
    *,
    default: int,
    min_value: int = 1,
    max_value: int | None = None,
) -> int:
    if value in {None, ""}:
        parsed = default
    else:
        try:
            parsed = int(str(value).strip())
        except ValueError as exc:
            raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed < min_value:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر أو تساوي {min_value}.")

    if max_value is not None and parsed > max_value:
        raise ValueError(f"قيمة {field_name} يجب ألا تتجاوز {max_value}.")

    return parsed


def _parse_date(value: Any, field_name: str) -> date | None:
    if value in {None, ""}:
        return None

    if isinstance(value, date):
        return value

    raw = str(value).strip()

    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} غير صحيحة. استخدم الصيغة YYYY-MM-DD.") from exc


def _validate_date_range(start_date: date | None, end_date: date | None) -> None:
    if start_date and end_date and start_date > end_date:
        raise ValueError("تاريخ البداية لا يمكن أن يكون بعد تاريخ النهاية.")


def _json_body(request) -> dict[str, Any]:
    try:
        body = request.body.decode("utf-8") if request.body else "{}"
        data = json.loads(body or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError("صيغة JSON غير صحيحة.") from exc

    if not isinstance(data, dict):
        raise ValueError("جسم الطلب يجب أن يكون JSON Object.")

    return data


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any], status: int = 200) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _validation_error_payload(exc: ValidationError) -> dict[str, Any]:
    if hasattr(exc, "message_dict"):
        return {
            "ok": False,
            "message": "تعذر حفظ البيانات.",
            "errors": exc.message_dict,
        }

    return {
        "ok": False,
        "message": "تعذر حفظ البيانات.",
        "errors": exc.messages if hasattr(exc, "messages") else [str(exc)],
    }


def _status_label(status: str, choices) -> str:
    for value, label in choices:
        if value == status:
            return label
    return status


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _iso_date(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return None


def _iso_datetime(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return None


# ============================================================
# 🧾 Serializers
# ============================================================

def _journal_entries_count_for_fiscal_year(fiscal_year: FiscalYear) -> int:
    return JournalEntry.objects.filter(period__fiscal_year=fiscal_year).count()


def _periods_count_for_fiscal_year(fiscal_year: FiscalYear) -> int:
    return fiscal_year.periods.count()


def _journal_entries_count_for_period(period: AccountingPeriod) -> int:
    return period.journal_entries.count()


def _serialize_fiscal_year(
    fiscal_year: FiscalYear,
    *,
    include_counts: bool = True,
) -> dict[str, Any]:
    periods_count = _periods_count_for_fiscal_year(fiscal_year) if include_counts else 0
    journal_entries_count = _journal_entries_count_for_fiscal_year(fiscal_year) if include_counts else 0

    return {
        "id": fiscal_year.id,
        "name": fiscal_year.name,
        "start_date": _iso_date(fiscal_year.start_date),
        "end_date": _iso_date(fiscal_year.end_date),
        "status": fiscal_year.status,
        "status_label": _status_label(fiscal_year.status, FiscalYearStatus.choices),
        "is_current": fiscal_year.is_current,
        "description": fiscal_year.description,
        "closed_at": _iso_datetime(fiscal_year.closed_at),
        "periods_count": periods_count,
        "journal_entries_count": journal_entries_count,
        "created_at": _iso_datetime(fiscal_year.created_at),
        "updated_at": _iso_datetime(fiscal_year.updated_at),
    }


def _serialize_period(
    period: AccountingPeriod,
    *,
    include_counts: bool = True,
) -> dict[str, Any]:
    fiscal_year = period.fiscal_year
    journal_entries_count = _journal_entries_count_for_period(period) if include_counts else 0

    return {
        "id": period.id,
        "name": period.name,
        "fiscal_year_id": period.fiscal_year_id,
        "fiscal_year": {
            "id": fiscal_year.id,
            "name": fiscal_year.name,
            "start_date": _iso_date(fiscal_year.start_date),
            "end_date": _iso_date(fiscal_year.end_date),
            "status": fiscal_year.status,
            "status_label": _status_label(fiscal_year.status, FiscalYearStatus.choices),
            "is_current": fiscal_year.is_current,
        } if fiscal_year else None,
        "start_date": _iso_date(period.start_date),
        "end_date": _iso_date(period.end_date),
        "status": period.status,
        "status_label": _status_label(period.status, AccountingPeriodStatus.choices),
        "is_adjustment_period": period.is_adjustment_period,
        "closed_at": _iso_datetime(period.closed_at),
        "journal_entries_count": journal_entries_count,
        "can_post": period.status == AccountingPeriodStatus.OPEN,
        "created_at": _iso_datetime(period.created_at),
        "updated_at": _iso_datetime(period.updated_at),
    }


def _choices_payload() -> dict[str, Any]:
    return {
        "fiscal_year_statuses": [
            {"value": value, "label": label}
            for value, label in FiscalYearStatus.choices
        ],
        "period_statuses": [
            {"value": value, "label": label}
            for value, label in AccountingPeriodStatus.choices
        ],
        "fiscal_year_orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_FISCAL_YEAR_ORDERING.keys()
        ],
        "period_orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_PERIOD_ORDERING.keys()
        ],
    }


# ============================================================
# 📅 Fiscal Years Query
# ============================================================

def _parse_fiscal_year_ordering(value: str | None, default: str = "-start_date") -> str:
    raw = (value or "").strip() or default

    if raw not in ALLOWED_FISCAL_YEAR_ORDERING:
        allowed = ", ".join(ALLOWED_FISCAL_YEAR_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_FISCAL_YEAR_ORDERING[raw]


def _build_fiscal_years_queryset(request):
    search = (request.GET.get("search") or "").strip()
    status = (request.GET.get("status") or "").strip()
    is_current = _parse_bool(request.GET.get("is_current"), default=None)

    start_from = _parse_date(request.GET.get("start_from"), "start_from")
    start_to = _parse_date(request.GET.get("start_to"), "start_to")
    _validate_date_range(start_from, start_to)

    qs = FiscalYear.objects.all()

    if search:
        qs = qs.filter(
            Q(name__icontains=search)
            | Q(description__icontains=search)
        )

    if status:
        valid_statuses = {choice[0] for choice in FiscalYearStatus.choices}
        if status not in valid_statuses:
            raise ValueError("حالة السنة المالية غير صحيحة.")
        qs = qs.filter(status=status)

    if is_current is not None:
        qs = qs.filter(is_current=is_current)

    if start_from:
        qs = qs.filter(start_date__gte=start_from)

    if start_to:
        qs = qs.filter(start_date__lte=start_to)

    return {
        "qs": qs,
        "filters": {
            "search": search or None,
            "status": status or None,
            "is_current": is_current,
            "start_from": _iso_date(start_from),
            "start_to": _iso_date(start_to),
        },
    }


def _build_fiscal_years_summary(fiscal_years: list[FiscalYear]) -> dict[str, Any]:
    total = len(fiscal_years)
    open_count = sum(1 for item in fiscal_years if item.status == FiscalYearStatus.OPEN)
    closed_count = sum(1 for item in fiscal_years if item.status == FiscalYearStatus.CLOSED)
    archived_count = sum(1 for item in fiscal_years if item.status == FiscalYearStatus.ARCHIVED)
    current_count = sum(1 for item in fiscal_years if item.is_current)

    return {
        "total_fiscal_years": total,
        "open_fiscal_years": open_count,
        "closed_fiscal_years": closed_count,
        "archived_fiscal_years": archived_count,
        "current_fiscal_years": current_count,
    }


def _apply_payload_to_fiscal_year(fiscal_year: FiscalYear, data: dict[str, Any]) -> FiscalYear:
    if "name" in data:
        fiscal_year.name = _clean_text(data.get("name"))

    if "start_date" in data:
        fiscal_year.start_date = _parse_date(data.get("start_date"), "start_date")

    if "end_date" in data:
        fiscal_year.end_date = _parse_date(data.get("end_date"), "end_date")

    if "status" in data:
        status = _clean_text(data.get("status"))
        valid_statuses = {choice[0] for choice in FiscalYearStatus.choices}
        if status not in valid_statuses:
            raise ValidationError({"status": "حالة السنة المالية غير صحيحة."})
        fiscal_year.status = status

    if "is_current" in data:
        is_current = data.get("is_current")
        if not isinstance(is_current, bool):
            raise ValidationError({"is_current": "قيمة is_current يجب أن تكون true أو false."})
        fiscal_year.is_current = is_current

    if "description" in data:
        fiscal_year.description = _clean_text(data.get("description"))

    return fiscal_year


# ============================================================
# 📆 Periods Query
# ============================================================

def _parse_period_ordering(value: str | None, default: str = "-start_date") -> str:
    raw = (value or "").strip() or default

    if raw not in ALLOWED_PERIOD_ORDERING:
        allowed = ", ".join(ALLOWED_PERIOD_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_PERIOD_ORDERING[raw]


def _build_periods_queryset(request):
    search = (request.GET.get("search") or "").strip()
    status = (request.GET.get("status") or "").strip()
    fiscal_year_id = _parse_int(request.GET.get("fiscal_year_id"), "fiscal_year_id")
    is_adjustment_period = _parse_bool(request.GET.get("is_adjustment_period"), default=None)
    can_post = _parse_bool(request.GET.get("can_post"), default=None)

    date_from = _parse_date(request.GET.get("date_from"), "date_from")
    date_to = _parse_date(request.GET.get("date_to"), "date_to")
    _validate_date_range(date_from, date_to)

    qs = AccountingPeriod.objects.select_related("fiscal_year").all()

    if search:
        qs = qs.filter(
            Q(name__icontains=search)
            | Q(fiscal_year__name__icontains=search)
        )

    if status:
        valid_statuses = {choice[0] for choice in AccountingPeriodStatus.choices}
        if status not in valid_statuses:
            raise ValueError("حالة الفترة المحاسبية غير صحيحة.")
        qs = qs.filter(status=status)

    if fiscal_year_id:
        qs = qs.filter(fiscal_year_id=fiscal_year_id)

    if is_adjustment_period is not None:
        qs = qs.filter(is_adjustment_period=is_adjustment_period)

    if can_post is True:
        qs = qs.filter(status=AccountingPeriodStatus.OPEN)

    if can_post is False:
        qs = qs.exclude(status=AccountingPeriodStatus.OPEN)

    if date_from:
        qs = qs.filter(start_date__gte=date_from)

    if date_to:
        qs = qs.filter(end_date__lte=date_to)

    return {
        "qs": qs,
        "filters": {
            "search": search or None,
            "status": status or None,
            "fiscal_year_id": fiscal_year_id,
            "is_adjustment_period": is_adjustment_period,
            "can_post": can_post,
            "date_from": _iso_date(date_from),
            "date_to": _iso_date(date_to),
        },
    }


def _build_periods_summary(periods: list[AccountingPeriod]) -> dict[str, Any]:
    total = len(periods)
    open_count = sum(1 for item in periods if item.status == AccountingPeriodStatus.OPEN)
    closed_count = sum(1 for item in periods if item.status == AccountingPeriodStatus.CLOSED)
    locked_count = sum(1 for item in periods if item.status == AccountingPeriodStatus.LOCKED)
    adjustment_count = sum(1 for item in periods if item.is_adjustment_period)

    return {
        "total_periods": total,
        "open_periods": open_count,
        "closed_periods": closed_count,
        "locked_periods": locked_count,
        "adjustment_periods": adjustment_count,
        "posting_periods": open_count,
    }


def _resolve_fiscal_year(fiscal_year_id: Any) -> FiscalYear:
    try:
        parsed_id = int(fiscal_year_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({"fiscal_year_id": "معرف السنة المالية غير صحيح."}) from exc

    fiscal_year = FiscalYear.objects.filter(id=parsed_id).first()

    if not fiscal_year:
        raise ValidationError({"fiscal_year_id": "السنة المالية المطلوبة غير موجودة."})

    return fiscal_year


def _apply_payload_to_period(period: AccountingPeriod, data: dict[str, Any]) -> AccountingPeriod:
    if "fiscal_year_id" in data:
        period.fiscal_year = _resolve_fiscal_year(data.get("fiscal_year_id"))

    if "name" in data:
        period.name = _clean_text(data.get("name"))

    if "start_date" in data:
        period.start_date = _parse_date(data.get("start_date"), "start_date")

    if "end_date" in data:
        period.end_date = _parse_date(data.get("end_date"), "end_date")

    if "status" in data:
        status = _clean_text(data.get("status"))
        valid_statuses = {choice[0] for choice in AccountingPeriodStatus.choices}
        if status not in valid_statuses:
            raise ValidationError({"status": "حالة الفترة المحاسبية غير صحيحة."})
        period.status = status

    if "is_adjustment_period" in data:
        is_adjustment_period = data.get("is_adjustment_period")
        if not isinstance(is_adjustment_period, bool):
            raise ValidationError({"is_adjustment_period": "قيمة is_adjustment_period يجب أن تكون true أو false."})
        period.is_adjustment_period = is_adjustment_period

    return period


# ============================================================
# 📤 Excel Builders
# ============================================================

def _build_fiscal_years_excel(
    fiscal_years: list[FiscalYear],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Fiscal Years")

    ws.append(["Fiscal Years"])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    meta_rows = [
        ("Search", filters.get("search")),
        ("Status", filters.get("status")),
        ("Is Current", filters.get("is_current")),
        ("Start From", filters.get("start_from")),
        ("Start To", filters.get("start_to")),
        ("Total Fiscal Years", summary.get("total_fiscal_years")),
        ("Open Fiscal Years", summary.get("open_fiscal_years")),
        ("Closed Fiscal Years", summary.get("closed_fiscal_years")),
        ("Archived Fiscal Years", summary.get("archived_fiscal_years")),
        ("Current Fiscal Years", summary.get("current_fiscal_years")),
    ]

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    start_row = current_row + 1

    headers = [
        "ID",
        "Name",
        "Start Date",
        "End Date",
        "Status",
        "Status Label",
        "Is Current",
        "Description",
        "Closed At",
        "Periods Count",
        "Journal Entries Count",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for fiscal_year in fiscal_years:
        row = _serialize_fiscal_year(fiscal_year)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("name"))
        ws.cell(row=current_row, column=3, value=row.get("start_date"))
        ws.cell(row=current_row, column=4, value=row.get("end_date"))
        ws.cell(row=current_row, column=5, value=row.get("status"))
        ws.cell(row=current_row, column=6, value=row.get("status_label"))
        ws.cell(row=current_row, column=7, value=str(row.get("is_current")))
        ws.cell(row=current_row, column=8, value=row.get("description"))
        ws.cell(row=current_row, column=9, value=row.get("closed_at"))
        ws.cell(row=current_row, column=10, value=row.get("periods_count"))
        ws.cell(row=current_row, column=11, value=row.get("journal_entries_count"))
        ws.cell(row=current_row, column=12, value=row.get("created_at"))
        ws.cell(row=current_row, column=13, value=row.get("updated_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "fiscal_years.xlsx")


def _build_periods_excel(
    periods: list[AccountingPeriod],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Accounting Periods")

    ws.append(["Accounting Periods"])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)

    meta_rows = [
        ("Search", filters.get("search")),
        ("Status", filters.get("status")),
        ("Fiscal Year ID", filters.get("fiscal_year_id")),
        ("Is Adjustment Period", filters.get("is_adjustment_period")),
        ("Can Post", filters.get("can_post")),
        ("Date From", filters.get("date_from")),
        ("Date To", filters.get("date_to")),
        ("Total Periods", summary.get("total_periods")),
        ("Open Periods", summary.get("open_periods")),
        ("Closed Periods", summary.get("closed_periods")),
        ("Locked Periods", summary.get("locked_periods")),
        ("Adjustment Periods", summary.get("adjustment_periods")),
    ]

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    start_row = current_row + 1

    headers = [
        "ID",
        "Name",
        "Fiscal Year ID",
        "Fiscal Year",
        "Start Date",
        "End Date",
        "Status",
        "Status Label",
        "Is Adjustment Period",
        "Can Post",
        "Closed At",
        "Journal Entries Count",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for period in periods:
        row = _serialize_period(period)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("name"))
        ws.cell(row=current_row, column=3, value=row.get("fiscal_year_id"))
        ws.cell(
            row=current_row,
            column=4,
            value=(row.get("fiscal_year") or {}).get("name") if row.get("fiscal_year") else "",
        )
        ws.cell(row=current_row, column=5, value=row.get("start_date"))
        ws.cell(row=current_row, column=6, value=row.get("end_date"))
        ws.cell(row=current_row, column=7, value=row.get("status"))
        ws.cell(row=current_row, column=8, value=row.get("status_label"))
        ws.cell(row=current_row, column=9, value=str(row.get("is_adjustment_period")))
        ws.cell(row=current_row, column=10, value=str(row.get("can_post")))
        ws.cell(row=current_row, column=11, value=row.get("closed_at"))
        ws.cell(row=current_row, column=12, value=row.get("journal_entries_count"))
        ws.cell(row=current_row, column=13, value=row.get("created_at"))
        ws.cell(row=current_row, column=14, value=row.get("updated_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "accounting_periods.xlsx")


# ============================================================
# 📅 Fiscal Years API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "POST"])
def accounting_fiscal_years_api(request):
    """
    GET:
      قائمة السنوات المالية.

    POST:
      إنشاء سنة مالية.
    """
    if request.method == "POST":
        return _create_fiscal_year(request)

    return _list_fiscal_years(request)


def _list_fiscal_years(request):
    try:
        query_data = _build_fiscal_years_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_fiscal_year_ordering(request.GET.get("ordering"), default="-start_date")

        if ordering.lstrip("-") == "start_date":
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by(ordering, "-start_date")

        fiscal_years = list(qs)
        summary = _build_fiscal_years_summary(fiscal_years)
        paginator = Paginator(fiscal_years, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        return _success_response(
            {
                "filters": {
                    **query_data["filters"],
                    "ordering": ordering,
                },
                "summary": summary,
                "choices": _choices_payload(),
                "pagination": {
                    "page": page_obj.number,
                    "page_size": page_size,
                    "total_pages": paginator.num_pages,
                    "total_items": paginator.count,
                    "has_next": page_obj.has_next(),
                    "has_previous": page_obj.has_previous(),
                    "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                    "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
                },
                "results": [
                    _serialize_fiscal_year(item)
                    for item in page_obj.object_list
                ],
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load fiscal years: %s", exc)
        return _error_response("تعذر تحميل السنوات المالية.", status=500)


@transaction.atomic
def _create_fiscal_year(request):
    try:
        data = _json_body(request)
        fiscal_year = FiscalYear()
        _apply_payload_to_fiscal_year(fiscal_year, data)
        fiscal_year.full_clean()
        fiscal_year.save()

        return _success_response(
            {
                "fiscal_year": _serialize_fiscal_year(fiscal_year),
                "message": "تم إنشاء السنة المالية بنجاح.",
            },
            status=201,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to create fiscal year: %s", exc)
        return _error_response("تعذر إنشاء السنة المالية.", status=500)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT"])
def accounting_fiscal_year_detail_api(request, fiscal_year_id: int):
    """
    GET:
      تفاصيل سنة مالية.

    PATCH/PUT:
      تحديث سنة مالية.
    """
    if request.method in {"PATCH", "PUT"}:
        return _update_fiscal_year(request, fiscal_year_id)

    return _detail_fiscal_year(request, fiscal_year_id)


def _get_fiscal_year_or_none(fiscal_year_id: int) -> FiscalYear | None:
    return FiscalYear.objects.filter(id=fiscal_year_id).first()


def _detail_fiscal_year(request, fiscal_year_id: int):
    try:
        fiscal_year = _get_fiscal_year_or_none(fiscal_year_id)

        if not fiscal_year:
            return _error_response("السنة المالية المطلوبة غير موجودة.", status=404)

        periods = list(fiscal_year.periods.order_by("start_date"))

        return _success_response(
            {
                "fiscal_year": _serialize_fiscal_year(fiscal_year),
                "periods": [
                    _serialize_period(period)
                    for period in periods
                ],
            }
        )

    except Exception as exc:
        logger.exception("Failed to load fiscal year detail: %s", exc)
        return _error_response("تعذر تحميل تفاصيل السنة المالية.", status=500)


@transaction.atomic
def _update_fiscal_year(request, fiscal_year_id: int):
    try:
        data = _json_body(request)
        fiscal_year = _get_fiscal_year_or_none(fiscal_year_id)

        if not fiscal_year:
            return _error_response("السنة المالية المطلوبة غير موجودة.", status=404)

        _apply_payload_to_fiscal_year(fiscal_year, data)
        fiscal_year.full_clean()
        fiscal_year.save()

        return _success_response(
            {
                "fiscal_year": _serialize_fiscal_year(fiscal_year),
                "message": "تم تحديث السنة المالية بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update fiscal year: %s", exc)
        return _error_response("تعذر تحديث السنة المالية.", status=500)


@csrf_exempt
@require_http_methods(["POST"])
def accounting_fiscal_year_status_api(request, fiscal_year_id: int):
    """
    تغيير حالة السنة المالية.
    body:
    {
      "status": "OPEN" | "CLOSED" | "ARCHIVED"
    }
    """
    try:
        data = _json_body(request)
        status = _clean_text(data.get("status"))

        valid_statuses = {choice[0] for choice in FiscalYearStatus.choices}
        if status not in valid_statuses:
            raise ValueError("حالة السنة المالية غير صحيحة.")

        fiscal_year = _get_fiscal_year_or_none(fiscal_year_id)

        if not fiscal_year:
            return _error_response("السنة المالية المطلوبة غير موجودة.", status=404)

        fiscal_year.status = status
        fiscal_year.closed_at = timezone.now() if status == FiscalYearStatus.CLOSED else None
        fiscal_year.full_clean()
        fiscal_year.save(update_fields=["status", "closed_at", "updated_at"])

        return _success_response(
            {
                "fiscal_year": _serialize_fiscal_year(fiscal_year),
                "message": "تم تحديث حالة السنة المالية بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update fiscal year status: %s", exc)
        return _error_response("تعذر تحديث حالة السنة المالية.", status=500)


@csrf_exempt
@require_http_methods(["POST"])
def accounting_fiscal_year_set_current_api(request, fiscal_year_id: int):
    """
    تعيين سنة مالية كسنة حالية.
    """
    try:
        fiscal_year = _get_fiscal_year_or_none(fiscal_year_id)

        if not fiscal_year:
            return _error_response("السنة المالية المطلوبة غير موجودة.", status=404)

        with transaction.atomic():
            FiscalYear.objects.exclude(id=fiscal_year.id).update(is_current=False)
            fiscal_year.is_current = True
            fiscal_year.save(update_fields=["is_current", "updated_at"])

        return _success_response(
            {
                "fiscal_year": _serialize_fiscal_year(fiscal_year),
                "message": "تم تعيين السنة المالية الحالية بنجاح.",
            }
        )

    except Exception as exc:
        logger.exception("Failed to set fiscal year as current: %s", exc)
        return _error_response("تعذر تعيين السنة المالية الحالية.", status=500)


@require_GET
def accounting_fiscal_years_excel_api(request):
    """
    تصدير السنوات المالية إلى Excel.
    """
    try:
        query_data = _build_fiscal_years_queryset(request)
        ordering = _parse_fiscal_year_ordering(request.GET.get("ordering"), default="-start_date")

        if ordering.lstrip("-") == "start_date":
            qs = query_data["qs"].order_by(ordering)
        else:
            qs = query_data["qs"].order_by(ordering, "-start_date")

        fiscal_years = list(qs)
        summary = _build_fiscal_years_summary(fiscal_years)

        return _build_fiscal_years_excel(
            fiscal_years,
            query_data["filters"],
            summary,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export fiscal years excel: %s", exc)
        return _error_response("تعذر تصدير السنوات المالية إلى Excel.", status=500)


# ============================================================
# 📆 Periods API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "POST"])
def accounting_periods_api(request):
    """
    GET:
      قائمة الفترات المحاسبية.

    POST:
      إنشاء فترة محاسبية.
    """
    if request.method == "POST":
        return _create_period(request)

    return _list_periods(request)


def _list_periods(request):
    try:
        query_data = _build_periods_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_period_ordering(request.GET.get("ordering"), default="-start_date")

        if ordering.lstrip("-") == "start_date":
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by(ordering, "-start_date")

        periods = list(qs)
        summary = _build_periods_summary(periods)
        paginator = Paginator(periods, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        return _success_response(
            {
                "filters": {
                    **query_data["filters"],
                    "ordering": ordering,
                },
                "summary": summary,
                "choices": _choices_payload(),
                "pagination": {
                    "page": page_obj.number,
                    "page_size": page_size,
                    "total_pages": paginator.num_pages,
                    "total_items": paginator.count,
                    "has_next": page_obj.has_next(),
                    "has_previous": page_obj.has_previous(),
                    "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                    "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
                },
                "results": [
                    _serialize_period(item)
                    for item in page_obj.object_list
                ],
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load accounting periods: %s", exc)
        return _error_response("تعذر تحميل الفترات المحاسبية.", status=500)


@transaction.atomic
def _create_period(request):
    try:
        data = _json_body(request)

        period = AccountingPeriod()
        _apply_payload_to_period(period, data)
        period.full_clean()
        period.save()

        return _success_response(
            {
                "period": _serialize_period(period),
                "message": "تم إنشاء الفترة المحاسبية بنجاح.",
            },
            status=201,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to create accounting period: %s", exc)
        return _error_response("تعذر إنشاء الفترة المحاسبية.", status=500)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT"])
def accounting_period_detail_api(request, period_id: int):
    """
    GET:
      تفاصيل فترة محاسبية.

    PATCH/PUT:
      تحديث فترة محاسبية.
    """
    if request.method in {"PATCH", "PUT"}:
        return _update_period(request, period_id)

    return _detail_period(request, period_id)


def _get_period_or_none(period_id: int) -> AccountingPeriod | None:
    return (
        AccountingPeriod.objects.select_related("fiscal_year")
        .filter(id=period_id)
        .first()
    )


def _detail_period(request, period_id: int):
    try:
        period = _get_period_or_none(period_id)

        if not period:
            return _error_response("الفترة المحاسبية المطلوبة غير موجودة.", status=404)

        recent_entries = list(
            JournalEntry.objects.filter(period_id=period.id)
            .order_by("-entry_date", "-id")[:20]
        )

        return _success_response(
            {
                "period": _serialize_period(period),
                "recent_journal_entries": [
                    {
                        "id": entry.id,
                        "entry_number": entry.entry_number,
                        "entry_date": _iso_date(entry.entry_date),
                        "status": entry.status,
                        "posting_source": entry.posting_source,
                        "reference": entry.reference,
                        "source_type": entry.source_type,
                        "source_id": entry.source_id,
                        "source_number": entry.source_number,
                        "total_debit": entry.total_debit,
                        "total_credit": entry.total_credit,
                        "currency": entry.currency,
                    }
                    for entry in recent_entries
                ],
            }
        )

    except Exception as exc:
        logger.exception("Failed to load accounting period detail: %s", exc)
        return _error_response("تعذر تحميل تفاصيل الفترة المحاسبية.", status=500)


@transaction.atomic
def _update_period(request, period_id: int):
    try:
        data = _json_body(request)
        period = _get_period_or_none(period_id)

        if not period:
            return _error_response("الفترة المحاسبية المطلوبة غير موجودة.", status=404)

        _apply_payload_to_period(period, data)
        period.full_clean()
        period.save()

        return _success_response(
            {
                "period": _serialize_period(period),
                "message": "تم تحديث الفترة المحاسبية بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update accounting period: %s", exc)
        return _error_response("تعذر تحديث الفترة المحاسبية.", status=500)


@csrf_exempt
@require_http_methods(["POST"])
def accounting_period_status_api(request, period_id: int):
    """
    تغيير حالة الفترة المحاسبية.
    body:
    {
      "status": "OPEN" | "CLOSED" | "LOCKED"
    }
    """
    try:
        data = _json_body(request)
        status = _clean_text(data.get("status"))

        valid_statuses = {choice[0] for choice in AccountingPeriodStatus.choices}
        if status not in valid_statuses:
            raise ValueError("حالة الفترة المحاسبية غير صحيحة.")

        period = _get_period_or_none(period_id)

        if not period:
            return _error_response("الفترة المحاسبية المطلوبة غير موجودة.", status=404)

        period.status = status
        period.closed_at = timezone.now() if status in {
            AccountingPeriodStatus.CLOSED,
            AccountingPeriodStatus.LOCKED,
        } else None

        period.full_clean()
        period.save(update_fields=["status", "closed_at", "updated_at"])

        return _success_response(
            {
                "period": _serialize_period(period),
                "message": "تم تحديث حالة الفترة المحاسبية بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update accounting period status: %s", exc)
        return _error_response("تعذر تحديث حالة الفترة المحاسبية.", status=500)


@require_GET
def accounting_periods_excel_api(request):
    """
    تصدير الفترات المحاسبية إلى Excel.
    """
    try:
        query_data = _build_periods_queryset(request)
        ordering = _parse_period_ordering(request.GET.get("ordering"), default="-start_date")

        if ordering.lstrip("-") == "start_date":
            qs = query_data["qs"].order_by(ordering)
        else:
            qs = query_data["qs"].order_by(ordering, "-start_date")

        periods = list(qs)
        summary = _build_periods_summary(periods)

        return _build_periods_excel(
            periods,
            query_data["filters"],
            summary,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export accounting periods excel: %s", exc)
        return _error_response("تعذر تصدير الفترات المحاسبية إلى Excel.", status=500)
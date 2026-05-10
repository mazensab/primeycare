# ============================================================
# 📂 api/accounting/detail.py
# 🧠 Accounting Account Detail API — Primey Care V2
# ------------------------------------------------------------
# ✅ API تفصيلي للحساب المحاسبي
# ✅ متوافق مع Accounting Backend الجديد
# ✅ يرجع:
#    - بيانات الحساب الأساسية والتشغيلية
#    - ملخص الحركات
#    - إجمالي المدين
#    - إجمالي الدائن
#    - صافي الحركة
#    - الرصيد حسب طبيعة الحساب
#    - سطور الحركات المرتبطة بالحساب
# ✅ يدعم:
#    - date_from
#    - date_to
#    - posted_only
#    - include_metadata
# ✅ Excel Export
# ============================================================

from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import Account, JournalEntryLine, JournalEntryStatus


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Helpers
# ============================================================

def _money(value: Decimal | int | float | str | None) -> Decimal:
    amount = Decimal(str(value or "0.00"))
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    return default


def _parse_date(value: str | None, field_name: str) -> date | None:
    if not value:
        return None

    raw_value = str(value).strip()

    try:
        return date.fromisoformat(raw_value)
    except ValueError as exc:
        raise ValueError(
            f"قيمة {field_name} غير صحيحة. استخدم الصيغة YYYY-MM-DD."
        ) from exc


def _validate_date_range(date_from: date | None, date_to: date | None) -> None:
    if date_from and date_to and date_from > date_to:
        raise ValueError("لا يمكن أن يكون date_from أكبر من date_to.")


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any]) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=200,
        json_dumps_params={"ensure_ascii": False},
    )


def _safe_attr(obj: Any, attr_name: str, default: Any = None) -> Any:
    try:
        return getattr(obj, attr_name, default)
    except Exception:
        return default


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_meta_rows(ws, title: str, meta_rows: list[tuple[str, Any]], *, merge_to_column: int = 8) -> int:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_to_column)

    current_row = 3

    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(
            row=current_row,
            column=2,
            value=str(value) if value is not None else "",
        )
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    return current_row + 1


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _resolve_account_display_name(account: Account | None) -> str | None:
    if not account:
        return None

    return (
        _safe_attr(account, "name", None)
        or _safe_attr(account, "name_ar", None)
        or _safe_attr(account, "name_en", None)
        or f"Account #{account.pk}"
    )


def _resolve_account_nature(account: Account | None) -> str:
    return str(_safe_attr(account, "nature", "") or "").upper()


def _resolve_account_nature_label(account: Account | None) -> str | None:
    if not account:
        return None

    if hasattr(account, "get_nature_display"):
        return account.get_nature_display()

    nature = _resolve_account_nature(account)

    if nature == "DEBIT":
        return "مدين"

    if nature == "CREDIT":
        return "دائن"

    return None


def _can_post(account: Account) -> bool:
    return bool(
        _safe_attr(account, "is_active", True)
        and not _safe_attr(account, "is_group", False)
    )


def _can_manual_post(account: Account) -> bool:
    return bool(
        _can_post(account)
        and _safe_attr(account, "allow_manual_posting", True)
    )


def _balance_side_from_nature(
    *,
    balance_amount: Decimal,
    account_nature: str,
) -> str:
    if balance_amount == Decimal("0.00"):
        return "ZERO"

    if account_nature == "CREDIT":
        return "CREDIT"

    return "DEBIT"


def _signed_balance_by_nature(
    *,
    total_debit: Decimal,
    total_credit: Decimal,
    account_nature: str,
) -> Decimal:
    if account_nature == "CREDIT":
        return _money(total_credit - total_debit)

    return _money(total_debit - total_credit)


def _serialize_parent_account(account: Account | None) -> dict[str, Any] | None:
    if not account:
        return None

    return {
        "id": account.id,
        "code": _safe_attr(account, "code", None),
        "name": _resolve_account_display_name(account),
        "name_ar": _safe_attr(account, "name", None),
        "name_en": _safe_attr(account, "name_en", ""),
        "account_type": _safe_attr(account, "account_type", None),
        "nature": _safe_attr(account, "nature", None),
        "level": int(_safe_attr(account, "level", 1) or 1),
        "is_group": bool(_safe_attr(account, "is_group", False)),
        "is_active": bool(_safe_attr(account, "is_active", True)),
    }


def _serialize_account(account: Account, *, include_metadata: bool = False) -> dict[str, Any]:
    children_count = 0

    try:
        children_count = account.children.count()
    except Exception:
        children_count = 0

    parent = _safe_attr(account, "parent", None)

    payload = {
        "id": account.id,
        "code": _safe_attr(account, "code", None),
        "name": _resolve_account_display_name(account),
        "name_ar": _safe_attr(account, "name", None),
        "name_en": _safe_attr(account, "name_en", ""),
        "description": _safe_attr(account, "description", ""),
        "account_type": _safe_attr(account, "account_type", None),
        "account_type_label": (
            account.get_account_type_display()
            if hasattr(account, "get_account_type_display")
            else None
        ),
        "nature": _resolve_account_nature(account),
        "nature_label": _resolve_account_nature_label(account),
        "parent_id": _safe_attr(account, "parent_id", None),
        "parent_code": _safe_attr(parent, "code", None),
        "parent_name": _resolve_account_display_name(parent) if parent else None,
        "parent": _serialize_parent_account(parent),
        "level": int(_safe_attr(account, "level", 1) or 1),
        "is_group": bool(_safe_attr(account, "is_group", False)),
        "is_active": bool(_safe_attr(account, "is_active", True)),
        "allow_manual_posting": bool(_safe_attr(account, "allow_manual_posting", True)),
        "is_system": bool(_safe_attr(account, "is_system", False)),
        "currency": _safe_attr(account, "currency", "SAR"),
        "opening_balance": _money(_safe_attr(account, "opening_balance", "0.00")),
        "can_post": _can_post(account),
        "can_manual_post": _can_manual_post(account),
        "children_count": children_count,
        "has_children": children_count > 0,
        "created_at": (
            _safe_attr(account, "created_at", None).isoformat()
            if _safe_attr(account, "created_at", None)
            else None
        ),
        "updated_at": (
            _safe_attr(account, "updated_at", None).isoformat()
            if _safe_attr(account, "updated_at", None)
            else None
        ),
    }

    if include_metadata:
        payload["metadata"] = _safe_attr(account, "metadata", {}) or {}

    return payload


def _serialize_cost_center(line: JournalEntryLine) -> dict[str, Any] | None:
    cost_center = _safe_attr(line, "cost_center", None)

    if not cost_center:
        return None

    return {
        "id": _safe_attr(cost_center, "id", None),
        "code": _safe_attr(cost_center, "code", None),
        "name": _safe_attr(cost_center, "name", None),
        "name_en": _safe_attr(cost_center, "name_en", ""),
        "level": _safe_attr(cost_center, "level", None),
        "is_group": bool(_safe_attr(cost_center, "is_group", False)),
        "status": _safe_attr(cost_center, "status", None),
    }


def _serialize_tax_rate(line: JournalEntryLine) -> dict[str, Any] | None:
    tax_rate = _safe_attr(line, "tax_rate", None)

    if not tax_rate:
        return None

    return {
        "id": _safe_attr(tax_rate, "id", None),
        "code": _safe_attr(tax_rate, "code", None),
        "name": _safe_attr(tax_rate, "name", None),
        "tax_type": _safe_attr(tax_rate, "tax_type", None),
        "rate": _money(_safe_attr(tax_rate, "rate", "0.00")),
        "is_active": bool(_safe_attr(tax_rate, "is_active", True)),
        "is_default": bool(_safe_attr(tax_rate, "is_default", False)),
    }


def _serialize_line(line: JournalEntryLine, *, include_metadata: bool = False) -> dict[str, Any]:
    journal_entry = _safe_attr(line, "journal_entry", None)
    period = _safe_attr(journal_entry, "period", None)

    payload = {
        "id": line.id,
        "journal_entry_id": _safe_attr(journal_entry, "id", None),
        "journal_entry_number": _safe_attr(journal_entry, "entry_number", None),
        "entry_date": (
            _safe_attr(journal_entry, "entry_date", None).isoformat()
            if _safe_attr(journal_entry, "entry_date", None)
            else None
        ),
        "entry_status": _safe_attr(journal_entry, "status", None),
        "entry_status_label": (
            journal_entry.get_status_display()
            if journal_entry and hasattr(journal_entry, "get_status_display")
            else None
        ),
        "posting_source": _safe_attr(journal_entry, "posting_source", None),
        "posting_source_label": (
            journal_entry.get_posting_source_display()
            if journal_entry and hasattr(journal_entry, "get_posting_source_display")
            else None
        ),
        "reference": _safe_attr(journal_entry, "reference", None),
        "external_reference": _safe_attr(journal_entry, "external_reference", None),
        "source_type": _safe_attr(journal_entry, "source_type", ""),
        "source_id": _safe_attr(journal_entry, "source_id", ""),
        "source_number": _safe_attr(journal_entry, "source_number", ""),
        "period": {
            "id": _safe_attr(period, "id", None),
            "name": _safe_attr(period, "name", None),
            "status": _safe_attr(period, "status", None),
        } if period else None,
        "description": _safe_attr(line, "description", None),
        "entry_description": _safe_attr(journal_entry, "description", None),
        "debit_amount": _money(_safe_attr(line, "debit_amount", "0.00")),
        "credit_amount": _money(_safe_attr(line, "credit_amount", "0.00")),
        "tax_amount": _money(_safe_attr(line, "tax_amount", "0.00")),
        "cost_center": _serialize_cost_center(line),
        "cost_center_id": _safe_attr(line, "cost_center_id", None),
        "tax_rate": _serialize_tax_rate(line),
        "tax_rate_id": _safe_attr(line, "tax_rate_id", None),
        "party_type": _safe_attr(line, "party_type", ""),
        "party_id": _safe_attr(line, "party_id", ""),
        "source_line_id": _safe_attr(line, "source_line_id", ""),
        "sort_order": _safe_attr(line, "sort_order", 0),
        "created_at": (
            _safe_attr(line, "created_at", None).isoformat()
            if _safe_attr(line, "created_at", None)
            else None
        ),
        "updated_at": (
            _safe_attr(line, "updated_at", None).isoformat()
            if _safe_attr(line, "updated_at", None)
            else None
        ),
    }

    if include_metadata:
        payload["metadata"] = _safe_attr(line, "metadata", {}) or {}

    return payload


def _build_account_detail_payload(account_id: int, request) -> dict[str, Any]:
    date_from = _parse_date(request.GET.get("date_from"), "date_from")
    date_to = _parse_date(request.GET.get("date_to"), "date_to")
    posted_only = _parse_bool(request.GET.get("posted_only"), default=True)
    include_metadata = _parse_bool(request.GET.get("include_metadata"), default=False)

    _validate_date_range(date_from, date_to)

    account = (
        Account.objects.select_related("parent")
        .filter(id=account_id)
        .first()
    )

    if not account:
        raise ValueError("الحساب المطلوب غير موجود.")

    lines_qs = (
        JournalEntryLine.objects.select_related(
            "journal_entry",
            "journal_entry__period",
            "account",
            "cost_center",
            "tax_rate",
        )
        .filter(account_id=account.id)
    )

    if posted_only:
        lines_qs = lines_qs.filter(
            journal_entry__status__in=[
                JournalEntryStatus.POSTED,
                JournalEntryStatus.REVERSED,
            ]
        )

    if date_from:
        lines_qs = lines_qs.filter(journal_entry__entry_date__gte=date_from)

    if date_to:
        lines_qs = lines_qs.filter(journal_entry__entry_date__lte=date_to)

    lines = list(
        lines_qs.order_by(
            "journal_entry__entry_date",
            "journal_entry__id",
            "sort_order",
            "id",
        )
    )

    total_debit = _money(
        sum(
            (
                _money(_safe_attr(line, "debit_amount", "0.00"))
                for line in lines
            ),
            Decimal("0.00"),
        )
    )
    total_credit = _money(
        sum(
            (
                _money(_safe_attr(line, "credit_amount", "0.00"))
                for line in lines
            ),
            Decimal("0.00"),
        )
    )
    total_tax = _money(
        sum(
            (
                _money(_safe_attr(line, "tax_amount", "0.00"))
                for line in lines
            ),
            Decimal("0.00"),
        )
    )

    account_nature = _resolve_account_nature(account)
    net_movement = _money(total_debit - total_credit)
    balance_amount = _signed_balance_by_nature(
        total_debit=total_debit,
        total_credit=total_credit,
        account_nature=account_nature,
    )

    transactions = [
        _serialize_line(line, include_metadata=include_metadata)
        for line in lines
    ]

    return {
        "account": _serialize_account(account, include_metadata=include_metadata),
        "filters": {
            "date_from": date_from.isoformat() if date_from else None,
            "date_to": date_to.isoformat() if date_to else None,
            "posted_only": posted_only,
            "include_metadata": include_metadata,
        },
        "summary": {
            "transaction_count": len(lines),
            "total_debit": total_debit,
            "total_credit": total_credit,
            "total_tax": total_tax,
            "net_movement": net_movement,
            "balance_amount": balance_amount,
            "balance_side": _balance_side_from_nature(
                balance_amount=balance_amount,
                account_nature=account_nature,
            ),
            "posted_lines": sum(
                1
                for line in lines
                if _safe_attr(_safe_attr(line, "journal_entry", None), "status", None)
                == JournalEntryStatus.POSTED
            ),
            "reversed_lines": sum(
                1
                for line in lines
                if _safe_attr(_safe_attr(line, "journal_entry", None), "status", None)
                == JournalEntryStatus.REVERSED
            ),
            "draft_lines": sum(
                1
                for line in lines
                if _safe_attr(_safe_attr(line, "journal_entry", None), "status", None)
                == JournalEntryStatus.DRAFT
            ),
        },
        "transactions": transactions,
    }


def _build_account_detail_excel(payload: dict[str, Any]) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Account Detail")

    account = payload["account"]
    filters = payload["filters"]
    summary = payload["summary"]

    start_row = _add_meta_rows(
        ws,
        "Account Detail",
        [
            ("Account ID", account.get("id")),
            ("Account Code", account.get("code")),
            ("Account Name", account.get("name")),
            ("Account Name EN", account.get("name_en")),
            ("Account Type", account.get("account_type")),
            ("Nature", account.get("nature")),
            ("Currency", account.get("currency")),
            ("Opening Balance", account.get("opening_balance")),
            ("Is Group", account.get("is_group")),
            ("Is Active", account.get("is_active")),
            ("Allow Manual Posting", account.get("allow_manual_posting")),
            ("Is System", account.get("is_system")),
            ("Can Post", account.get("can_post")),
            ("Date From", filters.get("date_from")),
            ("Date To", filters.get("date_to")),
            ("Posted Only", filters.get("posted_only")),
            ("Transaction Count", summary.get("transaction_count")),
            ("Total Debit", summary.get("total_debit")),
            ("Total Credit", summary.get("total_credit")),
            ("Total Tax", summary.get("total_tax")),
            ("Net Movement", summary.get("net_movement")),
            ("Balance Amount", summary.get("balance_amount")),
            ("Balance Side", summary.get("balance_side")),
        ],
        merge_to_column=18,
    )

    headers = [
        "ID",
        "Journal Entry ID",
        "Journal Entry Number",
        "Entry Date",
        "Entry Status",
        "Posting Source",
        "Reference",
        "External Reference",
        "Source Type",
        "Source ID",
        "Source Number",
        "Period",
        "Line Description",
        "Entry Description",
        "Debit Amount",
        "Credit Amount",
        "Tax Rate",
        "Tax Amount",
        "Cost Center",
        "Party Type",
        "Party ID",
        "Source Line ID",
        "Sort Order",
        "Created At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for row in payload["transactions"]:
        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("journal_entry_id"))
        ws.cell(row=current_row, column=3, value=row.get("journal_entry_number"))
        ws.cell(row=current_row, column=4, value=row.get("entry_date"))
        ws.cell(row=current_row, column=5, value=row.get("entry_status"))
        ws.cell(row=current_row, column=6, value=row.get("posting_source"))
        ws.cell(row=current_row, column=7, value=row.get("reference"))
        ws.cell(row=current_row, column=8, value=row.get("external_reference"))
        ws.cell(row=current_row, column=9, value=row.get("source_type"))
        ws.cell(row=current_row, column=10, value=row.get("source_id"))
        ws.cell(row=current_row, column=11, value=row.get("source_number"))
        ws.cell(
            row=current_row,
            column=12,
            value=(row.get("period") or {}).get("name") if row.get("period") else "",
        )
        ws.cell(row=current_row, column=13, value=row.get("description"))
        ws.cell(row=current_row, column=14, value=row.get("entry_description"))
        ws.cell(row=current_row, column=15, value=float(row.get("debit_amount", 0) or 0))
        ws.cell(row=current_row, column=16, value=float(row.get("credit_amount", 0) or 0))
        ws.cell(
            row=current_row,
            column=17,
            value=(row.get("tax_rate") or {}).get("code") if row.get("tax_rate") else "",
        )
        ws.cell(row=current_row, column=18, value=float(row.get("tax_amount", 0) or 0))
        ws.cell(
            row=current_row,
            column=19,
            value=(row.get("cost_center") or {}).get("name") if row.get("cost_center") else "",
        )
        ws.cell(row=current_row, column=20, value=row.get("party_type"))
        ws.cell(row=current_row, column=21, value=row.get("party_id"))
        ws.cell(row=current_row, column=22, value=row.get("source_line_id"))
        ws.cell(row=current_row, column=23, value=row.get("sort_order"))
        ws.cell(row=current_row, column=24, value=row.get("created_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, f'account_{account.get("id")}_detail.xlsx')


# ============================================================
# 📘 Account Detail API
# ============================================================

@require_GET
def accounting_account_detail_api(request, account_id: int):
    """
    إرجاع بيانات حساب محاسبي مفرد مع ملخص الحركات وحركة الحساب.
    """
    try:
        payload = _build_account_detail_payload(account_id, request)
        return _success_response(payload)

    except ValueError as exc:
        status = 404 if str(exc) == "الحساب المطلوب غير موجود." else 400
        return _error_response(str(exc), status=status)

    except Exception as exc:
        logger.exception("Failed to load accounting account detail: %s", exc)
        return _error_response(
            "تعذر تحميل بيانات الحساب.",
            status=500,
        )


@require_GET
def accounting_account_detail_excel_api(request, account_id: int):
    """
    تصدير تفاصيل الحساب إلى Excel.
    """
    try:
        payload = _build_account_detail_payload(account_id, request)
        return _build_account_detail_excel(payload)

    except ValueError as exc:
        status = 404 if str(exc) == "الحساب المطلوب غير موجود." else 400
        return _error_response(str(exc), status=status)

    except Exception as exc:
        logger.exception("Failed to export accounting account detail excel: %s", exc)
        return _error_response(
            "تعذر تصدير تفاصيل الحساب إلى Excel.",
            status=500,
        )
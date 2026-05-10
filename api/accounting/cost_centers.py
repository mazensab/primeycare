# ============================================================
# 📂 api/accounting/cost_centers.py
# 🧠 Cost Centers API — Primey Care V2
# ------------------------------------------------------------
# ✅ قائمة مراكز التكلفة
# ✅ شجرة مراكز التكلفة
# ✅ تفاصيل مركز تكلفة
# ✅ إنشاء مركز تكلفة
# ✅ تحديث مركز تكلفة
# ✅ تفعيل / تعطيل مركز تكلفة
# ✅ Excel Export
# ✅ متوافق مع Accounting Backend الجديد
# ✅ يدعم:
#    - search
#    - status
#    - is_group
#    - can_post
#    - parent_id
#    - level
#    - tree
#    - page / page_size
#    - ordering
# ============================================================

from __future__ import annotations

import json
import logging
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import CostCenter, CostCenterStatus, JournalEntryLine


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Constants
# ============================================================

DEFAULT_PAGE = 1
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 500

ALLOWED_ORDERING = {
    "code": "code",
    "-code": "-code",
    "name": "name",
    "-name": "-name",
    "name_en": "name_en",
    "-name_en": "-name_en",
    "level": "level",
    "-level": "-level",
    "status": "status",
    "-status": "-status",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "updated_at": "updated_at",
    "-updated_at": "-updated_at",
}


# ============================================================
# 🔧 Helpers
# ============================================================

def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool | None = None) -> bool | None:
    if value in {None, ""}:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    raise ValueError("قيمة boolean غير صحيحة. استخدم true أو false.")


def _parse_int(value: str | None, field_name: str) -> int | None:
    if value in {None, ""}:
        return None

    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed <= 0:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر من صفر.")

    return parsed


def _parse_positive_int(
    value: str | None,
    field_name: str,
    *,
    default: int,
    min_value: int = 1,
    max_value: int | None = None,
) -> int:
    if value in {None, ""}:
        parsed = default
    else:
        try:
            parsed = int(str(value).strip())
        except ValueError as exc:
            raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed < min_value:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر أو تساوي {min_value}.")

    if max_value is not None and parsed > max_value:
        raise ValueError(f"قيمة {field_name} يجب ألا تتجاوز {max_value}.")

    return parsed


def _parse_ordering(value: str | None, default: str = "code") -> str:
    raw = (value or "").strip() or default

    if raw not in ALLOWED_ORDERING:
        allowed = ", ".join(ALLOWED_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_ORDERING[raw]


def _json_body(request) -> dict[str, Any]:
    try:
        body = request.body.decode("utf-8") if request.body else "{}"
        data = json.loads(body or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError("صيغة JSON غير صحيحة.") from exc

    if not isinstance(data, dict):
        raise ValueError("جسم الطلب يجب أن يكون JSON Object.")

    return data


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any], status: int = 200) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _safe_attr(obj: Any, attr_name: str, default: Any = None) -> Any:
    try:
        return getattr(obj, attr_name, default)
    except Exception:
        return default


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _status_label(status: str) -> str:
    for value, label in CostCenterStatus.choices:
        if value == status:
            return label
    return status


def _can_post(cost_center: CostCenter) -> bool:
    return bool(
        _safe_attr(cost_center, "status", None) == CostCenterStatus.ACTIVE
        and not bool(_safe_attr(cost_center, "is_group", False))
    )


def _children_count(cost_center: CostCenter) -> int:
    try:
        return cost_center.children.count()
    except Exception:
        return 0


def _journal_lines_count(cost_center: CostCenter) -> int:
    try:
        return cost_center.journal_lines.count()
    except Exception:
        return 0


def _serialize_parent(cost_center: CostCenter | None) -> dict[str, Any] | None:
    if not cost_center:
        return None

    return {
        "id": cost_center.id,
        "code": cost_center.code,
        "name": cost_center.name,
        "name_en": cost_center.name_en,
        "level": cost_center.level,
        "is_group": cost_center.is_group,
        "status": cost_center.status,
        "status_label": _status_label(cost_center.status),
    }


def _serialize_cost_center(
    cost_center: CostCenter,
    *,
    include_counts: bool = True,
    include_metadata: bool = False,
) -> dict[str, Any]:
    parent = _safe_attr(cost_center, "parent", None)

    children_count = _children_count(cost_center) if include_counts else 0
    journal_lines_count = _journal_lines_count(cost_center) if include_counts else 0

    payload = {
        "id": cost_center.id,
        "code": cost_center.code,
        "name": cost_center.name,
        "name_en": cost_center.name_en,
        "description": cost_center.description,
        "parent_id": cost_center.parent_id,
        "parent_code": _safe_attr(parent, "code", None),
        "parent_name": _safe_attr(parent, "name", None),
        "parent": _serialize_parent(parent),
        "level": cost_center.level,
        "is_group": cost_center.is_group,
        "status": cost_center.status,
        "status_label": _status_label(cost_center.status),
        "can_post": _can_post(cost_center),
        "children_count": children_count,
        "has_children": children_count > 0,
        "journal_lines_count": journal_lines_count,
        "has_journal_lines": journal_lines_count > 0,
        "created_at": cost_center.created_at.isoformat() if cost_center.created_at else None,
        "updated_at": cost_center.updated_at.isoformat() if cost_center.updated_at else None,
    }

    if include_metadata:
        payload["metadata"] = cost_center.metadata or {}

    return payload


def _build_queryset(request):
    search = (request.GET.get("search") or "").strip()
    status = (request.GET.get("status") or "").strip()
    is_group = _parse_bool(request.GET.get("is_group"), default=None)
    can_post = _parse_bool(request.GET.get("can_post"), default=None)
    parent_id = _parse_int(request.GET.get("parent_id"), "parent_id")
    level = _parse_int(request.GET.get("level"), "level")

    qs = CostCenter.objects.select_related("parent").all()

    if search:
        qs = qs.filter(
            Q(code__icontains=search)
            | Q(name__icontains=search)
            | Q(name_en__icontains=search)
            | Q(description__icontains=search)
            | Q(parent__code__icontains=search)
            | Q(parent__name__icontains=search)
            | Q(parent__name_en__icontains=search)
        )

    if status:
        valid_statuses = {choice[0] for choice in CostCenterStatus.choices}
        if status not in valid_statuses:
            raise ValueError("حالة مركز التكلفة غير صحيحة.")
        qs = qs.filter(status=status)

    if is_group is not None:
        qs = qs.filter(is_group=is_group)

    if can_post is True:
        qs = qs.filter(status=CostCenterStatus.ACTIVE, is_group=False)

    if can_post is False:
        qs = qs.exclude(status=CostCenterStatus.ACTIVE, is_group=False)

    if parent_id:
        qs = qs.filter(parent_id=parent_id)

    if level:
        qs = qs.filter(level=level)

    return {
        "qs": qs,
        "filters": {
            "search": search or None,
            "status": status or None,
            "is_group": is_group,
            "can_post": can_post,
            "parent_id": parent_id,
            "level": level,
        },
    }


def _build_tree(cost_centers: list[CostCenter]) -> list[dict[str, Any]]:
    center_map: dict[int, dict[str, Any]] = {}
    roots: list[dict[str, Any]] = []

    for cost_center in cost_centers:
        item = _serialize_cost_center(
            cost_center,
            include_counts=False,
            include_metadata=False,
        )
        item["children"] = []
        center_map[cost_center.id] = item

    for cost_center in cost_centers:
        item = center_map[cost_center.id]
        parent_id = cost_center.parent_id

        if parent_id and parent_id in center_map:
            center_map[parent_id]["children"].append(item)
        else:
            roots.append(item)

    return roots


def _build_summary(cost_centers: list[CostCenter]) -> dict[str, Any]:
    total = len(cost_centers)
    active = sum(1 for item in cost_centers if item.status == CostCenterStatus.ACTIVE)
    inactive = total - active
    group_count = sum(1 for item in cost_centers if item.is_group)
    posting_count = sum(1 for item in cost_centers if _can_post(item))

    by_status: dict[str, int] = {}
    by_level: dict[str, int] = {}

    for item in cost_centers:
        by_status[item.status] = by_status.get(item.status, 0) + 1
        level_key = str(item.level)
        by_level[level_key] = by_level.get(level_key, 0) + 1

    return {
        "total_cost_centers": total,
        "active_cost_centers": active,
        "inactive_cost_centers": inactive,
        "group_cost_centers": group_count,
        "posting_cost_centers": posting_count,
        "by_status": by_status,
        "by_level": by_level,
    }


def _choices_payload() -> dict[str, Any]:
    return {
        "statuses": [
            {"value": value, "label": label}
            for value, label in CostCenterStatus.choices
        ],
        "orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_ORDERING.keys()
        ],
    }


def _resolve_parent(parent_id: Any) -> CostCenter | None:
    if parent_id in {None, ""}:
        return None

    try:
        parsed_parent_id = int(parent_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({"parent_id": "معرف مركز التكلفة الأب غير صحيح."}) from exc

    parent = CostCenter.objects.filter(id=parsed_parent_id).first()

    if not parent:
        raise ValidationError({"parent_id": "مركز التكلفة الأب غير موجود."})

    if not parent.is_group:
        raise ValidationError({"parent_id": "مركز التكلفة الأب يجب أن يكون تجميعيًا."})

    return parent


def _apply_payload_to_cost_center(cost_center: CostCenter, data: dict[str, Any]) -> CostCenter:
    if "code" in data:
        cost_center.code = _clean_text(data.get("code"))

    if "name" in data:
        cost_center.name = _clean_text(data.get("name"))

    if "name_en" in data:
        cost_center.name_en = _clean_text(data.get("name_en"))

    if "description" in data:
        cost_center.description = _clean_text(data.get("description"))

    if "is_group" in data:
        is_group = data.get("is_group")
        if not isinstance(is_group, bool):
            raise ValidationError({"is_group": "قيمة is_group يجب أن تكون true أو false."})
        cost_center.is_group = is_group

    if "status" in data:
        status = _clean_text(data.get("status"))
        valid_statuses = {choice[0] for choice in CostCenterStatus.choices}
        if status not in valid_statuses:
            raise ValidationError({"status": "حالة مركز التكلفة غير صحيحة."})
        cost_center.status = status

    if "parent_id" in data:
        cost_center.parent = _resolve_parent(data.get("parent_id"))

    if "metadata" in data:
        metadata = data.get("metadata")
        if metadata in {None, ""}:
            cost_center.metadata = {}
        elif not isinstance(metadata, dict):
            raise ValidationError({"metadata": "metadata يجب أن تكون JSON Object."})
        else:
            cost_center.metadata = metadata

    return cost_center


def _validation_error_payload(exc: ValidationError) -> dict[str, Any]:
    if hasattr(exc, "message_dict"):
        return {
            "ok": False,
            "message": "تعذر حفظ مركز التكلفة.",
            "errors": exc.message_dict,
        }

    return {
        "ok": False,
        "message": "تعذر حفظ مركز التكلفة.",
        "errors": exc.messages if hasattr(exc, "messages") else [str(exc)],
    }


# ============================================================
# 📊 Excel
# ============================================================

def _build_cost_centers_excel(
    cost_centers: list[CostCenter],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Cost Centers")

    ws.append(["Cost Centers"])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)

    meta_rows = [
        ("Search", filters.get("search")),
        ("Status", filters.get("status")),
        ("Is Group", filters.get("is_group")),
        ("Can Post", filters.get("can_post")),
        ("Parent ID", filters.get("parent_id")),
        ("Level", filters.get("level")),
        ("Total Cost Centers", summary.get("total_cost_centers")),
        ("Active Cost Centers", summary.get("active_cost_centers")),
        ("Posting Cost Centers", summary.get("posting_cost_centers")),
        ("Group Cost Centers", summary.get("group_cost_centers")),
    ]

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    start_row = current_row + 1

    headers = [
        "ID",
        "Code",
        "Name",
        "Name EN",
        "Parent ID",
        "Parent Code",
        "Parent Name",
        "Level",
        "Is Group",
        "Status",
        "Status Label",
        "Can Post",
        "Children Count",
        "Journal Lines Count",
        "Description",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for cost_center in cost_centers:
        row = _serialize_cost_center(cost_center)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("code"))
        ws.cell(row=current_row, column=3, value=row.get("name"))
        ws.cell(row=current_row, column=4, value=row.get("name_en"))
        ws.cell(row=current_row, column=5, value=row.get("parent_id"))
        ws.cell(row=current_row, column=6, value=row.get("parent_code"))
        ws.cell(row=current_row, column=7, value=row.get("parent_name"))
        ws.cell(row=current_row, column=8, value=row.get("level"))
        ws.cell(row=current_row, column=9, value=str(row.get("is_group")))
        ws.cell(row=current_row, column=10, value=row.get("status"))
        ws.cell(row=current_row, column=11, value=row.get("status_label"))
        ws.cell(row=current_row, column=12, value=str(row.get("can_post")))
        ws.cell(row=current_row, column=13, value=row.get("children_count"))
        ws.cell(row=current_row, column=14, value=row.get("journal_lines_count"))
        ws.cell(row=current_row, column=15, value=row.get("description"))
        ws.cell(row=current_row, column=16, value=row.get("created_at"))
        ws.cell(row=current_row, column=17, value=row.get("updated_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "cost_centers.xlsx")


# ============================================================
# 📘 List / Create API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "POST"])
def accounting_cost_centers_api(request):
    """
    GET:
      قائمة مراكز التكلفة.

    POST:
      إنشاء مركز تكلفة.
    """
    if request.method == "POST":
        return _create_cost_center(request)

    return _list_cost_centers(request)


def _list_cost_centers(request):
    try:
        query_data = _build_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_ordering(request.GET.get("ordering"), default="code")
        as_tree = _parse_bool(request.GET.get("tree"), default=False)
        include_metadata = _parse_bool(request.GET.get("include_metadata"), default=False)

        if ordering.lstrip("-") == "code":
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by(ordering, "code")

        cost_centers_all = list(qs)
        summary = _build_summary(cost_centers_all)

        if as_tree:
            return _success_response(
                {
                    "filters": {
                        **query_data["filters"],
                        "ordering": ordering,
                        "tree": True,
                        "include_metadata": False,
                    },
                    "summary": summary,
                    "choices": _choices_payload(),
                    "results": _build_tree(cost_centers_all),
                }
            )

        paginator = Paginator(cost_centers_all, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        return _success_response(
            {
                "filters": {
                    **query_data["filters"],
                    "ordering": ordering,
                    "tree": False,
                    "include_metadata": include_metadata,
                },
                "summary": summary,
                "choices": _choices_payload(),
                "pagination": {
                    "page": page_obj.number,
                    "page_size": page_size,
                    "total_pages": paginator.num_pages,
                    "total_items": paginator.count,
                    "has_next": page_obj.has_next(),
                    "has_previous": page_obj.has_previous(),
                    "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                    "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
                },
                "results": [
                    _serialize_cost_center(
                        item,
                        include_metadata=include_metadata,
                    )
                    for item in page_obj.object_list
                ],
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load cost centers: %s", exc)
        return _error_response("تعذر تحميل مراكز التكلفة.", status=500)


@transaction.atomic
def _create_cost_center(request):
    try:
        data = _json_body(request)

        cost_center = CostCenter()
        _apply_payload_to_cost_center(cost_center, data)
        cost_center.full_clean()
        cost_center.save()

        return _success_response(
            {
                "cost_center": _serialize_cost_center(
                    cost_center,
                    include_metadata=True,
                ),
                "message": "تم إنشاء مركز التكلفة بنجاح.",
            },
            status=201,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to create cost center: %s", exc)
        return _error_response("تعذر إنشاء مركز التكلفة.", status=500)


# ============================================================
# 📘 Detail / Update API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT"])
def accounting_cost_center_detail_api(request, cost_center_id: int):
    """
    GET:
      تفاصيل مركز تكلفة.

    PATCH/PUT:
      تحديث مركز تكلفة.
    """
    if request.method in {"PATCH", "PUT"}:
        return _update_cost_center(request, cost_center_id)

    return _detail_cost_center(request, cost_center_id)


def _get_cost_center_or_none(cost_center_id: int) -> CostCenter | None:
    return (
        CostCenter.objects.select_related("parent")
        .filter(id=cost_center_id)
        .first()
    )


def _detail_cost_center(request, cost_center_id: int):
    try:
        include_metadata = _parse_bool(request.GET.get("include_metadata"), default=False)

        cost_center = _get_cost_center_or_none(cost_center_id)

        if not cost_center:
            return _error_response("مركز التكلفة المطلوب غير موجود.", status=404)

        children = list(cost_center.children.order_by("code"))
        recent_lines = list(
            JournalEntryLine.objects.select_related(
                "journal_entry",
                "account",
            )
            .filter(cost_center_id=cost_center.id)
            .order_by("-journal_entry__entry_date", "-id")[:20]
        )

        payload = {
            "cost_center": _serialize_cost_center(
                cost_center,
                include_metadata=include_metadata,
            ),
            "children": [
                _serialize_cost_center(child, include_metadata=False)
                for child in children
            ],
            "recent_journal_lines": [
                {
                    "id": line.id,
                    "journal_entry_id": line.journal_entry_id,
                    "journal_entry_number": line.journal_entry.entry_number if line.journal_entry else None,
                    "entry_date": line.journal_entry.entry_date.isoformat() if line.journal_entry and line.journal_entry.entry_date else None,
                    "account_id": line.account_id,
                    "account_code": line.account.code if line.account else None,
                    "account_name": line.account.name if line.account else None,
                    "description": line.description,
                    "debit_amount": str(line.debit_amount),
                    "credit_amount": str(line.credit_amount),
                }
                for line in recent_lines
            ],
            "filters": {
                "include_metadata": include_metadata,
            },
        }

        return _success_response(payload)

    except Exception as exc:
        logger.exception("Failed to load cost center detail: %s", exc)
        return _error_response("تعذر تحميل تفاصيل مركز التكلفة.", status=500)


@transaction.atomic
def _update_cost_center(request, cost_center_id: int):
    try:
        data = _json_body(request)
        cost_center = _get_cost_center_or_none(cost_center_id)

        if not cost_center:
            return _error_response("مركز التكلفة المطلوب غير موجود.", status=404)

        old_is_group = cost_center.is_group

        _apply_payload_to_cost_center(cost_center, data)

        if old_is_group and cost_center.is_group is False and cost_center.children.exists():
            raise ValidationError(
                {"is_group": "لا يمكن تحويل مركز تكلفة لديه فروع إلى مركز غير تجميعي."}
            )

        cost_center.full_clean()
        cost_center.save()

        return _success_response(
            {
                "cost_center": _serialize_cost_center(
                    cost_center,
                    include_metadata=True,
                ),
                "message": "تم تحديث مركز التكلفة بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update cost center: %s", exc)
        return _error_response("تعذر تحديث مركز التكلفة.", status=500)


# ============================================================
# ✅ Status API
# ============================================================

@csrf_exempt
@require_http_methods(["POST"])
def accounting_cost_center_status_api(request, cost_center_id: int):
    """
    تغيير حالة مركز تكلفة.
    body:
    {
      "status": "ACTIVE" | "INACTIVE"
    }
    """
    try:
        data = _json_body(request)
        status = _clean_text(data.get("status"))

        valid_statuses = {choice[0] for choice in CostCenterStatus.choices}
        if status not in valid_statuses:
            raise ValueError("حالة مركز التكلفة غير صحيحة.")

        cost_center = _get_cost_center_or_none(cost_center_id)

        if not cost_center:
            return _error_response("مركز التكلفة المطلوب غير موجود.", status=404)

        cost_center.status = status
        cost_center.full_clean()
        cost_center.save(update_fields=["status", "updated_at"])

        return _success_response(
            {
                "cost_center": _serialize_cost_center(cost_center),
                "message": "تم تحديث حالة مركز التكلفة بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update cost center status: %s", exc)
        return _error_response("تعذر تحديث حالة مركز التكلفة.", status=500)


# ============================================================
# 📤 Excel API
# ============================================================

@require_GET
def accounting_cost_centers_excel_api(request):
    """
    تصدير مراكز التكلفة إلى Excel.
    """
    try:
        query_data = _build_queryset(request)
        ordering = _parse_ordering(request.GET.get("ordering"), default="code")

        if ordering.lstrip("-") == "code":
            qs = query_data["qs"].order_by(ordering)
        else:
            qs = query_data["qs"].order_by(ordering, "code")

        cost_centers = list(qs)
        summary = _build_summary(cost_centers)

        return _build_cost_centers_excel(
            cost_centers,
            query_data["filters"],
            summary,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export cost centers excel: %s", exc)
        return _error_response("تعذر تصدير مراكز التكلفة إلى Excel.", status=500)
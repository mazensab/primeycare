# ============================================================
# 📂 api/accounting/journals.py
# 🧠 Journal Entries List API — Primey Care V2
# ------------------------------------------------------------
# ✅ قائمة القيود اليومية
# ✅ متوافق مع Accounting Backend الجديد
# ✅ يدعم:
#    - date_from / date_to
#    - posting_source
#    - status
#    - reference
#    - external_reference
#    - source_type / source_id / source_number
#    - period_id
#    - is_auto_posted
#    - is_balanced
#    - search
#    - page / page_size
#    - ordering
# ✅ Excel Export
# ✅ يرجع:
#    - قائمة القيود
#    - الإجماليات
#    - حالة التوازن
#    - بيانات التصفح pagination
#    - choices
# ============================================================

from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from django.core.paginator import EmptyPage, Paginator
from django.db.models import Count, F, Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import (
    AccountingPeriod,
    JournalEntry,
    JournalEntryStatus,
    PostingSource,
)


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Constants
# ============================================================

DEFAULT_PAGE = 1
DEFAULT_PAGE_SIZE = 20
MAX_PAGE_SIZE = 200

ALLOWED_ORDERING = {
    "entry_date": "entry_date",
    "-entry_date": "-entry_date",
    "id": "id",
    "-id": "-id",
    "entry_number": "entry_number",
    "-entry_number": "-entry_number",
    "status": "status",
    "-status": "-status",
    "posting_source": "posting_source",
    "-posting_source": "-posting_source",
    "source_type": "source_type",
    "-source_type": "-source_type",
    "source_number": "source_number",
    "-source_number": "-source_number",
    "currency": "currency",
    "-currency": "-currency",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "updated_at": "updated_at",
    "-updated_at": "-updated_at",
    "posted_at": "posted_at",
    "-posted_at": "-posted_at",
    "cancelled_at": "cancelled_at",
    "-cancelled_at": "-cancelled_at",
    "reversed_at": "reversed_at",
    "-reversed_at": "-reversed_at",
    "total_debit": "total_debit",
    "-total_debit": "-total_debit",
    "total_credit": "total_credit",
    "-total_credit": "-total_credit",
}


# ============================================================
# 🔧 Helpers
# ============================================================

def _money(value: Decimal | int | float | str | None) -> Decimal:
    amount = Decimal(str(value or "0.00"))
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool | None = None) -> bool | None:
    if value in {None, ""}:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    raise ValueError("قيمة boolean غير صحيحة. استخدم true أو false.")


def _parse_date(value: str | None, field_name: str) -> date | None:
    if not value:
        return None

    raw_value = str(value).strip()

    try:
        return date.fromisoformat(raw_value)
    except ValueError as exc:
        raise ValueError(
            f"قيمة {field_name} غير صحيحة. استخدم الصيغة YYYY-MM-DD."
        ) from exc


def _parse_int(value: str | None, field_name: str) -> int | None:
    if value in {None, ""}:
        return None

    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed <= 0:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر من صفر.")

    return parsed


def _parse_positive_int(
    value: str | None,
    field_name: str,
    *,
    default: int,
    min_value: int = 1,
    max_value: int | None = None,
) -> int:
    if value in {None, ""}:
        parsed = default
    else:
        raw_value = str(value).strip()
        try:
            parsed = int(raw_value)
        except ValueError as exc:
            raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed < min_value:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر أو تساوي {min_value}.")

    if max_value is not None and parsed > max_value:
        raise ValueError(f"قيمة {field_name} يجب ألا تتجاوز {max_value}.")

    return parsed


def _parse_ordering(value: str | None, default: str = "-entry_date") -> str:
    raw_value = (value or "").strip() or default

    if raw_value not in ALLOWED_ORDERING:
        allowed = ", ".join(ALLOWED_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_ORDERING[raw_value]


def _validate_date_range(date_from: date | None, date_to: date | None) -> None:
    if date_from and date_to and date_from > date_to:
        raise ValueError("لا يمكن أن يكون date_from أكبر من date_to.")


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any]) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=200,
        json_dumps_params={"ensure_ascii": False},
    )


def _safe_attr(obj: Any, attr_name: str, default: Any = None) -> Any:
    try:
        return getattr(obj, attr_name, default)
    except Exception:
        return default


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_meta_rows(
    ws,
    title: str,
    meta_rows: list[tuple[str, Any]],
    *,
    merge_to_column: int = 12,
) -> int:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_to_column)

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    return current_row + 1


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _serialize_period(period: AccountingPeriod | None) -> dict[str, Any] | None:
    if not period:
        return None

    fiscal_year = _safe_attr(period, "fiscal_year", None)

    return {
        "id": period.id,
        "name": period.name,
        "status": period.status,
        "status_label": (
            period.get_status_display()
            if hasattr(period, "get_status_display")
            else period.status
        ),
        "start_date": period.start_date.isoformat() if period.start_date else None,
        "end_date": period.end_date.isoformat() if period.end_date else None,
        "fiscal_year": {
            "id": _safe_attr(fiscal_year, "id", None),
            "name": _safe_attr(fiscal_year, "name", None),
            "status": _safe_attr(fiscal_year, "status", None),
        } if fiscal_year else None,
    }


def _serialize_user(user) -> dict[str, Any] | None:
    if not user:
        return None

    display_name = (
        _safe_attr(user, "get_full_name", lambda: "")()
        if callable(_safe_attr(user, "get_full_name", None))
        else ""
    )

    if not display_name:
        display_name = (
            _safe_attr(user, "email", "")
            or _safe_attr(user, "username", "")
            or f"User #{_safe_attr(user, 'pk', '')}"
        )

    return {
        "id": _safe_attr(user, "id", None),
        "name": display_name,
        "email": _safe_attr(user, "email", ""),
        "username": _safe_attr(user, "username", ""),
    }


def _serialize_entry_ref(entry: JournalEntry | None) -> dict[str, Any] | None:
    if not entry:
        return None

    return {
        "id": entry.id,
        "entry_number": entry.entry_number,
        "entry_date": entry.entry_date.isoformat() if entry.entry_date else None,
        "status": entry.status,
        "posting_source": entry.posting_source,
        "reference": entry.reference,
        "external_reference": entry.external_reference,
    }


def _serialize_journal_entry(
    entry: JournalEntry,
    *,
    include_metadata: bool = False,
) -> dict[str, Any]:
    total_debit = _money(_safe_attr(entry, "total_debit", "0.00"))
    total_credit = _money(_safe_attr(entry, "total_credit", "0.00"))
    lines_count = int(_safe_attr(entry, "lines_count", 0) or 0)

    payload = {
        "id": entry.id,
        "entry_number": _safe_attr(entry, "entry_number", None),
        "entry_date": (
            _safe_attr(entry, "entry_date", None).isoformat()
            if _safe_attr(entry, "entry_date", None)
            else None
        ),
        "period": _serialize_period(_safe_attr(entry, "period", None)),
        "period_id": _safe_attr(entry, "period_id", None),
        "status": _safe_attr(entry, "status", None),
        "status_label": (
            entry.get_status_display()
            if hasattr(entry, "get_status_display")
            else _safe_attr(entry, "status", None)
        ),
        "posting_source": _safe_attr(entry, "posting_source", None),
        "posting_source_label": (
            entry.get_posting_source_display()
            if hasattr(entry, "get_posting_source_display")
            else _safe_attr(entry, "posting_source", None)
        ),
        "reference": _safe_attr(entry, "reference", None),
        "external_reference": _safe_attr(entry, "external_reference", None),
        "source_type": _safe_attr(entry, "source_type", ""),
        "source_id": _safe_attr(entry, "source_id", ""),
        "source_number": _safe_attr(entry, "source_number", ""),
        "description": _safe_attr(entry, "description", None),
        "notes": _safe_attr(entry, "notes", None),
        "currency": _safe_attr(entry, "currency", "SAR"),
        "total_debit": total_debit,
        "total_credit": total_credit,
        "is_balanced": total_debit == total_credit,
        "is_zero": total_debit == Decimal("0.00") and total_credit == Decimal("0.00"),
        "lines_count": lines_count,
        "is_auto_posted": bool(_safe_attr(entry, "is_auto_posted", False)),
        "reversal_of": _serialize_entry_ref(_safe_attr(entry, "reversal_of", None)),
        "reversal_of_id": _safe_attr(entry, "reversal_of_id", None),
        "reversed_entry": _serialize_entry_ref(_safe_attr(entry, "reversed_entry", None)),
        "reversed_entry_id": _safe_attr(entry, "reversed_entry_id", None),
        "created_by": _serialize_user(_safe_attr(entry, "created_by", None)),
        "posted_by": _serialize_user(_safe_attr(entry, "posted_by", None)),
        "cancelled_by": _serialize_user(_safe_attr(entry, "cancelled_by", None)),
        "posted_at": (
            _safe_attr(entry, "posted_at", None).isoformat()
            if _safe_attr(entry, "posted_at", None)
            else None
        ),
        "cancelled_at": (
            _safe_attr(entry, "cancelled_at", None).isoformat()
            if _safe_attr(entry, "cancelled_at", None)
            else None
        ),
        "reversed_at": (
            _safe_attr(entry, "reversed_at", None).isoformat()
            if _safe_attr(entry, "reversed_at", None)
            else None
        ),
        "created_at": (
            _safe_attr(entry, "created_at", None).isoformat()
            if _safe_attr(entry, "created_at", None)
            else None
        ),
        "updated_at": (
            _safe_attr(entry, "updated_at", None).isoformat()
            if _safe_attr(entry, "updated_at", None)
            else None
        ),
    }

    if include_metadata:
        payload["metadata"] = _safe_attr(entry, "metadata", {}) or {}

    return payload


def _validate_choice(value: str, allowed_values: set[str], error_message: str) -> None:
    if value and value not in allowed_values:
        raise ValueError(error_message)


def _build_journals_queryset(request):
    date_from = _parse_date(request.GET.get("date_from"), "date_from")
    date_to = _parse_date(request.GET.get("date_to"), "date_to")
    _validate_date_range(date_from, date_to)

    search = (request.GET.get("search") or "").strip()
    posting_source = (request.GET.get("posting_source") or "").strip()
    status = (request.GET.get("status") or "").strip()
    reference = (request.GET.get("reference") or "").strip()
    external_reference = (request.GET.get("external_reference") or "").strip()
    source_type = (request.GET.get("source_type") or "").strip()
    source_id = (request.GET.get("source_id") or "").strip()
    source_number = (request.GET.get("source_number") or "").strip()
    currency = (request.GET.get("currency") or "").strip().upper()
    period_id = _parse_int(request.GET.get("period_id"), "period_id")
    is_auto_posted = _parse_bool(request.GET.get("is_auto_posted"), default=None)
    is_balanced = _parse_bool(request.GET.get("is_balanced"), default=None)

    _validate_choice(
        posting_source,
        {choice[0] for choice in PostingSource.choices},
        "مصدر القيد غير صحيح.",
    )
    _validate_choice(
        status,
        {choice[0] for choice in JournalEntryStatus.choices},
        "حالة القيد غير صحيحة.",
    )

    qs = (
        JournalEntry.objects.select_related(
            "period",
            "period__fiscal_year",
            "reversal_of",
            "reversed_entry",
            "created_by",
            "posted_by",
            "cancelled_by",
        )
        .annotate(lines_count=Count("lines"))
        .all()
    )

    if search:
        qs = qs.filter(
            Q(entry_number__icontains=search)
            | Q(reference__icontains=search)
            | Q(external_reference__icontains=search)
            | Q(source_type__icontains=search)
            | Q(source_id__icontains=search)
            | Q(source_number__icontains=search)
            | Q(description__icontains=search)
            | Q(notes__icontains=search)
        )

    if date_from:
        qs = qs.filter(entry_date__gte=date_from)

    if date_to:
        qs = qs.filter(entry_date__lte=date_to)

    if posting_source:
        qs = qs.filter(posting_source=posting_source)

    if status:
        qs = qs.filter(status=status)

    if reference:
        qs = qs.filter(reference__icontains=reference)

    if external_reference:
        qs = qs.filter(external_reference__icontains=external_reference)

    if source_type:
        qs = qs.filter(source_type__icontains=source_type)

    if source_id:
        qs = qs.filter(source_id=source_id)

    if source_number:
        qs = qs.filter(source_number__icontains=source_number)

    if currency:
        qs = qs.filter(currency__iexact=currency)

    if period_id:
        qs = qs.filter(period_id=period_id)

    if is_auto_posted is not None:
        qs = qs.filter(is_auto_posted=is_auto_posted)

    if is_balanced is True:
        qs = qs.filter(total_debit=F("total_credit"))

    if is_balanced is False:
        qs = qs.exclude(total_debit=F("total_credit"))

    return {
        "qs": qs,
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "search": search,
            "posting_source": posting_source,
            "status": status,
            "reference": reference,
            "external_reference": external_reference,
            "source_type": source_type,
            "source_id": source_id,
            "source_number": source_number,
            "currency": currency,
            "period_id": period_id,
            "is_auto_posted": is_auto_posted,
            "is_balanced": is_balanced,
        },
    }


def _filters_payload(filters: dict[str, Any], ordering: str, include_metadata: bool = False) -> dict[str, Any]:
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")

    return {
        "date_from": date_from.isoformat() if date_from else None,
        "date_to": date_to.isoformat() if date_to else None,
        "search": filters.get("search") or None,
        "posting_source": filters.get("posting_source") or None,
        "status": filters.get("status") or None,
        "reference": filters.get("reference") or None,
        "external_reference": filters.get("external_reference") or None,
        "source_type": filters.get("source_type") or None,
        "source_id": filters.get("source_id") or None,
        "source_number": filters.get("source_number") or None,
        "currency": filters.get("currency") or None,
        "period_id": filters.get("period_id"),
        "is_auto_posted": filters.get("is_auto_posted"),
        "is_balanced": filters.get("is_balanced"),
        "ordering": ordering,
        "include_metadata": include_metadata,
    }


def _build_summary(entries: list[JournalEntry]) -> dict[str, Any]:
    total_entries = len(entries)
    total_debit = _money(
        sum(
            (_money(_safe_attr(item, "total_debit", "0.00")) for item in entries),
            Decimal("0.00"),
        )
    )
    total_credit = _money(
        sum(
            (_money(_safe_attr(item, "total_credit", "0.00")) for item in entries),
            Decimal("0.00"),
        )
    )

    balanced_entries_count = sum(
        1
        for item in entries
        if _money(_safe_attr(item, "total_debit", "0.00"))
        == _money(_safe_attr(item, "total_credit", "0.00"))
    )

    posted_entries = sum(1 for item in entries if item.status == JournalEntryStatus.POSTED)
    draft_entries = sum(1 for item in entries if item.status == JournalEntryStatus.DRAFT)
    cancelled_entries = sum(1 for item in entries if item.status == JournalEntryStatus.CANCELLED)
    reversed_entries = sum(1 for item in entries if item.status == JournalEntryStatus.REVERSED)
    auto_posted_entries = sum(1 for item in entries if bool(_safe_attr(item, "is_auto_posted", False)))

    by_status: dict[str, int] = {}
    by_source: dict[str, int] = {}
    by_currency: dict[str, int] = {}

    for entry in entries:
        by_status[entry.status] = by_status.get(entry.status, 0) + 1
        by_source[entry.posting_source] = by_source.get(entry.posting_source, 0) + 1
        by_currency[entry.currency] = by_currency.get(entry.currency, 0) + 1

    return {
        "total_entries": total_entries,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "balanced_entries_count": balanced_entries_count,
        "unbalanced_entries_count": total_entries - balanced_entries_count,
        "is_balanced_total": total_debit == total_credit,
        "posted_entries": posted_entries,
        "draft_entries": draft_entries,
        "cancelled_entries": cancelled_entries,
        "reversed_entries": reversed_entries,
        "auto_posted_entries": auto_posted_entries,
        "manual_entries": total_entries - auto_posted_entries,
        "by_status": by_status,
        "by_source": by_source,
        "by_currency": by_currency,
    }


def _choices_payload() -> dict[str, Any]:
    return {
        "statuses": [
            {"value": value, "label": label}
            for value, label in JournalEntryStatus.choices
        ],
        "posting_sources": [
            {"value": value, "label": label}
            for value, label in PostingSource.choices
        ],
        "orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_ORDERING.keys()
        ],
    }


def _build_journals_excel(
    entries: list[JournalEntry],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Journals")

    start_row = _add_meta_rows(
        ws,
        "Journal Entries",
        [
            ("Date From", filters.get("date_from")),
            ("Date To", filters.get("date_to")),
            ("Search", filters.get("search")),
            ("Posting Source", filters.get("posting_source")),
            ("Status", filters.get("status")),
            ("Reference", filters.get("reference")),
            ("External Reference", filters.get("external_reference")),
            ("Source Type", filters.get("source_type")),
            ("Source ID", filters.get("source_id")),
            ("Source Number", filters.get("source_number")),
            ("Currency", filters.get("currency")),
            ("Period ID", filters.get("period_id")),
            ("Is Auto Posted", filters.get("is_auto_posted")),
            ("Is Balanced", filters.get("is_balanced")),
            ("Ordering", filters.get("ordering")),
            ("Total Entries", summary.get("total_entries")),
            ("Total Debit", summary.get("total_debit")),
            ("Total Credit", summary.get("total_credit")),
            ("Balanced Entries", summary.get("balanced_entries_count")),
            ("Unbalanced Entries", summary.get("unbalanced_entries_count")),
            ("Posted Entries", summary.get("posted_entries")),
            ("Draft Entries", summary.get("draft_entries")),
            ("Cancelled Entries", summary.get("cancelled_entries")),
            ("Reversed Entries", summary.get("reversed_entries")),
        ],
        merge_to_column=24,
    )

    headers = [
        "ID",
        "Entry Number",
        "Entry Date",
        "Period",
        "Status",
        "Status Label",
        "Posting Source",
        "Posting Source Label",
        "Reference",
        "External Reference",
        "Source Type",
        "Source ID",
        "Source Number",
        "Description",
        "Notes",
        "Currency",
        "Total Debit",
        "Total Credit",
        "Is Balanced",
        "Lines Count",
        "Is Auto Posted",
        "Reversal Of",
        "Reversed Entry",
        "Posted At",
        "Cancelled At",
        "Reversed At",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for entry in entries:
        row = _serialize_journal_entry(entry)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("entry_number"))
        ws.cell(row=current_row, column=3, value=row.get("entry_date"))
        ws.cell(row=current_row, column=4, value=(row.get("period") or {}).get("name") if row.get("period") else "")
        ws.cell(row=current_row, column=5, value=row.get("status"))
        ws.cell(row=current_row, column=6, value=row.get("status_label"))
        ws.cell(row=current_row, column=7, value=row.get("posting_source"))
        ws.cell(row=current_row, column=8, value=row.get("posting_source_label"))
        ws.cell(row=current_row, column=9, value=row.get("reference"))
        ws.cell(row=current_row, column=10, value=row.get("external_reference"))
        ws.cell(row=current_row, column=11, value=row.get("source_type"))
        ws.cell(row=current_row, column=12, value=row.get("source_id"))
        ws.cell(row=current_row, column=13, value=row.get("source_number"))
        ws.cell(row=current_row, column=14, value=row.get("description"))
        ws.cell(row=current_row, column=15, value=row.get("notes"))
        ws.cell(row=current_row, column=16, value=row.get("currency"))
        ws.cell(row=current_row, column=17, value=float(row.get("total_debit", 0) or 0))
        ws.cell(row=current_row, column=18, value=float(row.get("total_credit", 0) or 0))
        ws.cell(row=current_row, column=19, value=str(row.get("is_balanced")))
        ws.cell(row=current_row, column=20, value=row.get("lines_count"))
        ws.cell(row=current_row, column=21, value=str(row.get("is_auto_posted")))
        ws.cell(
            row=current_row,
            column=22,
            value=(row.get("reversal_of") or {}).get("entry_number") if row.get("reversal_of") else "",
        )
        ws.cell(
            row=current_row,
            column=23,
            value=(row.get("reversed_entry") or {}).get("entry_number") if row.get("reversed_entry") else "",
        )
        ws.cell(row=current_row, column=24, value=row.get("posted_at"))
        ws.cell(row=current_row, column=25, value=row.get("cancelled_at"))
        ws.cell(row=current_row, column=26, value=row.get("reversed_at"))
        ws.cell(row=current_row, column=27, value=row.get("created_at"))
        ws.cell(row=current_row, column=28, value=row.get("updated_at"))

        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "journals.xlsx")


# ============================================================
# 📘 Journal Entries List API
# ============================================================

@require_GET
def accounting_journal_entries_api(request):
    """
    قائمة القيود اليومية.
    """
    try:
        query_data = _build_journals_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_ordering(request.GET.get("ordering"), default="-entry_date")
        include_metadata = _parse_bool(request.GET.get("include_metadata"), default=False)

        secondary_ordering = "-id"
        if ordering == "-id":
            secondary_ordering = "-entry_date"
        elif ordering == "id":
            secondary_ordering = "entry_date"

        qs = qs.order_by(ordering, secondary_ordering)
        entries_all = list(qs)
        summary = _build_summary(entries_all)

        paginator = Paginator(entries_all, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        payload = {
            "filters": _filters_payload(
                query_data["filters"],
                ordering,
                include_metadata=include_metadata,
            ),
            "summary": summary,
            "choices": _choices_payload(),
            "pagination": {
                "page": page_obj.number,
                "page_size": page_size,
                "total_pages": paginator.num_pages,
                "total_items": paginator.count,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
                "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
            },
            "results": [
                _serialize_journal_entry(entry, include_metadata=include_metadata)
                for entry in page_obj.object_list
            ],
        }

        return _success_response(payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load journal entries list: %s", exc)
        return _error_response(
            "تعذر تحميل قائمة القيود اليومية.",
            status=500,
        )


@require_GET
def accounting_journal_entries_excel_api(request):
    """
    تصدير القيود اليومية إلى Excel.
    """
    try:
        query_data = _build_journals_queryset(request)
        ordering = _parse_ordering(request.GET.get("ordering"), default="-entry_date")

        secondary_ordering = "-id"
        if ordering == "-id":
            secondary_ordering = "-entry_date"
        elif ordering == "id":
            secondary_ordering = "entry_date"

        qs = query_data["qs"].order_by(ordering, secondary_ordering)
        entries_all = list(qs)
        summary = _build_summary(entries_all)

        filters = _filters_payload(
            query_data["filters"],
            ordering,
            include_metadata=False,
        )

        return _build_journals_excel(entries_all, filters, summary)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export journal entries excel: %s", exc)
        return _error_response(
            "تعذر تصدير القيود اليومية إلى Excel.",
            status=500,
        )
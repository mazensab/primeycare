# ============================================================
# 📂 api/accounting/list.py
# 🧠 Accounting Reports API — Primey Care V1.3
# ------------------------------------------------------------
# ✅ طبقة API للتقارير المحاسبية
# ✅ تدعم:
#    - Trial Balance
#    - Profit & Loss
#    - Balance Sheet
# ✅ Export Hooks:
#    - Trial Balance Excel
#    - Profit & Loss Excel
#    - Balance Sheet Excel
# ✅ مبنية فوق accounting/services.py
# ✅ مع معالجة أخطاء واضحة وآمنة
# ✅ تدعم فلاتر التاريخ
# ------------------------------------------------------------

from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.services import (
    build_balance_sheet_payload,
    build_profit_and_loss_payload,
    build_trial_balance_payload,
)

logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Helpers
# ============================================================

def _decimal_to_string(value: Any) -> Any:
    """
    تحويل Decimal إلى string بشكل آمن داخل أي هيكل بيانات.
    """
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _parse_date(value: str | None, field_name: str) -> date | None:
    """
    تحويل التاريخ من query params بصيغة YYYY-MM-DD.
    """
    if not value:
        return None

    raw_value = str(value).strip()
    try:
        return date.fromisoformat(raw_value)
    except ValueError as exc:
        raise ValueError(
            f"قيمة {field_name} غير صحيحة. استخدم الصيغة YYYY-MM-DD."
        ) from exc


def _error_response(message: str, status: int = 400, extra: dict | None = None) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }
    if extra:
        payload.update(extra)
    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(report_code: str, data: dict) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "report_code": report_code,
            "data": _decimal_to_string(data),
        },
        status=200,
        json_dumps_params={"ensure_ascii": False},
    )


def _validate_date_range(date_from: date | None, date_to: date | None) -> None:
    """
    التحقق من منطقية نطاق التاريخ.
    """
    if date_from and date_to and date_from > date_to:
        raise ValueError("لا يمكن أن يكون date_from أكبر من date_to.")


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    clean = title
    for char in invalid_chars:
        clean = clean.replace(char, "-")
    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_meta_rows(ws, title: str, meta_rows: list[tuple[str, Any]]) -> int:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    return current_row + 1


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_trial_balance_excel(payload: dict[str, Any]) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Trial Balance")

    data = payload
    start_row = _add_meta_rows(
        ws,
        "Trial Balance",
        [
            ("Currency", data.get("currency")),
            ("Date From", data.get("date_from")),
            ("Date To", data.get("date_to")),
            ("Total Accounts", data.get("total_accounts")),
            ("Total Debit", data.get("total_debit")),
            ("Total Credit", data.get("total_credit")),
        ],
    )

    headers = [
        "Account ID",
        "Code",
        "Name",
        "Type",
        "Is Group",
        "Total Debit",
        "Total Credit",
        "Net Debit",
        "Net Credit",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1
    for row in data.get("rows", []):
        ws.cell(row=current_row, column=1, value=row.get("account_id"))
        ws.cell(row=current_row, column=2, value=row.get("account_code"))
        ws.cell(row=current_row, column=3, value=row.get("account_name"))
        ws.cell(row=current_row, column=4, value=row.get("account_type"))
        ws.cell(row=current_row, column=5, value=str(row.get("is_group")))
        ws.cell(row=current_row, column=6, value=float(row.get("total_debit", 0) or 0))
        ws.cell(row=current_row, column=7, value=float(row.get("total_credit", 0) or 0))
        ws.cell(row=current_row, column=8, value=float(row.get("net_debit", 0) or 0))
        ws.cell(row=current_row, column=9, value=float(row.get("net_credit", 0) or 0))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "trial_balance.xlsx")


def _build_profit_loss_excel(payload: dict[str, Any]) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Profit Loss")

    data = payload
    start_row = _add_meta_rows(
        ws,
        "Profit & Loss",
        [
            ("Currency", data.get("currency")),
            ("Date From", data.get("date_from")),
            ("Date To", data.get("date_to")),
            ("Revenue Total", data.get("revenue", {}).get("total_amount")),
            ("Expenses Total", data.get("expenses", {}).get("total_amount")),
            ("Net Profit", data.get("net_profit")),
        ],
    )

    headers = [
        "Section",
        "Account ID",
        "Code",
        "Name",
        "Type",
        "Amount",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for row in data.get("revenue", {}).get("rows", []):
        ws.cell(row=current_row, column=1, value="Revenue")
        ws.cell(row=current_row, column=2, value=row.get("account_id"))
        ws.cell(row=current_row, column=3, value=row.get("account_code"))
        ws.cell(row=current_row, column=4, value=row.get("account_name"))
        ws.cell(row=current_row, column=5, value=row.get("account_type"))
        ws.cell(row=current_row, column=6, value=float(row.get("amount", 0) or 0))
        current_row += 1

    for row in data.get("expenses", {}).get("rows", []):
        ws.cell(row=current_row, column=1, value="Expense")
        ws.cell(row=current_row, column=2, value=row.get("account_id"))
        ws.cell(row=current_row, column=3, value=row.get("account_code"))
        ws.cell(row=current_row, column=4, value=row.get("account_name"))
        ws.cell(row=current_row, column=5, value=row.get("account_type"))
        ws.cell(row=current_row, column=6, value=float(row.get("amount", 0) or 0))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "profit_loss.xlsx")


def _build_balance_sheet_excel(payload: dict[str, Any]) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Balance Sheet")

    data = payload
    start_row = _add_meta_rows(
        ws,
        "Balance Sheet",
        [
            ("Currency", data.get("currency")),
            ("As Of Date", data.get("as_of_date")),
            ("Assets Total", data.get("assets", {}).get("total_amount")),
            ("Liabilities Total", data.get("liabilities", {}).get("total_amount")),
            ("Equity Total", data.get("equity", {}).get("total_amount")),
            ("Total Liabilities and Equity", data.get("total_liabilities_and_equity")),
            ("Is Balanced", data.get("is_balanced")),
        ],
    )

    headers = [
        "Section",
        "Account ID",
        "Code",
        "Name",
        "Type",
        "Amount",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for section_name, key in [
        ("Assets", "assets"),
        ("Liabilities", "liabilities"),
        ("Equity", "equity"),
    ]:
        for row in data.get(key, {}).get("rows", []):
            ws.cell(row=current_row, column=1, value=section_name)
            ws.cell(row=current_row, column=2, value=row.get("account_id"))
            ws.cell(row=current_row, column=3, value=row.get("account_code"))
            ws.cell(row=current_row, column=4, value=row.get("account_name"))
            ws.cell(row=current_row, column=5, value=row.get("account_type"))
            ws.cell(row=current_row, column=6, value=float(row.get("amount", 0) or 0))
            current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "balance_sheet.xlsx")


# ============================================================
# 📊 Trial Balance API
# ============================================================

@require_GET
def accounting_trial_balance_api(request):
    """
    إرجاع ميزان المراجعة.
    يدعم:
    - include_zero_accounts
    - posted_only
    - date_from
    - date_to
    """
    try:
        include_zero_accounts = _parse_bool(
            request.GET.get("include_zero_accounts"),
            default=False,
        )
        posted_only = _parse_bool(
            request.GET.get("posted_only"),
            default=True,
        )

        date_from = _parse_date(request.GET.get("date_from"), "date_from")
        date_to = _parse_date(request.GET.get("date_to"), "date_to")
        _validate_date_range(date_from, date_to)

        payload = build_trial_balance_payload(
            date_from=date_from,
            date_to=date_to,
            include_zero_accounts=include_zero_accounts,
            posted_only=posted_only,
        )
        return _success_response("trial_balance", payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to build trial balance payload: %s", exc)
        return _error_response(
            "تعذر تحميل ميزان المراجعة.",
            status=500,
        )


@require_GET
def accounting_trial_balance_excel_api(request):
    """
    تصدير ميزان المراجعة إلى Excel.
    """
    try:
        include_zero_accounts = _parse_bool(
            request.GET.get("include_zero_accounts"),
            default=False,
        )
        posted_only = _parse_bool(
            request.GET.get("posted_only"),
            default=True,
        )

        date_from = _parse_date(request.GET.get("date_from"), "date_from")
        date_to = _parse_date(request.GET.get("date_to"), "date_to")
        _validate_date_range(date_from, date_to)

        payload = build_trial_balance_payload(
            date_from=date_from,
            date_to=date_to,
            include_zero_accounts=include_zero_accounts,
            posted_only=posted_only,
        )
        return _build_trial_balance_excel(payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export trial balance excel: %s", exc)
        return _error_response(
            "تعذر تصدير ميزان المراجعة إلى Excel.",
            status=500,
        )


# ============================================================
# 📈 Profit & Loss API
# ============================================================

@require_GET
def accounting_profit_loss_api(request):
    """
    إرجاع تقرير الأرباح والخسائر.
    يدعم:
    - include_zero_accounts
    - posted_only
    - date_from
    - date_to
    """
    try:
        include_zero_accounts = _parse_bool(
            request.GET.get("include_zero_accounts"),
            default=False,
        )
        posted_only = _parse_bool(
            request.GET.get("posted_only"),
            default=True,
        )

        date_from = _parse_date(request.GET.get("date_from"), "date_from")
        date_to = _parse_date(request.GET.get("date_to"), "date_to")
        _validate_date_range(date_from, date_to)

        payload = build_profit_and_loss_payload(
            date_from=date_from,
            date_to=date_to,
            include_zero_accounts=include_zero_accounts,
            posted_only=posted_only,
        )
        return _success_response("profit_loss", payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to build profit & loss payload: %s", exc)
        return _error_response(
            "تعذر تحميل تقرير الأرباح والخسائر.",
            status=500,
        )


@require_GET
def accounting_profit_loss_excel_api(request):
    """
    تصدير الأرباح والخسائر إلى Excel.
    """
    try:
        include_zero_accounts = _parse_bool(
            request.GET.get("include_zero_accounts"),
            default=False,
        )
        posted_only = _parse_bool(
            request.GET.get("posted_only"),
            default=True,
        )

        date_from = _parse_date(request.GET.get("date_from"), "date_from")
        date_to = _parse_date(request.GET.get("date_to"), "date_to")
        _validate_date_range(date_from, date_to)

        payload = build_profit_and_loss_payload(
            date_from=date_from,
            date_to=date_to,
            include_zero_accounts=include_zero_accounts,
            posted_only=posted_only,
        )
        return _build_profit_loss_excel(payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export profit & loss excel: %s", exc)
        return _error_response(
            "تعذر تصدير الأرباح والخسائر إلى Excel.",
            status=500,
        )


# ============================================================
# 🏦 Balance Sheet API
# ============================================================

@require_GET
def accounting_balance_sheet_api(request):
    """
    إرجاع المركز المالي.
    يدعم:
    - include_zero_accounts
    - posted_only
    - include_current_year_earnings
    - as_of_date
    """
    try:
        include_zero_accounts = _parse_bool(
            request.GET.get("include_zero_accounts"),
            default=False,
        )
        posted_only = _parse_bool(
            request.GET.get("posted_only"),
            default=True,
        )
        include_current_year_earnings = _parse_bool(
            request.GET.get("include_current_year_earnings"),
            default=True,
        )

        as_of_date = _parse_date(request.GET.get("as_of_date"), "as_of_date")

        payload = build_balance_sheet_payload(
            as_of_date=as_of_date,
            include_zero_accounts=include_zero_accounts,
            posted_only=posted_only,
            include_current_year_earnings=include_current_year_earnings,
        )
        return _success_response("balance_sheet", payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to build balance sheet payload: %s", exc)
        return _error_response(
            "تعذر تحميل المركز المالي.",
            status=500,
        )


@require_GET
def accounting_balance_sheet_excel_api(request):
    """
    تصدير المركز المالي إلى Excel.
    """
    try:
        include_zero_accounts = _parse_bool(
            request.GET.get("include_zero_accounts"),
            default=False,
        )
        posted_only = _parse_bool(
            request.GET.get("posted_only"),
            default=True,
        )
        include_current_year_earnings = _parse_bool(
            request.GET.get("include_current_year_earnings"),
            default=True,
        )

        as_of_date = _parse_date(request.GET.get("as_of_date"), "as_of_date")

        payload = build_balance_sheet_payload(
            as_of_date=as_of_date,
            include_zero_accounts=include_zero_accounts,
            posted_only=posted_only,
            include_current_year_earnings=include_current_year_earnings,
        )
        return _build_balance_sheet_excel(payload)

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export balance sheet excel: %s", exc)
        return _error_response(
            "تعذر تصدير المركز المالي إلى Excel.",
            status=500,
        )
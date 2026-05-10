# ============================================================
# 📂 api/accounting/ledger.py
# 🧠 General Ledger API — Primey Care V2
# ------------------------------------------------------------
# ✅ دفتر الأستاذ العام
# ✅ متوافق مع Accounting Backend الجديد
# ✅ يدعم:
#    - account_id اختياري
#    - date_from / date_to
#    - posted_only
#    - include_opening
#    - include_metadata
#    - cost_center_id
#    - period_id
#    - posting_source
#    - source_type / source_id / source_number
#    - party_type / party_id
#    - page / page_size
#    - ordering
# ✅ Excel Export
# ✅ يرجع:
#    - الفلاتر
#    - بيانات الحساب
#    - الرصيد الافتتاحي
#    - إجماليات الفترة
#    - الرصيد الجاري running balance
#    - الحركات مرتبة محاسبيًا
#    - بيانات التصفح pagination
# ------------------------------------------------------------
# ملاحظة:
# - يعتمد على Account.nature وليس normal_balance
#   لأن موديل المحاسبة الرسمي يستخدم nature.
# ============================================================

from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from django.core.paginator import EmptyPage, Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import (
    Account,
    AccountingPeriod,
    CostCenter,
    JournalEntryLine,
    JournalEntryStatus,
    PostingSource,
)


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Constants
# ============================================================

DEFAULT_PAGE = 1
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 500

POSTED_REPORT_STATUSES = {
    JournalEntryStatus.POSTED,
    JournalEntryStatus.REVERSED,
}

ALLOWED_ORDERING = {
    "entry_date": "journal_entry__entry_date",
    "-entry_date": "-journal_entry__entry_date",
    "id": "id",
    "-id": "-id",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "account_code": "account__code",
    "-account_code": "-account__code",
    "debit_amount": "debit_amount",
    "-debit_amount": "-debit_amount",
    "credit_amount": "credit_amount",
    "-credit_amount": "-credit_amount",
    "tax_amount": "tax_amount",
    "-tax_amount": "-tax_amount",
    "source_number": "journal_entry__source_number",
    "-source_number": "-journal_entry__source_number",
    "entry_number": "journal_entry__entry_number",
    "-entry_number": "-journal_entry__entry_number",
}


# ============================================================
# 🔧 Helpers
# ============================================================

def _money(value: Decimal | int | float | str | None) -> Decimal:
    amount = Decimal(str(value or "0.00"))
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool = False) -> bool:
    if value in {None, ""}:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    return default


def _parse_int(value: str | None, field_name: str) -> int | None:
    if value in {None, ""}:
        return None

    raw_value = str(value).strip()

    try:
        parsed = int(raw_value)
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed <= 0:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر من صفر.")

    return parsed


def _parse_positive_int(
    value: str | None,
    field_name: str,
    *,
    default: int,
    min_value: int = 1,
    max_value: int | None = None,
) -> int:
    if value in {None, ""}:
        parsed = default
    else:
        raw_value = str(value).strip()

        try:
            parsed = int(raw_value)
        except ValueError as exc:
            raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed < min_value:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر أو تساوي {min_value}.")

    if max_value is not None and parsed > max_value:
        raise ValueError(f"قيمة {field_name} يجب ألا تتجاوز {max_value}.")

    return parsed


def _parse_date(value: str | None, field_name: str) -> date | None:
    if not value:
        return None

    raw_value = str(value).strip()

    try:
        return date.fromisoformat(raw_value)
    except ValueError as exc:
        raise ValueError(
            f"قيمة {field_name} غير صحيحة. استخدم الصيغة YYYY-MM-DD."
        ) from exc


def _parse_ordering(value: str | None, default: str = "entry_date") -> str:
    raw_value = (value or "").strip() or default

    if raw_value not in ALLOWED_ORDERING:
        allowed = ", ".join(ALLOWED_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return raw_value


def _validate_date_range(date_from: date | None, date_to: date | None) -> None:
    if date_from and date_to and date_from > date_to:
        raise ValueError("لا يمكن أن يكون date_from أكبر من date_to.")


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any]) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=200,
        json_dumps_params={"ensure_ascii": False},
    )


def _safe_attr(obj: Any, attr_name: str, default: Any = None) -> Any:
    try:
        return getattr(obj, attr_name, default)
    except Exception:
        return default


def _call_display(obj: Any, method_name: str, default: Any = None) -> Any:
    try:
        method = getattr(obj, method_name, None)
        if callable(method):
            return method()
    except Exception:
        return default

    return default


def _iso_datetime(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return None


def _iso_date(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return None


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_meta_rows(
    ws,
    title: str,
    meta_rows: list[tuple[str, Any]],
    *,
    merge_to_column: int = 12,
) -> int:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_to_column)

    current_row = 3

    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(
            row=current_row,
            column=2,
            value=str(value) if value is not None else "",
        )
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    return current_row + 1


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# 🧾 Serializers
# ============================================================

def _resolve_account_name(account: Account | None) -> str | None:
    if not account:
        return None

    return (
        _safe_attr(account, "name", None)
        or _safe_attr(account, "name_ar", None)
        or _safe_attr(account, "name_en", None)
        or f"Account #{account.pk}"
    )


def _resolve_account_nature(account: Account | None) -> str:
    if not account:
        return "DEBIT"

    return str(_safe_attr(account, "nature", "") or "DEBIT").upper()


def _resolve_account_nature_label(account: Account | None) -> str | None:
    if not account:
        return None

    if hasattr(account, "get_nature_display"):
        return account.get_nature_display()

    nature = _resolve_account_nature(account)

    if nature == "DEBIT":
        return "مدين"

    if nature == "CREDIT":
        return "دائن"

    return None


def _account_signed_delta(
    *,
    account_nature: str | None,
    debit_amount: Decimal,
    credit_amount: Decimal,
) -> Decimal:
    if str(account_nature or "").upper() == "CREDIT":
        return _money(credit_amount - debit_amount)

    return _money(debit_amount - credit_amount)


def _build_order_by(ordering_key: str) -> list[str]:
    field_name = ALLOWED_ORDERING[ordering_key]

    if ordering_key.startswith("-"):
        return [field_name, "-journal_entry__id", "-sort_order", "-id"]

    return [field_name, "journal_entry__id", "sort_order", "id"]


def _serialize_account(account: Account | None) -> dict[str, Any] | None:
    if not account:
        return None

    return {
        "id": account.id,
        "code": _safe_attr(account, "code", None),
        "name": _resolve_account_name(account),
        "name_ar": _safe_attr(account, "name", None),
        "name_en": _safe_attr(account, "name_en", None),
        "account_type": _safe_attr(account, "account_type", None),
        "account_type_label": _call_display(
            account,
            "get_account_type_display",
            _safe_attr(account, "account_type", None),
        ),
        "nature": _resolve_account_nature(account),
        "nature_label": _resolve_account_nature_label(account),
        "is_group": bool(_safe_attr(account, "is_group", False)),
        "is_active": bool(_safe_attr(account, "is_active", True)),
        "allow_manual_posting": bool(_safe_attr(account, "allow_manual_posting", True)),
        "is_system": bool(_safe_attr(account, "is_system", False)),
        "parent_id": _safe_attr(account, "parent_id", None),
        "level": int(_safe_attr(account, "level", 1) or 1),
        "currency": _safe_attr(account, "currency", "SAR"),
        "opening_balance": _money(_safe_attr(account, "opening_balance", "0.00")),
        "can_post": bool(
            _safe_attr(account, "is_active", True)
            and not _safe_attr(account, "is_group", False)
        ),
    }


def _serialize_period(period: AccountingPeriod | None) -> dict[str, Any] | None:
    if not period:
        return None

    fiscal_year = _safe_attr(period, "fiscal_year", None)

    return {
        "id": period.id,
        "name": period.name,
        "start_date": _iso_date(period.start_date),
        "end_date": _iso_date(period.end_date),
        "status": period.status,
        "status_label": _call_display(period, "get_status_display", period.status),
        "fiscal_year": {
            "id": _safe_attr(fiscal_year, "id", None),
            "name": _safe_attr(fiscal_year, "name", None),
            "status": _safe_attr(fiscal_year, "status", None),
        } if fiscal_year else None,
    }


def _serialize_cost_center(cost_center: CostCenter | None) -> dict[str, Any] | None:
    if not cost_center:
        return None

    return {
        "id": cost_center.id,
        "code": cost_center.code,
        "name": cost_center.name,
        "name_en": cost_center.name_en,
        "level": cost_center.level,
        "is_group": cost_center.is_group,
        "status": cost_center.status,
        "status_label": _call_display(cost_center, "get_status_display", cost_center.status),
    }


def _serialize_tax_rate(tax_rate) -> dict[str, Any] | None:
    if not tax_rate:
        return None

    return {
        "id": _safe_attr(tax_rate, "id", None),
        "code": _safe_attr(tax_rate, "code", None),
        "name": _safe_attr(tax_rate, "name", None),
        "tax_type": _safe_attr(tax_rate, "tax_type", None),
        "tax_type_label": _call_display(tax_rate, "get_tax_type_display", _safe_attr(tax_rate, "tax_type", None)),
        "rate": _money(_safe_attr(tax_rate, "rate", "0.00")),
        "is_active": bool(_safe_attr(tax_rate, "is_active", True)),
        "is_default": bool(_safe_attr(tax_rate, "is_default", False)),
    }


def _serialize_line(
    line: JournalEntryLine,
    *,
    selected_account: Account | None = None,
    running_balance: Decimal,
    include_metadata: bool = False,
) -> dict[str, Any]:
    account = _safe_attr(line, "account", None)
    journal_entry = _safe_attr(line, "journal_entry", None)

    debit_amount = _money(_safe_attr(line, "debit_amount", "0.00"))
    credit_amount = _money(_safe_attr(line, "credit_amount", "0.00"))
    tax_amount = _money(_safe_attr(line, "tax_amount", "0.00"))

    account_for_delta = selected_account or account
    delta = _account_signed_delta(
        account_nature=_resolve_account_nature(account_for_delta),
        debit_amount=debit_amount,
        credit_amount=credit_amount,
    )

    payload = {
        "id": line.id,
        "journal_entry_id": _safe_attr(journal_entry, "id", None),
        "journal_entry_number": _safe_attr(journal_entry, "entry_number", None),
        "entry_date": _iso_date(_safe_attr(journal_entry, "entry_date", None)),
        "entry_status": _safe_attr(journal_entry, "status", None),
        "entry_status_label": _call_display(
            journal_entry,
            "get_status_display",
            _safe_attr(journal_entry, "status", None),
        ),
        "posting_source": _safe_attr(journal_entry, "posting_source", None),
        "posting_source_label": _call_display(
            journal_entry,
            "get_posting_source_display",
            _safe_attr(journal_entry, "posting_source", None),
        ),
        "reference": _safe_attr(journal_entry, "reference", None),
        "external_reference": _safe_attr(journal_entry, "external_reference", None),
        "source_type": _safe_attr(journal_entry, "source_type", ""),
        "source_id": _safe_attr(journal_entry, "source_id", ""),
        "source_number": _safe_attr(journal_entry, "source_number", ""),
        "entry_description": _safe_attr(journal_entry, "description", None),
        "period": _serialize_period(_safe_attr(journal_entry, "period", None)),
        "period_id": _safe_attr(journal_entry, "period_id", None),
        "account_id": _safe_attr(account, "id", None),
        "account_code": _safe_attr(account, "code", None),
        "account_name": _resolve_account_name(account),
        "account_type": _safe_attr(account, "account_type", None),
        "account_type_label": _call_display(
            account,
            "get_account_type_display",
            _safe_attr(account, "account_type", None),
        ),
        "nature": _resolve_account_nature(account),
        "nature_label": _resolve_account_nature_label(account),
        "account": _serialize_account(account),
        "cost_center": _serialize_cost_center(_safe_attr(line, "cost_center", None)),
        "cost_center_id": _safe_attr(line, "cost_center_id", None),
        "tax_rate": _serialize_tax_rate(_safe_attr(line, "tax_rate", None)),
        "tax_rate_id": _safe_attr(line, "tax_rate_id", None),
        "line_description": _safe_attr(line, "description", None),
        "debit_amount": debit_amount,
        "credit_amount": credit_amount,
        "tax_amount": tax_amount,
        "movement_amount": delta,
        "running_balance": running_balance,
        "party_type": _safe_attr(line, "party_type", ""),
        "party_id": _safe_attr(line, "party_id", ""),
        "source_line_id": _safe_attr(line, "source_line_id", ""),
        "sort_order": _safe_attr(line, "sort_order", 0),
        "created_at": _iso_datetime(_safe_attr(line, "created_at", None)),
        "updated_at": _iso_datetime(_safe_attr(line, "updated_at", None)),
    }

    if include_metadata:
        payload["metadata"] = _safe_attr(line, "metadata", {}) or {}
        payload["entry_metadata"] = _safe_attr(journal_entry, "metadata", {}) or {}

    return payload


# ============================================================
# 📘 Query Builder
# ============================================================

def _build_ledger_payload(request) -> dict[str, Any]:
    account_id = _parse_int(request.GET.get("account_id"), "account_id")
    date_from = _parse_date(request.GET.get("date_from"), "date_from")
    date_to = _parse_date(request.GET.get("date_to"), "date_to")
    posted_only = _parse_bool(request.GET.get("posted_only"), default=True)
    include_opening = _parse_bool(request.GET.get("include_opening"), default=True)
    include_metadata = _parse_bool(request.GET.get("include_metadata"), default=False)
    ordering_key = _parse_ordering(request.GET.get("ordering"), default="entry_date")

    cost_center_id = _parse_int(request.GET.get("cost_center_id"), "cost_center_id")
    period_id = _parse_int(request.GET.get("period_id"), "period_id")

    posting_source = (request.GET.get("posting_source") or "").strip()
    source_type = (request.GET.get("source_type") or "").strip()
    source_id = (request.GET.get("source_id") or "").strip()
    source_number = (request.GET.get("source_number") or "").strip()
    party_type = (request.GET.get("party_type") or "").strip()
    party_id = (request.GET.get("party_id") or "").strip()
    search = (request.GET.get("search") or "").strip()

    _validate_date_range(date_from, date_to)

    if posting_source and posting_source not in {choice[0] for choice in PostingSource.choices}:
        raise ValueError("مصدر القيد غير صحيح.")

    selected_account: Account | None = None
    selected_cost_center: CostCenter | None = None
    selected_period: AccountingPeriod | None = None

    if account_id:
        selected_account = (
            Account.objects.select_related("parent")
            .filter(id=account_id)
            .first()
        )

        if not selected_account:
            raise ValueError("الحساب المطلوب غير موجود.")

    if cost_center_id:
        selected_cost_center = CostCenter.objects.filter(id=cost_center_id).first()

        if not selected_cost_center:
            raise ValueError("مركز التكلفة المطلوب غير موجود.")

    if period_id:
        selected_period = AccountingPeriod.objects.filter(id=period_id).first()

        if not selected_period:
            raise ValueError("الفترة المحاسبية المطلوبة غير موجودة.")

    base_qs = JournalEntryLine.objects.select_related(
        "journal_entry",
        "journal_entry__period",
        "journal_entry__period__fiscal_year",
        "account",
        "cost_center",
        "tax_rate",
    )

    if posted_only:
        base_qs = base_qs.filter(journal_entry__status__in=POSTED_REPORT_STATUSES)

    if selected_account:
        base_qs = base_qs.filter(account_id=selected_account.id)

    if selected_cost_center:
        base_qs = base_qs.filter(cost_center_id=selected_cost_center.id)

    if selected_period:
        base_qs = base_qs.filter(journal_entry__period_id=selected_period.id)

    if posting_source:
        base_qs = base_qs.filter(journal_entry__posting_source=posting_source)

    if source_type:
        base_qs = base_qs.filter(journal_entry__source_type__icontains=source_type)

    if source_id:
        base_qs = base_qs.filter(journal_entry__source_id=source_id)

    if source_number:
        base_qs = base_qs.filter(journal_entry__source_number__icontains=source_number)

    if party_type:
        base_qs = base_qs.filter(party_type__icontains=party_type)

    if party_id:
        base_qs = base_qs.filter(party_id=party_id)

    if search:
        base_qs = base_qs.filter(
            Q(journal_entry__entry_number__icontains=search)
            | Q(journal_entry__reference__icontains=search)
            | Q(journal_entry__external_reference__icontains=search)
            | Q(journal_entry__source_type__icontains=search)
            | Q(journal_entry__source_id__icontains=search)
            | Q(journal_entry__source_number__icontains=search)
            | Q(journal_entry__description__icontains=search)
            | Q(description__icontains=search)
            | Q(account__code__icontains=search)
            | Q(account__name__icontains=search)
            | Q(account__name_en__icontains=search)
            | Q(cost_center__code__icontains=search)
            | Q(cost_center__name__icontains=search)
            | Q(cost_center__name_en__icontains=search)
            | Q(party_type__icontains=search)
            | Q(party_id__icontains=search)
        )

    opening_qs = base_qs

    if date_from:
        opening_qs = opening_qs.filter(journal_entry__entry_date__lt=date_from)
    else:
        opening_qs = base_qs.none()

    period_qs = base_qs

    if date_from:
        period_qs = period_qs.filter(journal_entry__entry_date__gte=date_from)

    if date_to:
        period_qs = period_qs.filter(journal_entry__entry_date__lte=date_to)

    period_lines = list(period_qs.order_by(*_build_order_by(ordering_key)))

    opening_debit = _money(
        sum(
            (
                _money(_safe_attr(line, "debit_amount", "0.00"))
                for line in opening_qs
            ),
            Decimal("0.00"),
        )
    )
    opening_credit = _money(
        sum(
            (
                _money(_safe_attr(line, "credit_amount", "0.00"))
                for line in opening_qs
            ),
            Decimal("0.00"),
        )
    )

    opening_balance = Decimal("0.00")

    if include_opening:
        if selected_account:
            opening_balance = _account_signed_delta(
                account_nature=_resolve_account_nature(selected_account),
                debit_amount=opening_debit,
                credit_amount=opening_credit,
            )
        else:
            opening_balance = _money(opening_debit - opening_credit)

    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    total_tax = Decimal("0.00")
    running_balance = _money(opening_balance)

    transactions_all: list[dict[str, Any]] = []

    for line in period_lines:
        account = _safe_attr(line, "account", None)

        debit_amount = _money(_safe_attr(line, "debit_amount", "0.00"))
        credit_amount = _money(_safe_attr(line, "credit_amount", "0.00"))
        tax_amount = _money(_safe_attr(line, "tax_amount", "0.00"))

        total_debit += debit_amount
        total_credit += credit_amount
        total_tax += tax_amount

        account_for_delta = selected_account or account
        delta = _account_signed_delta(
            account_nature=_resolve_account_nature(account_for_delta),
            debit_amount=debit_amount,
            credit_amount=credit_amount,
        )

        running_balance = _money(running_balance + delta)

        transactions_all.append(
            _serialize_line(
                line,
                selected_account=selected_account,
                running_balance=running_balance,
                include_metadata=include_metadata,
            )
        )

    total_debit = _money(total_debit)
    total_credit = _money(total_credit)
    total_tax = _money(total_tax)
    closing_balance = _money(running_balance)

    return {
        "filters": {
            "account_id": selected_account.id if selected_account else None,
            "cost_center_id": selected_cost_center.id if selected_cost_center else None,
            "period_id": selected_period.id if selected_period else None,
            "date_from": date_from.isoformat() if date_from else None,
            "date_to": date_to.isoformat() if date_to else None,
            "posted_only": posted_only,
            "include_opening": include_opening,
            "include_metadata": include_metadata,
            "posting_source": posting_source or None,
            "source_type": source_type or None,
            "source_id": source_id or None,
            "source_number": source_number or None,
            "party_type": party_type or None,
            "party_id": party_id or None,
            "search": search or None,
            "ordering": ordering_key,
        },
        "account": _serialize_account(selected_account),
        "cost_center": _serialize_cost_center(selected_cost_center),
        "period": _serialize_period(selected_period),
        "summary": {
            "transaction_count": len(transactions_all),
            "opening_debit": opening_debit,
            "opening_credit": opening_credit,
            "opening_balance": opening_balance if include_opening else Decimal("0.00"),
            "total_debit": total_debit,
            "total_credit": total_credit,
            "total_tax": total_tax,
            "net_movement": _money(total_debit - total_credit),
            "closing_balance": closing_balance,
            "is_period_balanced": total_debit == total_credit,
        },
        "transactions_all": transactions_all,
    }


# ============================================================
# 📊 Excel
# ============================================================

def _build_ledger_excel(payload: dict[str, Any]) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Ledger")

    filters = payload["filters"]
    account = payload["account"]
    cost_center = payload["cost_center"]
    period = payload["period"]
    summary = payload["summary"]

    start_row = _add_meta_rows(
        ws,
        "General Ledger",
        [
            ("Account ID", filters.get("account_id")),
            ("Account Code", account.get("code") if account else None),
            ("Account Name", account.get("name") if account else None),
            ("Cost Center ID", filters.get("cost_center_id")),
            ("Cost Center", cost_center.get("name") if cost_center else None),
            ("Period ID", filters.get("period_id")),
            ("Period", period.get("name") if period else None),
            ("Date From", filters.get("date_from")),
            ("Date To", filters.get("date_to")),
            ("Posted Only", filters.get("posted_only")),
            ("Include Opening", filters.get("include_opening")),
            ("Posting Source", filters.get("posting_source")),
            ("Source Type", filters.get("source_type")),
            ("Source ID", filters.get("source_id")),
            ("Source Number", filters.get("source_number")),
            ("Party Type", filters.get("party_type")),
            ("Party ID", filters.get("party_id")),
            ("Search", filters.get("search")),
            ("Ordering", filters.get("ordering")),
            ("Opening Balance", summary.get("opening_balance")),
            ("Total Debit", summary.get("total_debit")),
            ("Total Credit", summary.get("total_credit")),
            ("Total Tax", summary.get("total_tax")),
            ("Closing Balance", summary.get("closing_balance")),
            ("Transaction Count", summary.get("transaction_count")),
        ],
        merge_to_column=30,
    )

    headers = [
        "ID",
        "Journal Entry ID",
        "Journal Entry Number",
        "Entry Date",
        "Entry Status",
        "Posting Source",
        "Reference",
        "External Reference",
        "Source Type",
        "Source ID",
        "Source Number",
        "Period",
        "Entry Description",
        "Account ID",
        "Account Code",
        "Account Name",
        "Account Type",
        "Nature",
        "Cost Center",
        "Tax Rate",
        "Line Description",
        "Debit Amount",
        "Credit Amount",
        "Tax Amount",
        "Movement Amount",
        "Running Balance",
        "Party Type",
        "Party ID",
        "Source Line ID",
        "Sort Order",
        "Created At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for row in payload["transactions_all"]:
        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("journal_entry_id"))
        ws.cell(row=current_row, column=3, value=row.get("journal_entry_number"))
        ws.cell(row=current_row, column=4, value=row.get("entry_date"))
        ws.cell(row=current_row, column=5, value=row.get("entry_status"))
        ws.cell(row=current_row, column=6, value=row.get("posting_source"))
        ws.cell(row=current_row, column=7, value=row.get("reference"))
        ws.cell(row=current_row, column=8, value=row.get("external_reference"))
        ws.cell(row=current_row, column=9, value=row.get("source_type"))
        ws.cell(row=current_row, column=10, value=row.get("source_id"))
        ws.cell(row=current_row, column=11, value=row.get("source_number"))
        ws.cell(
            row=current_row,
            column=12,
            value=(row.get("period") or {}).get("name") if row.get("period") else "",
        )
        ws.cell(row=current_row, column=13, value=row.get("entry_description"))
        ws.cell(row=current_row, column=14, value=row.get("account_id"))
        ws.cell(row=current_row, column=15, value=row.get("account_code"))
        ws.cell(row=current_row, column=16, value=row.get("account_name"))
        ws.cell(row=current_row, column=17, value=row.get("account_type"))
        ws.cell(row=current_row, column=18, value=row.get("nature"))
        ws.cell(
            row=current_row,
            column=19,
            value=(row.get("cost_center") or {}).get("name") if row.get("cost_center") else "",
        )
        ws.cell(
            row=current_row,
            column=20,
            value=(row.get("tax_rate") or {}).get("code") if row.get("tax_rate") else "",
        )
        ws.cell(row=current_row, column=21, value=row.get("line_description"))
        ws.cell(row=current_row, column=22, value=float(row.get("debit_amount", 0) or 0))
        ws.cell(row=current_row, column=23, value=float(row.get("credit_amount", 0) or 0))
        ws.cell(row=current_row, column=24, value=float(row.get("tax_amount", 0) or 0))
        ws.cell(row=current_row, column=25, value=float(row.get("movement_amount", 0) or 0))
        ws.cell(row=current_row, column=26, value=float(row.get("running_balance", 0) or 0))
        ws.cell(row=current_row, column=27, value=row.get("party_type"))
        ws.cell(row=current_row, column=28, value=row.get("party_id"))
        ws.cell(row=current_row, column=29, value=row.get("source_line_id"))
        ws.cell(row=current_row, column=30, value=row.get("sort_order"))
        ws.cell(row=current_row, column=31, value=row.get("created_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "ledger.xlsx")


# ============================================================
# 📘 General Ledger API
# ============================================================

@require_GET
def accounting_general_ledger_api(request):
    """
    دفتر الأستاذ العام.
    """
    try:
        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )

        base_payload = _build_ledger_payload(request)
        transactions_all = base_payload["transactions_all"]

        paginator = Paginator(transactions_all, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        payload = {
            "filters": base_payload["filters"],
            "account": base_payload["account"],
            "cost_center": base_payload["cost_center"],
            "period": base_payload["period"],
            "summary": base_payload["summary"],
            "pagination": {
                "page": page_obj.number,
                "page_size": page_size,
                "total_pages": paginator.num_pages,
                "total_items": paginator.count,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
                "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
            },
            "transactions": list(page_obj.object_list),
        }

        return _success_response(payload)

    except ValueError as exc:
        not_found_messages = {
            "الحساب المطلوب غير موجود.",
            "مركز التكلفة المطلوب غير موجود.",
            "الفترة المحاسبية المطلوبة غير موجودة.",
        }
        status = 404 if str(exc) in not_found_messages else 400
        return _error_response(str(exc), status=status)

    except Exception as exc:
        logger.exception("Failed to load general ledger: %s", exc)
        return _error_response(
            "تعذر تحميل دفتر الأستاذ العام.",
            status=500,
        )


@require_GET
def accounting_general_ledger_excel_api(request):
    """
    تصدير دفتر الأستاذ العام إلى Excel.
    """
    try:
        payload = _build_ledger_payload(request)
        return _build_ledger_excel(payload)

    except ValueError as exc:
        not_found_messages = {
            "الحساب المطلوب غير موجود.",
            "مركز التكلفة المطلوب غير موجود.",
            "الفترة المحاسبية المطلوبة غير موجودة.",
        }
        status = 404 if str(exc) in not_found_messages else 400
        return _error_response(str(exc), status=status)

    except Exception as exc:
        logger.exception("Failed to export general ledger excel: %s", exc)
        return _error_response(
            "تعذر تصدير دفتر الأستاذ العام إلى Excel.",
            status=500,
        )
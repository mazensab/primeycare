# ============================================================
# 📂 api/accounting/taxes.py
# 🧠 Accounting Taxes API — Primey Care V2
# ------------------------------------------------------------
# ✅ إدارة نسب الضرائب Tax Rates
# ✅ إدارة الحركات الضريبية Tax Transactions
# ✅ قائمة / تفاصيل / إنشاء / تحديث
# ✅ تفعيل / تعطيل / تعيين ضريبة افتراضية
# ✅ Excel Export
# ✅ متوافق مع Accounting Backend الجديد
# ============================================================

from __future__ import annotations

import json
import logging
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from accounting.models import (
    Account,
    JournalEntry,
    JournalEntryLine,
    TaxDirection,
    TaxRate,
    TaxTransaction,
    TaxType,
)


logger = logging.getLogger(__name__)


# ============================================================
# 🔧 Constants
# ============================================================

DEFAULT_PAGE = 1
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 500

ALLOWED_TAX_RATE_ORDERING = {
    "code": "code",
    "-code": "-code",
    "name": "name",
    "-name": "-name",
    "tax_type": "tax_type",
    "-tax_type": "-tax_type",
    "rate": "rate",
    "-rate": "-rate",
    "is_active": "is_active",
    "-is_active": "-is_active",
    "is_default": "is_default",
    "-is_default": "-is_default",
    "created_at": "created_at",
    "-created_at": "-created_at",
    "updated_at": "updated_at",
    "-updated_at": "-updated_at",
}

ALLOWED_TAX_TRANSACTION_ORDERING = {
    "transaction_date": "transaction_date",
    "-transaction_date": "-transaction_date",
    "id": "id",
    "-id": "-id",
    "tax_amount": "tax_amount",
    "-tax_amount": "-tax_amount",
    "taxable_amount": "taxable_amount",
    "-taxable_amount": "-taxable_amount",
    "direction": "direction",
    "-direction": "-direction",
    "source_type": "source_type",
    "-source_type": "-source_type",
    "source_number": "source_number",
    "-source_number": "-source_number",
    "created_at": "created_at",
    "-created_at": "-created_at",
}


# ============================================================
# 🔧 Helpers
# ============================================================

def _money(value: Decimal | int | float | str | None) -> Decimal:
    amount = Decimal(str(value or "0.00"))
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_to_string(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, dict):
        return {key: _decimal_to_string(val) for key, val in value.items()}

    if isinstance(value, list):
        return [_decimal_to_string(item) for item in value]

    if isinstance(value, tuple):
        return tuple(_decimal_to_string(item) for item in value)

    return value


def _parse_bool(value: str | None, default: bool | None = None) -> bool | None:
    if value in {None, ""}:
        return default

    raw = str(value).strip().lower()

    if raw in {"1", "true", "yes", "on"}:
        return True

    if raw in {"0", "false", "no", "off"}:
        return False

    raise ValueError("قيمة boolean غير صحيحة. استخدم true أو false.")


def _parse_int(value: str | None, field_name: str) -> int | None:
    if value in {None, ""}:
        return None

    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed <= 0:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر من صفر.")

    return parsed


def _parse_positive_int(
    value: str | None,
    field_name: str,
    *,
    default: int,
    min_value: int = 1,
    max_value: int | None = None,
) -> int:
    if value in {None, ""}:
        parsed = default
    else:
        try:
            parsed = int(str(value).strip())
        except ValueError as exc:
            raise ValueError(f"قيمة {field_name} يجب أن تكون رقمًا صحيحًا.") from exc

    if parsed < min_value:
        raise ValueError(f"قيمة {field_name} يجب أن تكون أكبر أو تساوي {min_value}.")

    if max_value is not None and parsed > max_value:
        raise ValueError(f"قيمة {field_name} يجب ألا تتجاوز {max_value}.")

    return parsed


def _parse_date(value: Any, field_name: str) -> date | None:
    if value in {None, ""}:
        return None

    if isinstance(value, date):
        return value

    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"قيمة {field_name} غير صحيحة. استخدم الصيغة YYYY-MM-DD.") from exc


def _validate_date_range(date_from: date | None, date_to: date | None) -> None:
    if date_from and date_to and date_from > date_to:
        raise ValueError("لا يمكن أن يكون date_from أكبر من date_to.")


def _json_body(request) -> dict[str, Any]:
    try:
        body = request.body.decode("utf-8") if request.body else "{}"
        data = json.loads(body or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError("صيغة JSON غير صحيحة.") from exc

    if not isinstance(data, dict):
        raise ValueError("جسم الطلب يجب أن يكون JSON Object.")

    return data


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _error_response(
    message: str,
    status: int = 400,
    extra: dict | None = None,
) -> JsonResponse:
    payload = {
        "ok": False,
        "message": message,
    }

    if extra:
        payload.update(extra)

    return JsonResponse(
        payload,
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _success_response(data: dict[str, Any], status: int = 200) -> JsonResponse:
    return JsonResponse(
        {
            "ok": True,
            "data": _decimal_to_string(data),
        },
        status=status,
        json_dumps_params={"ensure_ascii": False},
    )


def _validation_error_payload(exc: ValidationError) -> dict[str, Any]:
    if hasattr(exc, "message_dict"):
        return {
            "ok": False,
            "message": "تعذر حفظ البيانات.",
            "errors": exc.message_dict,
        }

    return {
        "ok": False,
        "message": "تعذر حفظ البيانات.",
        "errors": exc.messages if hasattr(exc, "messages") else [str(exc)],
    }


def _status_label(value: str, choices) -> str:
    for choice_value, label in choices:
        if choice_value == value:
            return label
    return value


def _safe_attr(obj: Any, attr_name: str, default: Any = None) -> Any:
    try:
        return getattr(obj, attr_name, default)
    except Exception:
        return default


def _iso_date(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return None


def _iso_datetime(value: Any) -> str | None:
    if not value:
        return None

    try:
        return value.isoformat()
    except Exception:
        return None


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    clean = title

    for char in invalid_chars:
        clean = clean.replace(char, "-")

    return clean[:31]


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        first_real_cell = None
        max_length = 0

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                first_real_cell = cell
                break

        if first_real_cell is None:
            continue

        column_letter = first_real_cell.column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                continue

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_excel_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# 🧾 Account Serializers
# ============================================================

def _resolve_account(account_id: Any, field_name: str) -> Account | None:
    if account_id in {None, ""}:
        return None

    try:
        parsed_id = int(account_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({field_name: "معرف الحساب غير صحيح."}) from exc

    account = Account.objects.filter(id=parsed_id).first()

    if not account:
        raise ValidationError({field_name: "الحساب المطلوب غير موجود."})

    if account.is_group:
        raise ValidationError({field_name: "لا يمكن استخدام حساب تجميعي."})

    if not account.is_active:
        raise ValidationError({field_name: "لا يمكن استخدام حساب غير نشط."})

    return account


def _serialize_account(account: Account | None) -> dict[str, Any] | None:
    if not account:
        return None

    return {
        "id": account.id,
        "code": account.code,
        "name": account.name,
        "name_en": account.name_en,
        "account_type": account.account_type,
        "nature": account.nature,
        "is_group": account.is_group,
        "is_active": account.is_active,
        "currency": account.currency,
    }


def _serialize_journal_entry(entry: JournalEntry | None) -> dict[str, Any] | None:
    if not entry:
        return None

    return {
        "id": entry.id,
        "entry_number": entry.entry_number,
        "entry_date": _iso_date(entry.entry_date),
        "status": entry.status,
        "posting_source": entry.posting_source,
        "reference": entry.reference,
        "source_type": entry.source_type,
        "source_id": entry.source_id,
        "source_number": entry.source_number,
        "total_debit": entry.total_debit,
        "total_credit": entry.total_credit,
        "currency": entry.currency,
    }


def _serialize_journal_line(line: JournalEntryLine | None) -> dict[str, Any] | None:
    if not line:
        return None

    return {
        "id": line.id,
        "journal_entry_id": line.journal_entry_id,
        "account": _serialize_account(line.account),
        "description": line.description,
        "debit_amount": line.debit_amount,
        "credit_amount": line.credit_amount,
        "tax_amount": line.tax_amount,
        "party_type": line.party_type,
        "party_id": line.party_id,
        "source_line_id": line.source_line_id,
    }


# ============================================================
# 🧾 Tax Rate Serializers
# ============================================================

def _serialize_tax_rate(tax_rate: TaxRate) -> dict[str, Any]:
    return {
        "id": tax_rate.id,
        "code": tax_rate.code,
        "name": tax_rate.name,
        "tax_type": tax_rate.tax_type,
        "tax_type_label": _status_label(tax_rate.tax_type, TaxType.choices),
        "rate": tax_rate.rate,
        "sales_account_id": tax_rate.sales_account_id,
        "sales_account": _serialize_account(tax_rate.sales_account),
        "purchase_account_id": tax_rate.purchase_account_id,
        "purchase_account": _serialize_account(tax_rate.purchase_account),
        "is_active": tax_rate.is_active,
        "is_default": tax_rate.is_default,
        "description": tax_rate.description,
        "created_at": _iso_datetime(tax_rate.created_at),
        "updated_at": _iso_datetime(tax_rate.updated_at),
    }


def _build_tax_rates_queryset(request):
    search = (request.GET.get("search") or "").strip()
    tax_type = (request.GET.get("tax_type") or "").strip()
    is_active = _parse_bool(request.GET.get("is_active"), default=None)
    is_default = _parse_bool(request.GET.get("is_default"), default=None)

    qs = TaxRate.objects.select_related("sales_account", "purchase_account").all()

    if search:
        qs = qs.filter(
            Q(code__icontains=search)
            | Q(name__icontains=search)
            | Q(description__icontains=search)
            | Q(sales_account__code__icontains=search)
            | Q(sales_account__name__icontains=search)
            | Q(purchase_account__code__icontains=search)
            | Q(purchase_account__name__icontains=search)
        )

    if tax_type:
        valid_types = {choice[0] for choice in TaxType.choices}
        if tax_type not in valid_types:
            raise ValueError("نوع الضريبة غير صحيح.")
        qs = qs.filter(tax_type=tax_type)

    if is_active is not None:
        qs = qs.filter(is_active=is_active)

    if is_default is not None:
        qs = qs.filter(is_default=is_default)

    return {
        "qs": qs,
        "filters": {
            "search": search or None,
            "tax_type": tax_type or None,
            "is_active": is_active,
            "is_default": is_default,
        },
    }


def _parse_tax_rate_ordering(value: str | None, default: str = "code") -> str:
    raw = (value or "").strip() or default

    if raw not in ALLOWED_TAX_RATE_ORDERING:
        allowed = ", ".join(ALLOWED_TAX_RATE_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_TAX_RATE_ORDERING[raw]


def _build_tax_rates_summary(tax_rates: list[TaxRate]) -> dict[str, Any]:
    total = len(tax_rates)
    active = sum(1 for item in tax_rates if item.is_active)
    inactive = total - active
    defaults = sum(1 for item in tax_rates if item.is_default)

    by_type: dict[str, int] = {}

    for item in tax_rates:
        by_type[item.tax_type] = by_type.get(item.tax_type, 0) + 1

    return {
        "total_tax_rates": total,
        "active_tax_rates": active,
        "inactive_tax_rates": inactive,
        "default_tax_rates": defaults,
        "by_type": by_type,
    }


def _apply_payload_to_tax_rate(tax_rate: TaxRate, data: dict[str, Any]) -> TaxRate:
    if "code" in data:
        tax_rate.code = _clean_text(data.get("code"))

    if "name" in data:
        tax_rate.name = _clean_text(data.get("name"))

    if "tax_type" in data:
        tax_type = _clean_text(data.get("tax_type"))
        valid_types = {choice[0] for choice in TaxType.choices}
        if tax_type not in valid_types:
            raise ValidationError({"tax_type": "نوع الضريبة غير صحيح."})
        tax_rate.tax_type = tax_type

    if "rate" in data:
        tax_rate.rate = _money(data.get("rate"))

    if "sales_account_id" in data:
        tax_rate.sales_account = _resolve_account(data.get("sales_account_id"), "sales_account_id")

    if "purchase_account_id" in data:
        tax_rate.purchase_account = _resolve_account(data.get("purchase_account_id"), "purchase_account_id")

    if "is_active" in data:
        is_active = data.get("is_active")
        if not isinstance(is_active, bool):
            raise ValidationError({"is_active": "قيمة is_active يجب أن تكون true أو false."})
        tax_rate.is_active = is_active

    if "is_default" in data:
        is_default = data.get("is_default")
        if not isinstance(is_default, bool):
            raise ValidationError({"is_default": "قيمة is_default يجب أن تكون true أو false."})
        tax_rate.is_default = is_default

    if "description" in data:
        tax_rate.description = _clean_text(data.get("description"))

    return tax_rate


def _get_tax_rate_or_none(tax_rate_id: int) -> TaxRate | None:
    return (
        TaxRate.objects.select_related("sales_account", "purchase_account")
        .filter(id=tax_rate_id)
        .first()
    )


# ============================================================
# 🧾 Tax Transaction Serializers
# ============================================================

def _resolve_tax_rate(tax_rate_id: Any) -> TaxRate:
    try:
        parsed_id = int(tax_rate_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({"tax_rate_id": "معرف الضريبة غير صحيح."}) from exc

    tax_rate = TaxRate.objects.filter(id=parsed_id).first()

    if not tax_rate:
        raise ValidationError({"tax_rate_id": "الضريبة المطلوبة غير موجودة."})

    if not tax_rate.is_active:
        raise ValidationError({"tax_rate_id": "لا يمكن استخدام ضريبة غير نشطة."})

    return tax_rate


def _resolve_journal_entry(journal_entry_id: Any) -> JournalEntry | None:
    if journal_entry_id in {None, ""}:
        return None

    try:
        parsed_id = int(journal_entry_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({"journal_entry_id": "معرف القيد غير صحيح."}) from exc

    entry = JournalEntry.objects.filter(id=parsed_id).first()

    if not entry:
        raise ValidationError({"journal_entry_id": "القيد المطلوب غير موجود."})

    return entry


def _resolve_journal_line(journal_line_id: Any) -> JournalEntryLine | None:
    if journal_line_id in {None, ""}:
        return None

    try:
        parsed_id = int(journal_line_id)
    except (TypeError, ValueError) as exc:
        raise ValidationError({"journal_line_id": "معرف سطر القيد غير صحيح."}) from exc

    line = JournalEntryLine.objects.select_related("account").filter(id=parsed_id).first()

    if not line:
        raise ValidationError({"journal_line_id": "سطر القيد المطلوب غير موجود."})

    return line


def _serialize_tax_transaction(tax_transaction: TaxTransaction) -> dict[str, Any]:
    return {
        "id": tax_transaction.id,
        "tax_rate_id": tax_transaction.tax_rate_id,
        "tax_rate": _serialize_tax_rate(tax_transaction.tax_rate),
        "direction": tax_transaction.direction,
        "direction_label": _status_label(tax_transaction.direction, TaxDirection.choices),
        "transaction_date": _iso_date(tax_transaction.transaction_date),
        "taxable_amount": tax_transaction.taxable_amount,
        "tax_amount": tax_transaction.tax_amount,
        "currency": tax_transaction.currency,
        "source_type": tax_transaction.source_type,
        "source_id": tax_transaction.source_id,
        "source_number": tax_transaction.source_number,
        "journal_entry_id": tax_transaction.journal_entry_id,
        "journal_entry": _serialize_journal_entry(tax_transaction.journal_entry),
        "journal_line_id": tax_transaction.journal_line_id,
        "journal_line": _serialize_journal_line(tax_transaction.journal_line),
        "description": tax_transaction.description,
        "metadata": tax_transaction.metadata or {},
        "created_at": _iso_datetime(tax_transaction.created_at),
        "updated_at": _iso_datetime(tax_transaction.updated_at),
    }


def _build_tax_transactions_queryset(request):
    search = (request.GET.get("search") or "").strip()
    direction = (request.GET.get("direction") or "").strip()
    source_type = (request.GET.get("source_type") or "").strip()
    source_id = (request.GET.get("source_id") or "").strip()
    source_number = (request.GET.get("source_number") or "").strip()
    currency = (request.GET.get("currency") or "").strip().upper()

    tax_rate_id = _parse_int(request.GET.get("tax_rate_id"), "tax_rate_id")
    journal_entry_id = _parse_int(request.GET.get("journal_entry_id"), "journal_entry_id")

    date_from = _parse_date(request.GET.get("date_from"), "date_from")
    date_to = _parse_date(request.GET.get("date_to"), "date_to")
    _validate_date_range(date_from, date_to)

    qs = TaxTransaction.objects.select_related(
        "tax_rate",
        "tax_rate__sales_account",
        "tax_rate__purchase_account",
        "journal_entry",
        "journal_line",
        "journal_line__account",
    )

    if search:
        qs = qs.filter(
            Q(description__icontains=search)
            | Q(source_type__icontains=search)
            | Q(source_id__icontains=search)
            | Q(source_number__icontains=search)
            | Q(tax_rate__code__icontains=search)
            | Q(tax_rate__name__icontains=search)
            | Q(journal_entry__entry_number__icontains=search)
        )

    if direction:
        valid_directions = {choice[0] for choice in TaxDirection.choices}
        if direction not in valid_directions:
            raise ValueError("اتجاه الضريبة غير صحيح.")
        qs = qs.filter(direction=direction)

    if source_type:
        qs = qs.filter(source_type__icontains=source_type)

    if source_id:
        qs = qs.filter(source_id=source_id)

    if source_number:
        qs = qs.filter(source_number__icontains=source_number)

    if currency:
        qs = qs.filter(currency__iexact=currency)

    if tax_rate_id:
        qs = qs.filter(tax_rate_id=tax_rate_id)

    if journal_entry_id:
        qs = qs.filter(journal_entry_id=journal_entry_id)

    if date_from:
        qs = qs.filter(transaction_date__gte=date_from)

    if date_to:
        qs = qs.filter(transaction_date__lte=date_to)

    return {
        "qs": qs,
        "filters": {
            "search": search or None,
            "direction": direction or None,
            "source_type": source_type or None,
            "source_id": source_id or None,
            "source_number": source_number or None,
            "currency": currency or None,
            "tax_rate_id": tax_rate_id,
            "journal_entry_id": journal_entry_id,
            "date_from": _iso_date(date_from),
            "date_to": _iso_date(date_to),
        },
    }


def _parse_tax_transaction_ordering(value: str | None, default: str = "-transaction_date") -> str:
    raw = (value or "").strip() or default

    if raw not in ALLOWED_TAX_TRANSACTION_ORDERING:
        allowed = ", ".join(ALLOWED_TAX_TRANSACTION_ORDERING.keys())
        raise ValueError(f"قيمة ordering غير مدعومة. القيم المسموحة: {allowed}")

    return ALLOWED_TAX_TRANSACTION_ORDERING[raw]


def _build_tax_transactions_summary(transactions: list[TaxTransaction]) -> dict[str, Any]:
    total = len(transactions)
    total_taxable = _money(sum((_money(item.taxable_amount) for item in transactions), Decimal("0.00")))
    total_tax = _money(sum((_money(item.tax_amount) for item in transactions), Decimal("0.00")))

    output_tax = _money(
        sum(
            (_money(item.tax_amount) for item in transactions if item.direction == TaxDirection.OUTPUT),
            Decimal("0.00"),
        )
    )
    input_tax = _money(
        sum(
            (_money(item.tax_amount) for item in transactions if item.direction == TaxDirection.INPUT),
            Decimal("0.00"),
        )
    )

    by_direction: dict[str, int] = {}

    for item in transactions:
        by_direction[item.direction] = by_direction.get(item.direction, 0) + 1

    return {
        "total_transactions": total,
        "total_taxable_amount": total_taxable,
        "total_tax_amount": total_tax,
        "output_tax_amount": output_tax,
        "input_tax_amount": input_tax,
        "net_vat_payable": _money(output_tax - input_tax),
        "by_direction": by_direction,
    }


def _apply_payload_to_tax_transaction(
    tax_transaction: TaxTransaction,
    data: dict[str, Any],
) -> TaxTransaction:
    if "tax_rate_id" in data:
        tax_transaction.tax_rate = _resolve_tax_rate(data.get("tax_rate_id"))

    if "direction" in data:
        direction = _clean_text(data.get("direction"))
        valid_directions = {choice[0] for choice in TaxDirection.choices}
        if direction not in valid_directions:
            raise ValidationError({"direction": "اتجاه الضريبة غير صحيح."})
        tax_transaction.direction = direction

    if "transaction_date" in data:
        tax_transaction.transaction_date = _parse_date(data.get("transaction_date"), "transaction_date")

    if "taxable_amount" in data:
        tax_transaction.taxable_amount = _money(data.get("taxable_amount"))

    if "tax_amount" in data:
        tax_transaction.tax_amount = _money(data.get("tax_amount"))

    if "currency" in data:
        tax_transaction.currency = _clean_text(data.get("currency")).upper() or "SAR"

    if "source_type" in data:
        tax_transaction.source_type = _clean_text(data.get("source_type"))

    if "source_id" in data:
        tax_transaction.source_id = _clean_text(data.get("source_id"))

    if "source_number" in data:
        tax_transaction.source_number = _clean_text(data.get("source_number"))

    if "journal_entry_id" in data:
        tax_transaction.journal_entry = _resolve_journal_entry(data.get("journal_entry_id"))

    if "journal_line_id" in data:
        tax_transaction.journal_line = _resolve_journal_line(data.get("journal_line_id"))

    if "description" in data:
        tax_transaction.description = _clean_text(data.get("description"))

    if "metadata" in data:
        metadata = data.get("metadata")
        if metadata in {None, ""}:
            tax_transaction.metadata = {}
        elif not isinstance(metadata, dict):
            raise ValidationError({"metadata": "metadata يجب أن تكون JSON Object."})
        else:
            tax_transaction.metadata = metadata

    return tax_transaction


def _get_tax_transaction_or_none(tax_transaction_id: int) -> TaxTransaction | None:
    return (
        TaxTransaction.objects.select_related(
            "tax_rate",
            "tax_rate__sales_account",
            "tax_rate__purchase_account",
            "journal_entry",
            "journal_line",
            "journal_line__account",
        )
        .filter(id=tax_transaction_id)
        .first()
    )


# ============================================================
# 📤 Excel Builders
# ============================================================

def _build_tax_rates_excel(
    tax_rates: list[TaxRate],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Tax Rates")

    ws.append(["Tax Rates"])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)

    meta_rows = [
        ("Search", filters.get("search")),
        ("Tax Type", filters.get("tax_type")),
        ("Is Active", filters.get("is_active")),
        ("Is Default", filters.get("is_default")),
        ("Total Tax Rates", summary.get("total_tax_rates")),
        ("Active Tax Rates", summary.get("active_tax_rates")),
        ("Default Tax Rates", summary.get("default_tax_rates")),
    ]

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    start_row = current_row + 1

    headers = [
        "ID",
        "Code",
        "Name",
        "Tax Type",
        "Tax Type Label",
        "Rate",
        "Sales Account Code",
        "Sales Account",
        "Purchase Account Code",
        "Purchase Account",
        "Is Active",
        "Is Default",
        "Description",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for tax_rate in tax_rates:
        row = _serialize_tax_rate(tax_rate)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=row.get("code"))
        ws.cell(row=current_row, column=3, value=row.get("name"))
        ws.cell(row=current_row, column=4, value=row.get("tax_type"))
        ws.cell(row=current_row, column=5, value=row.get("tax_type_label"))
        ws.cell(row=current_row, column=6, value=float(row.get("rate", 0) or 0))
        ws.cell(row=current_row, column=7, value=(row.get("sales_account") or {}).get("code") if row.get("sales_account") else "")
        ws.cell(row=current_row, column=8, value=(row.get("sales_account") or {}).get("name") if row.get("sales_account") else "")
        ws.cell(row=current_row, column=9, value=(row.get("purchase_account") or {}).get("code") if row.get("purchase_account") else "")
        ws.cell(row=current_row, column=10, value=(row.get("purchase_account") or {}).get("name") if row.get("purchase_account") else "")
        ws.cell(row=current_row, column=11, value=str(row.get("is_active")))
        ws.cell(row=current_row, column=12, value=str(row.get("is_default")))
        ws.cell(row=current_row, column=13, value=row.get("description"))
        ws.cell(row=current_row, column=14, value=row.get("created_at"))
        ws.cell(row=current_row, column=15, value=row.get("updated_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "tax_rates.xlsx")


def _build_tax_transactions_excel(
    tax_transactions: list[TaxTransaction],
    filters: dict[str, Any],
    summary: dict[str, Any],
) -> HttpResponse:
    workbook = Workbook()
    ws = workbook.active
    ws.title = _safe_sheet_title("Tax Transactions")

    ws.append(["Tax Transactions"])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)

    meta_rows = [
        ("Search", filters.get("search")),
        ("Direction", filters.get("direction")),
        ("Tax Rate ID", filters.get("tax_rate_id")),
        ("Journal Entry ID", filters.get("journal_entry_id")),
        ("Source Type", filters.get("source_type")),
        ("Source ID", filters.get("source_id")),
        ("Source Number", filters.get("source_number")),
        ("Currency", filters.get("currency")),
        ("Date From", filters.get("date_from")),
        ("Date To", filters.get("date_to")),
        ("Total Transactions", summary.get("total_transactions")),
        ("Total Taxable Amount", summary.get("total_taxable_amount")),
        ("Total Tax Amount", summary.get("total_tax_amount")),
        ("Output Tax Amount", summary.get("output_tax_amount")),
        ("Input Tax Amount", summary.get("input_tax_amount")),
        ("Net VAT Payable", summary.get("net_vat_payable")),
    ]

    current_row = 3
    for label, value in meta_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

    start_row = current_row + 1

    headers = [
        "ID",
        "Tax Rate",
        "Direction",
        "Transaction Date",
        "Taxable Amount",
        "Tax Amount",
        "Currency",
        "Source Type",
        "Source ID",
        "Source Number",
        "Journal Entry",
        "Journal Line ID",
        "Description",
        "Created At",
        "Updated At",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        _apply_header_style(cell)

    current_row = start_row + 1

    for tax_transaction in tax_transactions:
        row = _serialize_tax_transaction(tax_transaction)

        ws.cell(row=current_row, column=1, value=row.get("id"))
        ws.cell(row=current_row, column=2, value=(row.get("tax_rate") or {}).get("code") if row.get("tax_rate") else "")
        ws.cell(row=current_row, column=3, value=row.get("direction"))
        ws.cell(row=current_row, column=4, value=row.get("transaction_date"))
        ws.cell(row=current_row, column=5, value=float(row.get("taxable_amount", 0) or 0))
        ws.cell(row=current_row, column=6, value=float(row.get("tax_amount", 0) or 0))
        ws.cell(row=current_row, column=7, value=row.get("currency"))
        ws.cell(row=current_row, column=8, value=row.get("source_type"))
        ws.cell(row=current_row, column=9, value=row.get("source_id"))
        ws.cell(row=current_row, column=10, value=row.get("source_number"))
        ws.cell(row=current_row, column=11, value=(row.get("journal_entry") or {}).get("entry_number") if row.get("journal_entry") else "")
        ws.cell(row=current_row, column=12, value=row.get("journal_line_id"))
        ws.cell(row=current_row, column=13, value=row.get("description"))
        ws.cell(row=current_row, column=14, value=row.get("created_at"))
        ws.cell(row=current_row, column=15, value=row.get("updated_at"))
        current_row += 1

    _auto_fit_columns(ws)
    return _build_excel_response(workbook, "tax_transactions.xlsx")


# ============================================================
# 🧾 Tax Rates API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "POST"])
def accounting_tax_rates_api(request):
    """
    GET:
      قائمة نسب الضرائب.

    POST:
      إنشاء ضريبة.
    """
    if request.method == "POST":
        return _create_tax_rate(request)

    return _list_tax_rates(request)


def _list_tax_rates(request):
    try:
        query_data = _build_tax_rates_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_tax_rate_ordering(request.GET.get("ordering"), default="code")

        if ordering.lstrip("-") == "code":
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by(ordering, "code")

        tax_rates = list(qs)
        summary = _build_tax_rates_summary(tax_rates)
        paginator = Paginator(tax_rates, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        return _success_response(
            {
                "filters": {
                    **query_data["filters"],
                    "ordering": ordering,
                },
                "summary": summary,
                "choices": _choices_payload(),
                "pagination": {
                    "page": page_obj.number,
                    "page_size": page_size,
                    "total_pages": paginator.num_pages,
                    "total_items": paginator.count,
                    "has_next": page_obj.has_next(),
                    "has_previous": page_obj.has_previous(),
                    "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                    "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
                },
                "results": [
                    _serialize_tax_rate(item)
                    for item in page_obj.object_list
                ],
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load tax rates: %s", exc)
        return _error_response("تعذر تحميل نسب الضرائب.", status=500)


@transaction.atomic
def _create_tax_rate(request):
    try:
        data = _json_body(request)

        tax_rate = TaxRate()
        _apply_payload_to_tax_rate(tax_rate, data)
        tax_rate.full_clean()
        tax_rate.save()

        return _success_response(
            {
                "tax_rate": _serialize_tax_rate(tax_rate),
                "message": "تم إنشاء الضريبة بنجاح.",
            },
            status=201,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to create tax rate: %s", exc)
        return _error_response("تعذر إنشاء الضريبة.", status=500)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT"])
def accounting_tax_rate_detail_api(request, tax_rate_id: int):
    """
    GET:
      تفاصيل ضريبة.

    PATCH/PUT:
      تحديث ضريبة.
    """
    if request.method in {"PATCH", "PUT"}:
        return _update_tax_rate(request, tax_rate_id)

    return _detail_tax_rate(request, tax_rate_id)


def _detail_tax_rate(request, tax_rate_id: int):
    try:
        tax_rate = _get_tax_rate_or_none(tax_rate_id)

        if not tax_rate:
            return _error_response("الضريبة المطلوبة غير موجودة.", status=404)

        recent_transactions = list(
            TaxTransaction.objects.select_related("journal_entry")
            .filter(tax_rate_id=tax_rate.id)
            .order_by("-transaction_date", "-id")[:20]
        )

        return _success_response(
            {
                "tax_rate": _serialize_tax_rate(tax_rate),
                "recent_transactions": [
                    {
                        "id": item.id,
                        "direction": item.direction,
                        "transaction_date": _iso_date(item.transaction_date),
                        "taxable_amount": item.taxable_amount,
                        "tax_amount": item.tax_amount,
                        "currency": item.currency,
                        "source_type": item.source_type,
                        "source_id": item.source_id,
                        "source_number": item.source_number,
                        "journal_entry_number": item.journal_entry.entry_number if item.journal_entry else None,
                    }
                    for item in recent_transactions
                ],
            }
        )

    except Exception as exc:
        logger.exception("Failed to load tax rate detail: %s", exc)
        return _error_response("تعذر تحميل تفاصيل الضريبة.", status=500)


@transaction.atomic
def _update_tax_rate(request, tax_rate_id: int):
    try:
        data = _json_body(request)
        tax_rate = _get_tax_rate_or_none(tax_rate_id)

        if not tax_rate:
            return _error_response("الضريبة المطلوبة غير موجودة.", status=404)

        _apply_payload_to_tax_rate(tax_rate, data)
        tax_rate.full_clean()
        tax_rate.save()

        return _success_response(
            {
                "tax_rate": _serialize_tax_rate(tax_rate),
                "message": "تم تحديث الضريبة بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update tax rate: %s", exc)
        return _error_response("تعذر تحديث الضريبة.", status=500)


@csrf_exempt
@require_http_methods(["POST"])
def accounting_tax_rate_status_api(request, tax_rate_id: int):
    """
    تغيير حالة الضريبة.
    body:
    {
      "is_active": true | false
    }
    """
    try:
        data = _json_body(request)

        if "is_active" not in data or not isinstance(data.get("is_active"), bool):
            raise ValueError("قيمة is_active مطلوبة ويجب أن تكون true أو false.")

        tax_rate = _get_tax_rate_or_none(tax_rate_id)

        if not tax_rate:
            return _error_response("الضريبة المطلوبة غير موجودة.", status=404)

        tax_rate.is_active = data["is_active"]
        tax_rate.full_clean()
        tax_rate.save(update_fields=["is_active", "updated_at"])

        return _success_response(
            {
                "tax_rate": _serialize_tax_rate(tax_rate),
                "message": "تم تحديث حالة الضريبة بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update tax rate status: %s", exc)
        return _error_response("تعذر تحديث حالة الضريبة.", status=500)


@csrf_exempt
@require_http_methods(["POST"])
def accounting_tax_rate_set_default_api(request, tax_rate_id: int):
    """
    تعيين ضريبة كافتراضية.
    """
    try:
        tax_rate = _get_tax_rate_or_none(tax_rate_id)

        if not tax_rate:
            return _error_response("الضريبة المطلوبة غير موجودة.", status=404)

        if not tax_rate.is_active:
            return _error_response("لا يمكن تعيين ضريبة غير نشطة كافتراضية.", status=400)

        with transaction.atomic():
            TaxRate.objects.exclude(id=tax_rate.id).update(is_default=False)
            tax_rate.is_default = True
            tax_rate.save(update_fields=["is_default", "updated_at"])

        return _success_response(
            {
                "tax_rate": _serialize_tax_rate(tax_rate),
                "message": "تم تعيين الضريبة الافتراضية بنجاح.",
            }
        )

    except Exception as exc:
        logger.exception("Failed to set default tax rate: %s", exc)
        return _error_response("تعذر تعيين الضريبة الافتراضية.", status=500)


@require_GET
def accounting_tax_rates_excel_api(request):
    """
    تصدير نسب الضرائب إلى Excel.
    """
    try:
        query_data = _build_tax_rates_queryset(request)
        ordering = _parse_tax_rate_ordering(request.GET.get("ordering"), default="code")

        if ordering.lstrip("-") == "code":
            qs = query_data["qs"].order_by(ordering)
        else:
            qs = query_data["qs"].order_by(ordering, "code")

        tax_rates = list(qs)
        summary = _build_tax_rates_summary(tax_rates)

        return _build_tax_rates_excel(
            tax_rates,
            query_data["filters"],
            summary,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export tax rates excel: %s", exc)
        return _error_response("تعذر تصدير نسب الضرائب إلى Excel.", status=500)


# ============================================================
# 🧾 Tax Transactions API
# ============================================================

@csrf_exempt
@require_http_methods(["GET", "POST"])
def accounting_tax_transactions_api(request):
    """
    GET:
      قائمة الحركات الضريبية.

    POST:
      إنشاء حركة ضريبية.
    """
    if request.method == "POST":
        return _create_tax_transaction(request)

    return _list_tax_transactions(request)


def _list_tax_transactions(request):
    try:
        query_data = _build_tax_transactions_queryset(request)
        qs = query_data["qs"]

        page = _parse_positive_int(
            request.GET.get("page"),
            "page",
            default=DEFAULT_PAGE,
            min_value=1,
        )
        page_size = _parse_positive_int(
            request.GET.get("page_size"),
            "page_size",
            default=DEFAULT_PAGE_SIZE,
            min_value=1,
            max_value=MAX_PAGE_SIZE,
        )
        ordering = _parse_tax_transaction_ordering(
            request.GET.get("ordering"),
            default="-transaction_date",
        )

        if ordering.lstrip("-") == "transaction_date":
            qs = qs.order_by(ordering, "-id")
        else:
            qs = qs.order_by(ordering, "-transaction_date", "-id")

        tax_transactions = list(qs)
        summary = _build_tax_transactions_summary(tax_transactions)
        paginator = Paginator(tax_transactions, page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            raise ValueError("رقم الصفحة المطلوب خارج النطاق.")

        return _success_response(
            {
                "filters": {
                    **query_data["filters"],
                    "ordering": ordering,
                },
                "summary": summary,
                "choices": _choices_payload(),
                "pagination": {
                    "page": page_obj.number,
                    "page_size": page_size,
                    "total_pages": paginator.num_pages,
                    "total_items": paginator.count,
                    "has_next": page_obj.has_next(),
                    "has_previous": page_obj.has_previous(),
                    "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
                    "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
                },
                "results": [
                    _serialize_tax_transaction(item)
                    for item in page_obj.object_list
                ],
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to load tax transactions: %s", exc)
        return _error_response("تعذر تحميل الحركات الضريبية.", status=500)


@transaction.atomic
def _create_tax_transaction(request):
    try:
        data = _json_body(request)

        tax_transaction = TaxTransaction()
        _apply_payload_to_tax_transaction(tax_transaction, data)
        tax_transaction.full_clean()
        tax_transaction.save()

        return _success_response(
            {
                "tax_transaction": _serialize_tax_transaction(tax_transaction),
                "message": "تم إنشاء الحركة الضريبية بنجاح.",
            },
            status=201,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to create tax transaction: %s", exc)
        return _error_response("تعذر إنشاء الحركة الضريبية.", status=500)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT"])
def accounting_tax_transaction_detail_api(request, tax_transaction_id: int):
    """
    GET:
      تفاصيل حركة ضريبية.

    PATCH/PUT:
      تحديث حركة ضريبية.
    """
    if request.method in {"PATCH", "PUT"}:
        return _update_tax_transaction(request, tax_transaction_id)

    return _detail_tax_transaction(request, tax_transaction_id)


def _detail_tax_transaction(request, tax_transaction_id: int):
    try:
        tax_transaction = _get_tax_transaction_or_none(tax_transaction_id)

        if not tax_transaction:
            return _error_response("الحركة الضريبية المطلوبة غير موجودة.", status=404)

        return _success_response(
            {
                "tax_transaction": _serialize_tax_transaction(tax_transaction),
            }
        )

    except Exception as exc:
        logger.exception("Failed to load tax transaction detail: %s", exc)
        return _error_response("تعذر تحميل تفاصيل الحركة الضريبية.", status=500)


@transaction.atomic
def _update_tax_transaction(request, tax_transaction_id: int):
    try:
        data = _json_body(request)
        tax_transaction = _get_tax_transaction_or_none(tax_transaction_id)

        if not tax_transaction:
            return _error_response("الحركة الضريبية المطلوبة غير موجودة.", status=404)

        _apply_payload_to_tax_transaction(tax_transaction, data)
        tax_transaction.full_clean()
        tax_transaction.save()

        return _success_response(
            {
                "tax_transaction": _serialize_tax_transaction(tax_transaction),
                "message": "تم تحديث الحركة الضريبية بنجاح.",
            }
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except ValidationError as exc:
        return JsonResponse(
            _validation_error_payload(exc),
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    except Exception as exc:
        logger.exception("Failed to update tax transaction: %s", exc)
        return _error_response("تعذر تحديث الحركة الضريبية.", status=500)


@require_GET
def accounting_tax_transactions_excel_api(request):
    """
    تصدير الحركات الضريبية إلى Excel.
    """
    try:
        query_data = _build_tax_transactions_queryset(request)
        ordering = _parse_tax_transaction_ordering(
            request.GET.get("ordering"),
            default="-transaction_date",
        )

        if ordering.lstrip("-") == "transaction_date":
            qs = query_data["qs"].order_by(ordering, "-id")
        else:
            qs = query_data["qs"].order_by(ordering, "-transaction_date", "-id")

        tax_transactions = list(qs)
        summary = _build_tax_transactions_summary(tax_transactions)

        return _build_tax_transactions_excel(
            tax_transactions,
            query_data["filters"],
            summary,
        )

    except ValueError as exc:
        return _error_response(str(exc), status=400)

    except Exception as exc:
        logger.exception("Failed to export tax transactions excel: %s", exc)
        return _error_response("تعذر تصدير الحركات الضريبية إلى Excel.", status=500)


# ============================================================
# 🧩 Choices
# ============================================================

def _choices_payload() -> dict[str, Any]:
    return {
        "tax_types": [
            {"value": value, "label": label}
            for value, label in TaxType.choices
        ],
        "tax_directions": [
            {"value": value, "label": label}
            for value, label in TaxDirection.choices
        ],
        "tax_rate_orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_TAX_RATE_ORDERING.keys()
        ],
        "tax_transaction_orderings": [
            {"value": value, "label": value}
            for value in ALLOWED_TAX_TRANSACTION_ORDERING.keys()
        ],
    }